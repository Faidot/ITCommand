from collections import defaultdict
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO

from django.db import transaction
from django.db.models import Count, Prefetch, Q
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response

from core.mixins import AuditLogMixin
from core.models import (
    BudgetCategory,
    Department,
    Expense,
    Notification,
    Role,
    SoftwareLicense,
    Subscription,
    SubscriptionAssignment,
    SubscriptionCategoryBudget,
    SubscriptionPayment,
    SubscriptionRenewal,
    SubscriptionSettings,
    User,
    VaultCredential,
    Vendor,
    VendorContract,
)
from core.permissions import HasModulePermission, has_role_permission
from core.serializers import (
    SubscriptionAssignmentSerializer,
    SubscriptionCategoryBudgetSerializer,
    SubscriptionDetailSerializer,
    SubscriptionRenewalSerializer,
    SubscriptionSerializer,
    SubscriptionSettingsSerializer,
)
# Reused rather than duplicated: the helper is generic over MONTHLY/YEARLY and
# already handles the Feb-29 edge case.
from core.views.licenses import next_expiry_for_cycle
from core.fx import convert_many, rate_as_of
from core.lov import get_values
from core.subscription_alerts import (
    refresh_subscription_alerts,
    retire_subscription_notifications,
)


TWOPLACES = Decimal("0.01")


def _money(value):
    return str(Decimal(value or 0).quantize(TWOPLACES, rounding=ROUND_HALF_UP))


def _safe_date(value):
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except (TypeError, ValueError):
        return None


def _excel_text(value):
    """Keep user-entered text from being interpreted as a spreadsheet formula."""
    value = str(value or "")
    if value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def apply_subscription_renewal(
    subscription, *, new_expiry_date, cost=Decimal("0.00"), seats_added=0, notes="", renewer=None
):
    """Record a renewal and advance the subscription's expiry.

    Returns the SubscriptionRenewal. Callers are responsible for refreshing
    alerts afterwards — see run_subscription_auto_renewals.
    """
    renewal = SubscriptionRenewal.objects.create(
        subscription=subscription,
        previous_expiry=subscription.expiry_date,
        new_expiry=new_expiry_date,
        cost=cost,
        seats_added=seats_added,
        notes=notes,
        renewed_by=renewer,
    )
    subscription.expiry_date = new_expiry_date
    if seats_added and subscription.seats_total is not None:
        subscription.seats_total = subscription.seats_total + seats_added
    subscription.save(update_fields=["expiry_date", "seats_total", "updated_at"])
    return renewal


def run_subscription_auto_renewals(actor=None):
    """Advance expiry on every active subscription with auto_renew=True whose
    expiry has passed. Repeats per subscription until the expiry is in the
    future, so one that is months overdue catches up in a single run.

    Returns a list of {id, name, previous_expiry, new_expiry, cycles_advanced}.
    """
    today = timezone.localdate()
    results = []
    qs = Subscription.objects.filter(
        status="ACTIVE",
        auto_renew=True,
        expiry_date__isnull=False,
        expiry_date__lt=today,
        billing_cycle__in=["MONTHLY", "YEARLY"],
    )
    for subscription in qs:
        cycles = 0
        previous = subscription.expiry_date
        with transaction.atomic():
            while (
                subscription.expiry_date
                and subscription.expiry_date < today
                and cycles < 60
            ):
                next_expiry = next_expiry_for_cycle(
                    subscription.expiry_date, subscription.billing_cycle
                )
                if not next_expiry:
                    break
                apply_subscription_renewal(
                    subscription,
                    new_expiry_date=next_expiry,
                    notes="Auto-renewed",
                    renewer=actor,
                )
                cycles += 1
        if cycles:
            # The expiry has moved, so any EXPIRY/RENEWAL notification raised
            # against the old date is now wrong. Subscriptions have no
            # is_resolved flag to flip (licenses do), so reconcile instead.
            refresh_subscription_alerts(subscription.pk)
            results.append(
                {
                    "id": subscription.pk,
                    "name": subscription.name,
                    "previous_expiry": str(previous),
                    "new_expiry": str(subscription.expiry_date),
                    "cycles_advanced": cycles,
                }
            )
    return results


# Seat notifications deliberately use their own type. The alert reconciler in
# core.subscription_alerts deletes *every* unread notification typed
# "SUBSCRIPTION" when notifications are switched off, which would take seat
# messages with it.
SEAT_NOTIFICATION_TYPE = "SUBSCRIPTION_SEAT"


def _notify_seat_change(user, message, subscription_id):
    if user:
        Notification.objects.create(
            user=user,
            message=message,
            notification_type=SEAT_NOTIFICATION_TYPE,
            link=f"/subscriptions/{subscription_id}",
        )


def _new_spend_bucket():
    return {
        "monthly_billed": Decimal("0"),
        "yearly_billed": Decimal("0"),
        "count": 0,
    }


def _add_subscription_cost(bucket, subscription):
    key = "monthly_billed" if subscription.billing_cycle == "MONTHLY" else "yearly_billed"
    bucket[key] += subscription.cost
    bucket["count"] += 1


def _normalized_spend(bucket):
    annual = bucket["monthly_billed"] * Decimal("12") + bucket["yearly_billed"]
    return annual / Decimal("12"), annual


class SubscriptionPagination(PageNumberPagination):
    page_size = 25
    page_size_query_param = "page_size"
    max_page_size = 100


class SubscriptionViewSet(AuditLogMixin, viewsets.ModelViewSet):
    serializer_class = SubscriptionSerializer
    permission_classes = [HasModulePermission]
    rbac_module = "subscriptions"
    pagination_class = SubscriptionPagination
    rbac_action_permissions = {
        "subscription_settings": {"PATCH": "edit", "PUT": "edit"},
    }

    def get_serializer_class(self):
        # Only retrieve pays for the nested rosters/history; the list stays lean.
        if self.action == "retrieve":
            return SubscriptionDetailSerializer
        return SubscriptionSerializer

    @staticmethod
    def _refresh_alerts_after_commit(subscription_id):
        transaction.on_commit(
            lambda pk=subscription_id: refresh_subscription_alerts(pk),
            robust=True,
        )

    def perform_create(self, serializer):
        super().perform_create(serializer)
        self._refresh_alerts_after_commit(serializer.instance.pk)

    def perform_update(self, serializer):
        super().perform_update(serializer)
        self._refresh_alerts_after_commit(serializer.instance.pk)

    def perform_destroy(self, instance):
        subscription_id = instance.pk
        super().perform_destroy(instance)
        transaction.on_commit(
            lambda pk=subscription_id: retire_subscription_notifications(pk),
            robust=True,
        )

    def perform_content_negotiation(self, request, force=False):
        # DRF normally reserves ?format= for renderer selection and returns a
        # 404 for formats without a registered renderer. This endpoint uses
        # the documented query parameter to choose a binary report instead.
        if getattr(self, "action", None) == "export":
            renderer = self.get_renderers()[0]
            return renderer, renderer.media_type
        return super().perform_content_negotiation(request, force=force)

    def get_queryset(self):
        qs = Subscription.objects.select_related(
            "department",
            "owner",
            "admin",
            "created_by",
            "vendor",
            "vendor_contract",
            "budget_category",
            "vault_credential",
            "linked_license__product",
        ).annotate(
            # Without this, serializing seats_used costs one COUNT per row.
            _seats_used=Count(
                "assignments", filter=Q(assignments__is_active=True), distinct=True
            )
        )
        if self.action == "retrieve":
            qs = qs.prefetch_related(
                Prefetch(
                    "assignments",
                    queryset=SubscriptionAssignment.objects.select_related(
                        "user", "user__department", "assigned_by"
                    ).order_by("-is_active", "-assigned_date"),
                ),
                Prefetch(
                    "renewals",
                    queryset=SubscriptionRenewal.objects.select_related("renewed_by"),
                ),
                Prefetch(
                    "expenses",
                    queryset=Expense.objects.order_by("-expense_date"),
                ),
                Prefetch(
                    "payments",
                    queryset=SubscriptionPayment.objects.select_related("card"),
                ),
            )
        params = self.request.query_params

        search = (params.get("search") or "").strip()
        if search:
            qs = qs.filter(
                Q(name__icontains=search)
                | Q(platform__icontains=search)
                | Q(plan_type__icontains=search)
                | Q(purpose__icontains=search)
                | Q(team__icontains=search)
                | Q(notes__icontains=search)
                | Q(owner__full_name__icontains=search)
                | Q(owner__email__icontains=search)
                | Q(admin__full_name__icontains=search)
                | Q(admin__email__icontains=search)
                | Q(vendor__name__icontains=search)
            )

        for field in ("category", "billing_cycle", "currency"):
            value = params.get(field)
            if value:
                qs = qs.filter(**{field: value.upper()})

        for field in ("department", "owner", "admin", "vendor", "budget_category"):
            value = params.get(field)
            if value and value.isdigit():
                qs = qs.filter(**{f"{field}_id": int(value)})

        auto_renew = params.get("auto_renew")
        if auto_renew and auto_renew.lower() in {"true", "false"}:
            qs = qs.filter(auto_renew=auto_renew.lower() == "true")

        today = timezone.localdate()
        effective_status = (params.get("status") or "").upper()
        if effective_status == "ACTIVE":
            qs = qs.filter(status="ACTIVE", start_date__lte=today, expiry_date__gte=today)
        elif effective_status == "EXPIRED":
            qs = qs.filter(status="ACTIVE", expiry_date__lt=today)
        elif effective_status == "SCHEDULED":
            qs = qs.filter(status="ACTIVE", start_date__gt=today)
        elif effective_status in {"PAUSED", "CANCELLED"}:
            qs = qs.filter(status=effective_status)

        expiry_from = _safe_date(params.get("expiry_from"))
        expiry_to = _safe_date(params.get("expiry_to"))
        if expiry_from:
            qs = qs.filter(expiry_date__gte=expiry_from)
        if expiry_to:
            qs = qs.filter(expiry_date__lte=expiry_to)

        expiring_within = params.get("expiring_within")
        if expiring_within:
            try:
                days = max(0, min(int(expiring_within), 3650))
                qs = qs.filter(expiry_date__gte=today, expiry_date__lte=today + timedelta(days=days))
            except ValueError:
                pass

        ordering = params.get("ordering", "expiry_date")
        allowed_ordering = {
            "name", "-name", "platform", "-platform", "cost", "-cost",
            "start_date", "-start_date", "expiry_date", "-expiry_date",
            "created_at", "-created_at", "updated_at", "-updated_at",
        }
        return qs.order_by(ordering if ordering in allowed_ordering else "expiry_date", "name")

    @action(detail=False, methods=["get"])
    def dashboard(self, request):
        subscriptions = list(self.get_queryset())
        settings = SubscriptionSettings.get_solo()
        today = timezone.localdate()

        try:
            days = max(1, min(int(request.query_params.get("days", 60)), 3650))
        except ValueError:
            days = 60

        status_counts = {
            "ACTIVE": 0,
            "EXPIRED": 0,
            "SCHEDULED": 0,
            "PAUSED": 0,
            "CANCELLED": 0,
        }
        spend_by_currency = defaultdict(_new_spend_bucket)
        category_spend = defaultdict(_new_spend_bucket)
        upcoming = []

        for subscription in subscriptions:
            effective = subscription.effective_status
            status_counts[effective] += 1
            if effective == "ACTIVE":
                currency_totals = spend_by_currency[subscription.currency]
                _add_subscription_cost(currency_totals, subscription)

                category_totals = category_spend[(subscription.category, subscription.currency)]
                _add_subscription_cost(category_totals, subscription)

                if today <= subscription.expiry_date <= today + timedelta(days=days):
                    upcoming.append(subscription)

        by_currency = []
        for currency, values in sorted(spend_by_currency.items()):
            monthly_spend, yearly_spend = _normalized_spend(values)
            by_currency.append({
                "currency": currency,
                "monthly_spend": _money(monthly_spend),
                "yearly_spend": _money(yearly_spend),
                "count": values["count"],
            })
        category_labels = dict(Subscription.CATEGORY_CHOICES)
        by_category = []
        for (category, currency), values in sorted(category_spend.items()):
            monthly_spend, yearly_spend = _normalized_spend(values)
            by_category.append({
                "category": category,
                "category_label": category_labels.get(category, category.title()),
                "currency": currency,
                "monthly_spend": _money(monthly_spend),
                "yearly_spend": _money(yearly_spend),
                "count": values["count"],
            })

        # Converted view: one headline number across every currency, with the
        # per-currency breakdown kept above and anything unconvertible listed
        # explicitly rather than dropped or counted as 1:1.
        monthly_pairs, yearly_pairs = [], []
        for currency, values in spend_by_currency.items():
            monthly, yearly = _normalized_spend(values)
            monthly_pairs.append((monthly, currency))
            yearly_pairs.append((yearly, currency))
        converted_monthly, report_currency, unconvertible = convert_many(monthly_pairs)
        converted_yearly, _, _ = convert_many(yearly_pairs)
        converted = {
            "currency": report_currency,
            "monthly_spend": _money(converted_monthly),
            "yearly_spend": _money(converted_yearly),
            "rates_as_of": rate_as_of(report_currency),
            "unconvertible": unconvertible,
            "is_complete": not unconvertible,
        }

        # Budget spend = every subscription converted into the budget currency,
        # not just the ones already priced in it. A USD subscription is converted
        # to PKR (or vice-versa) and added in, so the budget reflects total spend.
        budget_currency = settings.budget_currency
        primary_monthly, _, budget_unconvertible = convert_many(monthly_pairs, to_currency=budget_currency)
        primary_yearly, _, _ = convert_many(yearly_pairs, to_currency=budget_currency)
        monthly_threshold = settings.monthly_budget_threshold
        yearly_threshold = settings.yearly_budget_threshold

        def usage_percent(spend, threshold):
            if threshold is None or threshold <= 0:
                return None
            return float((spend / threshold * Decimal("100")).quantize(TWOPLACES))

        # Per-category budgets: convert each category's spend (across currencies)
        # into the budget currency and compare against its allocation.
        cat_pairs = defaultdict(lambda: {"monthly": [], "yearly": []})
        for (category, currency), values in category_spend.items():
            m, y = _normalized_spend(values)
            cat_pairs[category]["monthly"].append((m, currency))
            cat_pairs[category]["yearly"].append((y, currency))

        category_budget_rows = []
        allocations = {b.category: b for b in SubscriptionCategoryBudget.objects.all()}
        # Categories come from the admin-managed LOV so ones added in Settings
        # are budgetable; fall back to a stored code's own label if it's missing.
        lov_categories = [(str(code).upper(), label) for code, label in get_values("subscription_category")]
        lov_seen = {code for code, _ in lov_categories}
        for code in allocations:
            if code not in lov_seen:
                lov_categories.append((code, code.title()))
        for code, label in lov_categories:
            allocation = allocations.get(code)
            if not allocation or (allocation.monthly_threshold is None and allocation.yearly_threshold is None):
                continue
            m_spend, _, _ = convert_many(cat_pairs[code]["monthly"], to_currency=budget_currency)
            y_spend, _, _ = convert_many(cat_pairs[code]["yearly"], to_currency=budget_currency)
            category_budget_rows.append({
                "category": code,
                "category_label": label,
                "currency": budget_currency,
                "monthly_spend": _money(m_spend),
                "yearly_spend": _money(y_spend),
                "monthly_threshold": _money(allocation.monthly_threshold) if allocation.monthly_threshold is not None else None,
                "yearly_threshold": _money(allocation.yearly_threshold) if allocation.yearly_threshold is not None else None,
                "monthly_usage_percent": usage_percent(m_spend, allocation.monthly_threshold),
                "yearly_usage_percent": usage_percent(y_spend, allocation.yearly_threshold),
                "monthly_exceeded": bool(allocation.monthly_threshold is not None and m_spend >= allocation.monthly_threshold),
                "yearly_exceeded": bool(allocation.yearly_threshold is not None and y_spend >= allocation.yearly_threshold),
            })

        return Response(
            {
                "default_currency": settings.budget_currency,
                "category_budgets": category_budget_rows,
                "converted": converted,
                "monthly_spend": _money(primary_monthly),
                "yearly_spend": _money(primary_yearly),
                "spend_by_currency": by_currency,
                "status_counts": status_counts,
                "active_count": status_counts["ACTIVE"],
                "expired_count": status_counts["EXPIRED"],
                "upcoming_count": len(upcoming),
                "upcoming_renewals": SubscriptionSerializer(
                    sorted(upcoming, key=lambda item: item.expiry_date),
                    many=True,
                    context={"request": request},
                ).data,
                "spend_by_category": by_category,
                "budget": {
                    "currency": settings.budget_currency,
                    "monthly_threshold": (
                        _money(monthly_threshold) if monthly_threshold is not None else None
                    ),
                    "yearly_threshold": (
                        _money(yearly_threshold) if yearly_threshold is not None else None
                    ),
                    "monthly_spend": _money(primary_monthly),
                    "yearly_spend": _money(primary_yearly),
                    "monthly_usage_percent": usage_percent(
                        primary_monthly, monthly_threshold
                    ),
                    "yearly_usage_percent": usage_percent(
                        primary_yearly, yearly_threshold
                    ),
                    "monthly_exceeded": bool(
                        monthly_threshold is not None
                        and monthly_threshold > 0
                        and primary_monthly >= monthly_threshold
                    ),
                    "yearly_exceeded": bool(
                        yearly_threshold is not None
                        and yearly_threshold > 0
                        and primary_yearly >= yearly_threshold
                    ),
                    # Spend now aggregates every currency converted into the
                    # budget currency; anything with no rate is listed here.
                    "converted_from_all_currencies": True,
                    "rates_as_of": rate_as_of(budget_currency),
                    "unconvertible": budget_unconvertible,
                },
            }
        )

    # ─── Seat assignments ──────────────────────────────────────────

    @action(detail=True, methods=["post"], url_path="assign")
    @transaction.atomic
    def assign_seat(self, request, pk=None):
        # Lock the row so two concurrent assigns cannot both read the same
        # seats_available and overfill the subscription.
        subscription = Subscription.objects.select_for_update().get(
            pk=self.get_object().pk
        )
        user_id = request.data.get("user_id")
        if not user_id:
            return Response(
                {"detail": "user_id is required."}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            target_user = User.objects.get(id=user_id)
        except (User.DoesNotExist, ValueError, TypeError):
            return Response(
                {"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND
            )

        available = subscription.seats_available
        if available is not None and available <= 0:
            return Response(
                {"detail": "No seats available for this subscription."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if SubscriptionAssignment.objects.filter(
            subscription=subscription, user=target_user, is_active=True
        ).exists():
            return Response(
                {
                    "detail": (
                        f"{target_user.full_name} already has an active seat on "
                        f"this subscription."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        assignment = SubscriptionAssignment.objects.create(
            subscription=subscription,
            user=target_user,
            assigned_by=request.user,
            notes=request.data.get("notes", ""),
        )
        _notify_seat_change(
            target_user,
            f"You have been given a seat on {subscription.name}.",
            subscription.pk,
        )
        self.log_action(
            "ASSIGN_SUBSCRIPTION_SEAT",
            subscription,
            {"action": "assign_seat", "user_id": target_user.pk},
        )
        return Response(
            SubscriptionAssignmentSerializer(assignment).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=["post"], url_path=r"revoke/(?P<user_id>\d+)")
    def revoke_seat(self, request, pk=None, user_id=None):
        subscription = self.get_object()
        try:
            target_user = User.objects.get(id=user_id)
        except (User.DoesNotExist, ValueError, TypeError):
            return Response(
                {"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND
            )

        assignment = SubscriptionAssignment.objects.filter(
            subscription=subscription, user=target_user, is_active=True
        ).first()
        if not assignment:
            return Response(
                {"detail": "No active seat found for this user."},
                status=status.HTTP_404_NOT_FOUND,
            )

        assignment.is_active = False
        assignment.revoked_date = timezone.now()
        assignment.save(update_fields=["is_active", "revoked_date"])

        _notify_seat_change(
            target_user,
            f"Your seat on {subscription.name} has been revoked.",
            subscription.pk,
        )
        self.log_action(
            "REVOKE_SUBSCRIPTION_SEAT",
            subscription,
            {"action": "revoke_seat", "user_id": target_user.pk},
        )
        return Response({"detail": "Seat revoked successfully."})

    @action(detail=True, methods=["get"], url_path="assignments")
    def list_assignments(self, request, pk=None):
        subscription = self.get_object()
        assignments = (
            subscription.assignments.select_related("user", "user__department", "assigned_by")
            .all()
            .order_by("-is_active", "-assigned_date")
        )
        return Response(SubscriptionAssignmentSerializer(assignments, many=True).data)

    # ─── Renewals ──────────────────────────────────────────────────

    @action(detail=True, methods=["get"], url_path="renewals")
    def list_renewals(self, request, pk=None):
        subscription = self.get_object()
        return Response(
            SubscriptionRenewalSerializer(
                subscription.renewals.select_related("renewed_by").all(), many=True
            ).data
        )

    @action(detail=True, methods=["get"], url_path="suggest_next_expiry")
    def suggest_next_expiry(self, request, pk=None):
        subscription = self.get_object()
        next_expiry = next_expiry_for_cycle(
            subscription.expiry_date, subscription.billing_cycle
        )
        return Response(
            {
                "current_expiry": str(subscription.expiry_date)
                if subscription.expiry_date
                else None,
                "billing_cycle": subscription.billing_cycle,
                "suggested_expiry": str(next_expiry) if next_expiry else None,
            }
        )

    @action(detail=True, methods=["post"], url_path="renew")
    def renew(self, request, pk=None):
        """Extend a subscription's expiry and capture a renewal record.

        Body:
            new_expiry (date, required)
            cost (decimal, optional, default 0)
            seats_added (int, optional, default 0)
            notes (str, optional)
        """
        subscription = self.get_object()

        new_expiry = request.data.get("new_expiry")
        if not new_expiry:
            return Response(
                {"detail": "new_expiry is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        new_expiry_date = (
            _safe_date(new_expiry) if isinstance(new_expiry, str) else new_expiry
        )
        if not new_expiry_date:
            return Response(
                {"detail": "new_expiry must be YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if subscription.expiry_date and new_expiry_date <= subscription.expiry_date:
            return Response(
                {
                    "detail": (
                        f"new_expiry ({new_expiry_date}) must be after the current "
                        f"expiry ({subscription.expiry_date})."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            cost = Decimal(str(request.data.get("cost") or "0"))
            seats_added = int(request.data.get("seats_added") or 0)
        except (TypeError, ValueError, InvalidOperation):
            return Response(
                {"detail": "cost / seats_added must be numeric."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if cost < 0:
            return Response(
                {"detail": "cost cannot be negative."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if seats_added < 0:
            return Response(
                {"detail": "seats_added cannot be negative."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        previous_expiry = subscription.expiry_date
        with transaction.atomic():
            renewal = apply_subscription_renewal(
                subscription,
                new_expiry_date=new_expiry_date,
                cost=cost,
                seats_added=seats_added,
                notes=request.data.get("notes", ""),
                renewer=request.user,
            )
            self.log_action(
                "RENEW",
                subscription,
                changes={
                    "previous_expiry": str(previous_expiry) if previous_expiry else None,
                    "new_expiry": str(new_expiry_date),
                    "cost": str(cost),
                    "seats_added": seats_added,
                    "renewal_id": renewal.pk,
                },
            )
        # Reminders were raised against the old expiry; reconcile them.
        self._refresh_alerts_after_commit(subscription.pk)

        return Response(
            SubscriptionRenewalSerializer(renewal).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=False, methods=["post"], url_path="process_auto_renewals")
    def process_auto_renewals(self, request):
        results = run_subscription_auto_renewals(actor=request.user)
        for result in results:
            subscription = Subscription.objects.filter(pk=result["id"]).first()
            if subscription:
                self.log_action(
                    "AUTO_RENEW",
                    subscription,
                    changes={
                        "previous_expiry": result["previous_expiry"],
                        "new_expiry": result["new_expiry"],
                        "cycles_advanced": result["cycles_advanced"],
                    },
                )
        return Response({"renewed": results, "count": len(results)})

    # ─── Bulk actions ──────────────────────────────────────────────

    BULK_STATUS_ACTIONS = {
        "pause": "PAUSED",
        "resume": "ACTIVE",
        "cancel": "CANCELLED",
    }

    @action(detail=False, methods=["post"], url_path="bulk_action")
    def bulk_action(self, request):
        """Apply one change to many subscriptions.

        Body: { ids: [int, ...], action: str, value?: any }

        Note: unlike other bulk endpoints in this codebase, this deliberately
        does NOT use queryset.update(). Subscriptions wire alert reconciliation
        into perform_update/perform_destroy, and a bare .update() would skip it,
        leaving live reminders pointing at cancelled or deleted rows.
        """
        ids = request.data.get("ids") or []
        operation = request.data.get("action")
        value = request.data.get("value")

        if not isinstance(ids, list) or not ids:
            return Response(
                {"detail": "ids[] is required."}, status=status.HTTP_400_BAD_REQUEST
            )
        if not all(isinstance(pk, int) and not isinstance(pk, bool) for pk in ids):
            return Response(
                {"detail": "ids[] must contain integers."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        subscriptions = list(Subscription.objects.filter(pk__in=ids))
        if not subscriptions:
            return Response(
                {"detail": "No matching subscriptions."},
                status=status.HTTP_404_NOT_FOUND,
            )

        if operation == "delete":
            return self._bulk_delete(subscriptions)

        update_fields = []
        changes = {}

        if operation in self.BULK_STATUS_ACTIONS:
            new_status = self.BULK_STATUS_ACTIONS[operation]
            update_fields = ["status"]
            changes = {"status": new_status}
            for subscription in subscriptions:
                subscription.status = new_status
        elif operation in {"auto_renew_on", "auto_renew_off"}:
            enabled = operation == "auto_renew_on"
            update_fields = ["auto_renew"]
            changes = {"auto_renew": enabled}
            for subscription in subscriptions:
                subscription.auto_renew = enabled
        elif operation in {
            "set_owner",
            "set_admin",
            "set_department",
            "set_vendor",
            "set_budget_category",
        }:
            field, model = {
                "set_owner": ("owner", User),
                "set_admin": ("admin", User),
                "set_department": ("department", Department),
                "set_vendor": ("vendor", Vendor),
                "set_budget_category": ("budget_category", BudgetCategory),
            }[operation]
            related = None
            if value not in (None, "", "none"):
                related = model.objects.filter(pk=value).first()
                if not related:
                    return Response(
                        {"detail": f"No matching {field.replace('_', ' ')}."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
            update_fields = [field]
            changes = {field: related.pk if related else None}
            for subscription in subscriptions:
                setattr(subscription, field, related)
        elif operation == "set_category":
            valid = {choice for choice, _ in Subscription.CATEGORY_CHOICES}
            if value not in valid:
                return Response(
                    {"detail": "Invalid category."}, status=status.HTTP_400_BAD_REQUEST
                )
            update_fields = ["category"]
            changes = {"category": value}
            for subscription in subscriptions:
                subscription.category = value
        else:
            return Response(
                {"detail": "Unknown action."}, status=status.HTTP_400_BAD_REQUEST
            )

        with transaction.atomic():
            for subscription in subscriptions:
                subscription.save(update_fields=update_fields + ["updated_at"])
                self.log_action("UPDATE", subscription, changes)

        # Status and auto-renew changes move which reminders are due, and a
        # vendor/owner change moves who receives them.
        for subscription in subscriptions:
            self._refresh_alerts_after_commit(subscription.pk)

        return Response(
            {"detail": "OK", "affected": len(subscriptions), "action": operation}
        )

    def _bulk_delete(self, subscriptions):
        deleted, blocked = [], []
        for subscription in subscriptions:
            reason = self._blocked_reason(subscription)
            if reason:
                blocked.append(
                    {"id": subscription.pk, "name": subscription.name, "reason": reason}
                )
                continue
            subscription_id = subscription.pk
            with transaction.atomic():
                self.log_action("DELETE", subscription, {"bulk": True})
                subscription.delete()
            deleted.append(subscription_id)
            transaction.on_commit(
                lambda pk=subscription_id: retire_subscription_notifications(pk),
                robust=True,
            )
        return Response(
            {
                "deleted_count": len(deleted),
                "blocked_count": len(blocked),
                "deleted": deleted,
                "blocked": blocked,
            }
        )

    # ─── Delete protection ─────────────────────────────────────────

    @staticmethod
    def _blocked_reason(subscription):
        active = subscription.assignments.filter(is_active=True).count()
        if active:
            return f"{active} active seat assignment(s)"
        return None

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        reason = self._blocked_reason(instance)
        if reason:
            return Response(
                {
                    "detail": (
                        f"Cannot delete {instance.name}: {reason}. Revoke seats first."
                    )
                },
                status=status.HTTP_409_CONFLICT,
            )
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=["get"])
    def options(self, request):
        """Subscription-scoped selector data, without requiring Users RBAC."""
        can_select_users = has_role_permission(
            request.user, "subscriptions", "add"
        ) or has_role_permission(request.user, "subscriptions", "edit")
        users = []
        if can_select_users:
            users = list(
                User.objects.filter(is_active=True)
                .order_by("full_name", "email")
                .values("id", "full_name", "email", "role")
            )
            roles = {
                role.slug: role
                for role in Role.objects.filter(
                    slug__in={user["role"] for user in users if user["role"]}
                )
            }
            for user in users:
                role_slug = user.pop("role", None)
                role = roles.get(role_slug)
                user["can_receive_subscription_alerts"] = bool(
                    role_slug == "SUPERADMIN"
                    or (role and role.can("subscriptions", "view"))
                )
        departments = Department.objects.order_by("name").values("id", "name")

        # Cross-module selector data. Vendors/budgets follow the subscriptions
        # write permission; vault and licenses are gated on their own modules so
        # a subscriptions editor cannot enumerate either one.
        vendors = []
        contracts = []
        budget_categories = []
        if can_select_users:
            vendors = list(Vendor.objects.order_by("name").values("id", "name"))
            contracts = list(
                VendorContract.objects.order_by("title").values(
                    "id", "title", "contract_number", "vendor"
                )
            )
            budget_categories = list(
                BudgetCategory.objects.order_by("name").values("id", "name")
            )

        vault_credentials = []
        if has_role_permission(request.user, "vault", "view"):
            vault_credentials = list(
                VaultCredential.objects.order_by("title").values("id", "title")
            )

        licenses = []
        if has_role_permission(request.user, "licenses", "view"):
            licenses = [
                {"id": pk, "name": name}
                for pk, name in SoftwareLicense.objects.order_by(
                    "product__name"
                ).values_list("id", "product__name")
            ]

        return Response(
            {
                "users": users,
                "departments": list(departments),
                # Admin-managed currency list, so a currency added in Django
                # admin is immediately selectable here.
                "currencies": [
                    {"value": code, "label": label}
                    for code, label in get_values("currency")
                ],
                "vendors": vendors,
                "contracts": contracts,
                "budget_categories": budget_categories,
                "vault_credentials": vault_credentials,
                "licenses": licenses,
                "categories": [
                    {"value": value, "label": label}
                    for value, label in Subscription.CATEGORY_CHOICES
                ],
                "billing_cycles": [
                    {"value": value, "label": label}
                    for value, label in Subscription.BILLING_CYCLE_CHOICES
                ],
                "statuses": [
                    {"value": value, "label": label}
                    for value, label in Subscription.STATUS_CHOICES
                ],
            }
        )

    @action(
        detail=False,
        methods=["get", "patch", "put"],
        url_path="settings",
        url_name="settings",
    )
    def subscription_settings(self, request):
        subscription_settings = SubscriptionSettings.get_solo()
        if request.method == "GET":
            return Response(SubscriptionSettingsSerializer(subscription_settings).data)

        serializer = SubscriptionSettingsSerializer(
            subscription_settings,
            data=request.data,
            partial=request.method == "PATCH",
        )
        serializer.is_valid(raise_exception=True)
        serializer.save(updated_by=request.user)
        self.log_action("UPDATE", subscription_settings, serializer.data)
        return Response(serializer.data)

    @staticmethod
    def _category_choices():
        """Categories offered for budgets — sourced from the admin-managed
        `subscription_category` list of values, so ones added in Settings appear
        here too."""
        return [(str(code).upper(), label) for code, label in get_values("subscription_category")]

    def _category_budgets_payload(self):
        existing = {b.category: b for b in SubscriptionCategoryBudget.objects.all()}
        rows = []
        for code, label in self._category_choices():
            b = existing.get(code)
            rows.append({
                "category": code, "category_label": label,
                "monthly_threshold": _money(b.monthly_threshold) if b and b.monthly_threshold is not None else None,
                "yearly_threshold": _money(b.yearly_threshold) if b and b.yearly_threshold is not None else None,
            })
        return {"budgets": rows}

    @action(detail=False, methods=["get", "put"], url_path="category-budgets", url_name="category-budgets")
    def category_budgets(self, request):
        """List or replace per-category subscription spend limits.

        GET  -> every category with its (possibly null) monthly/yearly limit.
        PUT  -> {"budgets": [{category, monthly_threshold, yearly_threshold}]}
                Upserts the rows given; a null/blank threshold clears that limit.
        """
        if request.method == "GET":
            return Response(self._category_budgets_payload())

        payload = request.data.get("budgets", [])
        if not isinstance(payload, list):
            return Response({"detail": "budgets must be a list."}, status=status.HTTP_400_BAD_REQUEST)

        valid = {code for code, _ in self._category_choices()}

        def parse(v):
            if v in (None, ""):
                return None
            try:
                d = Decimal(str(v))
            except (InvalidOperation, TypeError):
                raise ValueError("Thresholds must be numbers.")
            if d <= 0:
                raise ValueError("Thresholds must be greater than zero, or blank to disable.")
            return d.quantize(TWOPLACES)

        try:
            for item in payload:
                category = (item.get("category") or "").strip().upper()
                if category not in valid:
                    return Response({"detail": f"Unknown category '{category}'."}, status=status.HTTP_400_BAD_REQUEST)
                monthly = parse(item.get("monthly_threshold"))
                yearly = parse(item.get("yearly_threshold"))
                if monthly is None and yearly is None:
                    SubscriptionCategoryBudget.objects.filter(category=category).delete()
                else:
                    SubscriptionCategoryBudget.objects.update_or_create(
                        category=category,
                        defaults={"monthly_threshold": monthly, "yearly_threshold": yearly},
                    )
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(self._category_budgets_payload())

    @action(detail=False, methods=["get"])
    def export(self, request):
        export_format = (request.query_params.get("format") or "xlsx").lower()
        subscriptions = list(self.get_queryset())
        if export_format == "xlsx":
            return self._export_xlsx(subscriptions)
        if export_format == "pdf":
            return self._export_pdf(subscriptions)
        return Response(
            {"format": "Use 'xlsx' or 'pdf'."}, status=status.HTTP_400_BAD_REQUEST
        )

    @staticmethod
    def _export_xlsx(subscriptions):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Subscriptions"
        headers = [
            "Name", "Platform", "Plan", "Category", "Status", "Cost",
            "Currency", "Billing cycle", "Monthly cost", "Annual cost",
            "Start date", "Expiry date", "Days remaining", "Auto renew",
            "Department", "Team", "Owner", "Administrator", "Purpose", "Notes",
        ]
        sheet.append(headers)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for subscription in subscriptions:
            sheet.append(
                [
                    _excel_text(subscription.name),
                    _excel_text(subscription.platform),
                    _excel_text(subscription.plan_type),
                    _excel_text(subscription.get_category_display()),
                    subscription.effective_status.title(),
                    float(subscription.cost),
                    subscription.currency,
                    subscription.get_billing_cycle_display(),
                    float(subscription.monthly_cost),
                    float(subscription.annual_cost),
                    subscription.start_date,
                    subscription.expiry_date,
                    subscription.days_until_expiry,
                    "Yes" if subscription.auto_renew else "No",
                    _excel_text(subscription.department.name if subscription.department else ""),
                    _excel_text(subscription.team),
                    _excel_text(subscription.owner.full_name if subscription.owner else ""),
                    _excel_text(subscription.admin.full_name if subscription.admin else ""),
                    _excel_text(subscription.purpose),
                    _excel_text(subscription.notes),
                ]
            )

        for row in sheet.iter_rows(min_row=2):
            row[5].number_format = "#,##0.00"
            row[8].number_format = "#,##0.00"
            row[9].number_format = "#,##0.00"
            row[10].number_format = "yyyy-mm-dd"
            row[11].number_format = "yyyy-mm-dd"
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for index, column in enumerate(sheet.columns, 1):
            max_length = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[get_column_letter(index)].width = min(max(max_length + 2, 12), 42)

        output = BytesIO()
        workbook.save(output)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="subscriptions-{timezone.localdate().isoformat()}.xlsx"'
        )
        return response

    @staticmethod
    def _export_pdf(subscriptions):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        output = BytesIO()
        document = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=10 * mm,
            leftMargin=10 * mm,
            topMargin=10 * mm,
            bottomMargin=10 * mm,
            title="Software subscription spend report",
        )
        styles = getSampleStyleSheet()
        story = [
            Paragraph("Software Subscription Spend Report", styles["Title"]),
            Paragraph(
                f"Generated {timezone.localtime().strftime('%Y-%m-%d %H:%M')} · "
                f"{len(subscriptions)} subscription(s)",
                styles["Normal"],
            ),
            Spacer(1, 5 * mm),
        ]
        rows = [[
            "Name", "Platform / plan", "Category", "Status", "Billing",
            "Cost", "Monthly", "Annual", "Expiry", "Owner", "Purpose",
        ]]
        for subscription in subscriptions:
            purpose = subscription.purpose
            if len(purpose) > 65:
                purpose = purpose[:62] + "..."
            rows.append(
                [
                    subscription.name,
                    f"{subscription.platform}\n{subscription.plan_type}".strip(),
                    subscription.get_category_display(),
                    subscription.effective_status.title(),
                    subscription.get_billing_cycle_display(),
                    f"{subscription.currency} {_money(subscription.cost)}",
                    f"{subscription.currency} {_money(subscription.monthly_cost)}",
                    f"{subscription.currency} {_money(subscription.annual_cost)}",
                    subscription.expiry_date.isoformat(),
                    subscription.owner.full_name if subscription.owner else "",
                    purpose,
                ]
            )
        table = Table(
            rows,
            repeatRows=1,
            colWidths=[25*mm, 31*mm, 20*mm, 17*mm, 18*mm, 23*mm, 23*mm, 23*mm, 21*mm, 28*mm, 46*mm],
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F6F9")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(table)
        document.build(story)
        response = HttpResponse(output.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="subscriptions-{timezone.localdate().isoformat()}.pdf"'
        )
        return response
