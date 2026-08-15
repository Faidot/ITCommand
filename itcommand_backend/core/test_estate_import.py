"""Bulk import for the Digital Estate.

The promise this feature makes is narrow and worth pinning down: **a sheet
either imports completely or not at all, and you learn about every problem
before anything is written.** Most of these tests exist to stop that promise
quietly regressing.
"""
from decimal import Decimal
from io import BytesIO

from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook
from rest_framework import status
from rest_framework.test import APIClient

from core import estate_import
from core.models import Department, Property, Provider, ProviderAccount, Service
from core.test_helpers import create_role, create_user


def sheet(spec_key, rows, headers=None):
    """An in-memory .xlsx shaped like the template, with `rows` as data.

    Row 1 headers, 2 guidance, 3 example — the same three the template writes,
    so the fixture exercises the real skipping logic rather than a tidier one.
    """
    spec = estate_import.build_specs()[spec_key]
    wb = Workbook()
    ws = wb.active
    ws.append(headers or [c.name for c in spec.columns])
    ws.append([c.help for c in spec.columns])
    ws.append([c.example for c in spec.columns])
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = f"{spec_key}.xlsx"
    return buf


class TemplateTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = create_user("import-admin@example.invalid", "ADMIN")
        self.client.force_authenticate(self.admin)

    def test_a_template_is_produced_for_every_resource(self):
        for key in estate_import.build_specs():
            response = self.client.get(reverse("estate_import_template"), {"resource": key})
            self.assertEqual(response.status_code, status.HTTP_200_OK, key)
            self.assertIn("spreadsheetml", response["Content-Type"])
            self.assertIn(f"estate-{key}-template.xlsx", response["Content-Disposition"])

    def test_the_template_headers_match_what_the_importer_expects(self):
        """If these drift, every download becomes an unusable sheet."""
        spec = estate_import.build_specs()["services"]
        wb = load_workbook(BytesIO(estate_import.build_template(spec)))
        headers = [c.value for c in next(wb[wb.sheetnames[0]].iter_rows(max_row=1))]
        self.assertEqual(headers, [c.name for c in spec.columns])

    def test_the_reference_sheet_lists_the_accepted_codes(self):
        """Otherwise "what can I put here?" becomes a support question."""
        spec = estate_import.build_specs()["services"]
        wb = load_workbook(BytesIO(estate_import.build_template(spec)))
        text = "\n".join(
            str(c.value) for row in wb["Reference"].iter_rows() for c in row if c.value
        )
        self.assertIn("MONTHLY", text)
        self.assertIn("ACTIVE", text)

    def test_an_unknown_resource_is_rejected(self):
        response = self.client.get(reverse("estate_import_template"), {"resource": "wombats"})
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)


class PropertyImportTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = create_user("prop-admin@example.invalid", "ADMIN")
        self.client.force_authenticate(self.admin)
        self.owner = create_user("owner@example.invalid", create_role("P_OWNER", view=True).slug)
        Department.objects.create(name="IT")

    def post(self, url_name, rows, **extra):
        return self.client.post(
            reverse(url_name),
            {"resource": "properties", "file": sheet("properties", rows), **extra},
            format="multipart",
        )

    def test_a_clean_sheet_validates_and_imports(self):
        rows = [["terafort.com", "INFRA", "owner@example.invalid", "IT", "Yes", ""]]

        check = self.post("estate_import_validate", rows)
        self.assertEqual(check.status_code, status.HTTP_200_OK)
        self.assertTrue(check.data["can_commit"])
        self.assertEqual(check.data["to_create"], 1)
        self.assertEqual(Property.objects.count(), 0, "validation must not write")

        done = self.post("estate_import_commit", rows)
        self.assertEqual(done.status_code, status.HTTP_200_OK)
        self.assertEqual(done.data["created"], 1)

        prop = Property.objects.get()
        self.assertEqual(prop.name, "terafort.com")
        self.assertEqual(prop.owner, self.owner)
        self.assertEqual(prop.department.name, "IT")

    def test_one_bad_row_stops_the_whole_sheet(self):
        """The failure this feature exists to prevent."""
        rows = [
            ["good-one.com", "INFRA", "owner@example.invalid", "IT", "Yes", ""],
            ["bad-one.com", "NOT_A_KIND", "owner@example.invalid", "IT", "Yes", ""],
            ["good-two.com", "INFRA", "owner@example.invalid", "IT", "Yes", ""],
        ]
        response = self.post("estate_import_commit", rows)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(
            Property.objects.count(), 0,
            "the two valid rows must not be imported without the third",
        )

    def test_every_problem_is_reported_at_once(self):
        """Fixing a sheet one error per upload would be miserable."""
        rows = [
            ["a.com", "NOT_A_KIND", "owner@example.invalid", "IT", "Yes", ""],
            ["b.com", "INFRA", "nobody@example.invalid", "IT", "Yes", ""],
            ["c.com", "INFRA", "owner@example.invalid", "Nonexistent", "Yes", ""],
        ]
        response = self.post("estate_import_validate", rows)

        self.assertEqual(response.data["invalid"], 3)
        blob = str(response.data["rows"])
        self.assertIn("Kind", blob)
        self.assertIn("Owner email", blob)
        self.assertIn("Department", blob)

    def test_a_missing_required_cell_is_named(self):
        response = self.post("estate_import_validate", [["", "INFRA", "", "", "", ""]])
        self.assertIn("Name is required.", response.data["rows"][0]["errors"])

    def test_an_existing_row_updates_instead_of_duplicating(self):
        Property.objects.create(name="terafort.com", kind="INFRA")
        rows = [["terafort.com", "INFRA", "owner@example.invalid", "IT", "Yes", "note"]]

        check = self.post("estate_import_validate", rows)
        self.assertEqual(check.data["to_update"], 1)
        self.assertEqual(check.data["to_create"], 0)

        self.post("estate_import_commit", rows)
        self.assertEqual(Property.objects.count(), 1)
        self.assertEqual(Property.objects.get().owner, self.owner)

    def test_the_same_record_twice_in_one_sheet_is_rejected(self):
        rows = [
            ["dup.com", "INFRA", "", "", "Yes", ""],
            ["dup.com", "INFRA", "", "", "Yes", ""],
        ]
        response = self.post("estate_import_validate", rows)
        self.assertFalse(response.data["can_commit"])
        self.assertIn("same record as row", str(response.data["rows"][1]["errors"]))

    def test_blank_rows_are_skipped_not_rejected(self):
        rows = [
            ["one.com", "INFRA", "", "", "Yes", ""],
            ["", "", "", "", "", ""],
            ["two.com", "INFRA", "", "", "Yes", ""],
        ]
        response = self.post("estate_import_validate", rows)
        self.assertEqual(response.data["total"], 2)
        self.assertTrue(response.data["can_commit"])

    def test_a_reordered_header_row_is_refused_with_an_explanation(self):
        response = self.client.post(
            reverse("estate_import_validate"),
            {
                "resource": "properties",
                "file": sheet("properties", [], headers=["Nope", "Kind", "Owner email"]),
            },
            format="multipart",
        )
        self.assertFalse(response.data["can_commit"])
        self.assertIn("Missing column", str(response.data["sheet_errors"]))


class ServiceImportTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.client.force_authenticate(create_user("svc-admin@example.invalid", "ADMIN"))
        self.provider = Provider.objects.create(name="Cloudflare", slug="cloudflare")
        self.account = ProviderAccount.objects.create(
            provider=self.provider, account_email="ops@example.invalid"
        )
        self.property = Property.objects.create(name="terafort.com", kind="INFRA")

    def rows(self, **over):
        row = {
            "provider": "Cloudflare", "email": "ops@example.invalid", "type": "DNS",
            "identifier": "terafort.com DNS", "property": "terafort.com",
            "status": "ACTIVE", "renewal": "2027-01-31", "auto": "Yes",
            "cost": "1200", "currency": "PKR", "cycle": "YEARLY",
            "console": "", "descriptor": "", "notes": "",
        }
        row.update(over)
        return [list(row.values())]

    def post(self, url_name, rows):
        return self.client.post(
            reverse(url_name),
            {"resource": "services", "file": sheet("services", rows)},
            format="multipart",
        )

    def test_a_service_imports_with_money_intact(self):
        response = self.post("estate_import_commit", self.rows())
        self.assertEqual(response.status_code, status.HTTP_200_OK, response.data)

        service = Service.objects.get()
        self.assertEqual(service.identifier, "terafort.com DNS")
        self.assertEqual(service.cost, Decimal("1200"))
        self.assertEqual(service.provider_account, self.account)
        self.assertEqual(service.property, self.property)

    def test_the_provider_is_taken_from_the_account_not_asked_for_twice(self):
        """Two columns for one fact is two columns that can disagree."""
        self.post("estate_import_commit", self.rows())
        self.assertEqual(Service.objects.get().provider, self.provider)

    def test_an_account_that_does_not_belong_to_the_provider_is_caught(self):
        other = Provider.objects.create(name="AWS", slug="aws")
        ProviderAccount.objects.create(provider=other, account_email="aws@example.invalid")

        response = self.post("estate_import_validate", self.rows(email="aws@example.invalid"))
        self.assertFalse(response.data["can_commit"])
        self.assertIn("is not an account at Cloudflare", str(response.data["rows"][0]["errors"]))

    def test_a_thousands_separator_in_cost_is_accepted(self):
        """It is what a spreadsheet produces once somebody formats the column."""
        self.post("estate_import_commit", self.rows(cost="1,200.50"))
        self.assertEqual(Service.objects.get().cost, Decimal("1200.50"))

    def test_common_date_formats_are_accepted(self):
        for raw in ("2027-01-31", "31/01/2027", "31-01-2027"):
            Service.objects.all().delete()
            response = self.post("estate_import_commit", self.rows(renewal=raw))
            self.assertEqual(response.status_code, status.HTTP_200_OK, raw)
            self.assertEqual(str(Service.objects.get().renewal_date), "2027-01-31", raw)

    def test_a_nonsense_date_is_rejected_with_the_expected_format(self):
        response = self.post("estate_import_validate", self.rows(renewal="next tuesday"))
        self.assertIn("YYYY-MM-DD", str(response.data["rows"][0]["errors"]))

    def test_a_human_label_is_accepted_as_well_as_the_code(self):
        """The Reference sheet shows both, so both have to work."""
        response = self.post("estate_import_validate", self.rows(cycle="Yearly"))
        self.assertTrue(response.data["can_commit"], response.data["rows"])

    def test_a_blank_property_imports_as_an_orphan(self):
        self.post("estate_import_commit", self.rows(property=""))
        self.assertIsNone(Service.objects.get().property)

    def test_reimporting_the_same_sheet_updates_rather_than_duplicating(self):
        self.post("estate_import_commit", self.rows())
        self.post("estate_import_commit", self.rows(cost="1500"))

        self.assertEqual(Service.objects.count(), 1)
        self.assertEqual(Service.objects.get().cost, Decimal("1500"))


class ImportPermissionTests(TestCase):
    """Bulk creation across a module is not the same authority as one edit."""

    def setUp(self):
        self.url = reverse("estate_import_validate")

    def test_an_estate_editor_who_is_not_an_admin_is_refused(self):
        client = APIClient()
        client.force_authenticate(
            create_user("estate-editor@example.invalid",
                        create_role("EST_EDIT", view=True, add=True, edit=True).slug)
        )
        for name in ("estate_import_options", "estate_import_template"):
            self.assertEqual(
                client.get(reverse(name), {"resource": "properties"}).status_code,
                status.HTTP_403_FORBIDDEN, name,
            )
        response = client.post(
            self.url,
            {"resource": "properties", "file": sheet("properties", [])},
            format="multipart",
        )
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_an_anonymous_caller_is_refused(self):
        self.assertEqual(
            APIClient().get(reverse("estate_import_options")).status_code,
            status.HTTP_401_UNAUTHORIZED,
        )

    def test_an_admin_may_import(self):
        client = APIClient()
        client.force_authenticate(create_user("ok-admin@example.invalid", "ADMIN"))
        self.assertEqual(
            client.get(reverse("estate_import_options")).status_code, status.HTTP_200_OK
        )


class UploadGuardTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.client.force_authenticate(create_user("guard@example.invalid", "ADMIN"))

    def test_a_missing_file_is_a_clear_400(self):
        response = self.client.post(
            reverse("estate_import_validate"), {"resource": "properties"}, format="multipart"
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("Attach", response.data["detail"])

    def test_a_csv_is_refused_rather_than_half_parsed(self):
        csv = BytesIO(b"Name,Kind\nfoo,DOMAIN\n")
        csv.name = "rows.csv"
        response = self.client.post(
            reverse("estate_import_validate"),
            {"resource": "properties", "file": csv},
            format="multipart",
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn(".xlsx", response.data["detail"])

    def test_a_file_that_is_not_a_workbook_is_reported_not_crashed(self):
        junk = BytesIO(b"this is not a spreadsheet")
        junk.name = "rows.xlsx"
        response = self.client.post(
            reverse("estate_import_validate"),
            {"resource": "properties", "file": junk},
            format="multipart",
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertFalse(response.data["can_commit"])
        self.assertIn("could not be read", str(response.data["sheet_errors"]))


class MasterSheetTests(TestCase):
    """One row that creates its own provider, account and property."""

    def setUp(self):
        self.client = APIClient()
        self.client.force_authenticate(create_user("master@example.invalid", "ADMIN"))
        self.owner = create_user("ops-owner@example.invalid", create_role("M_OWNER", view=True).slug)

    def rows(self, **over):
        row = {
            "provider": "Cloudflare", "email": "ops@example.invalid", "type": "DNS",
            "identifier": "terafort.com DNS", "property": "terafort.com",
            "kind": "INFRA", "status": "ACTIVE", "renewal": "2027-01-31",
            "auto": "Yes", "cost": "1200", "currency": "PKR", "cycle": "YEARLY",
            "owner": "ops-owner@example.invalid", "mfa": "APP",
            "console": "", "notes": "",
        }
        row.update(over)
        return [list(row.values())]

    def post(self, url_name, rows):
        return self.client.post(
            reverse(url_name),
            {"resource": "master", "file": sheet("master", rows)},
            format="multipart",
        )

    def test_one_row_creates_the_whole_chain(self):
        """The point of the master sheet: a blank system in one upload."""
        response = self.post("estate_import_commit", self.rows())
        self.assertEqual(response.status_code, status.HTTP_200_OK, response.data)

        provider = Provider.objects.get(name="Cloudflare")
        account = ProviderAccount.objects.get(account_email="ops@example.invalid")
        prop = Property.objects.get(name="terafort.com")
        service = Service.objects.get()

        self.assertEqual(account.provider, provider)
        self.assertEqual(account.owner, self.owner)
        self.assertEqual(account.mfa_type, "APP")
        self.assertEqual(service.provider_account, account)
        self.assertEqual(service.property, prop)
        self.assertEqual(prop.kind, "INFRA")
        self.assertEqual(service.cost, Decimal("1200"))

    def test_validation_says_what_it_will_create_before_writing(self):
        """A misspelled provider silently making a second one is the hazard."""
        response = self.post("estate_import_validate", self.rows())

        self.assertTrue(response.data["can_commit"])
        blob = str(response.data["will_create"])
        self.assertIn("Cloudflare", blob)
        self.assertIn("ops@example.invalid", blob)
        self.assertIn("terafort.com", blob)
        self.assertEqual(Provider.objects.count(), 0, "validation must not write")

    def test_repeated_names_in_one_sheet_create_one_record_each(self):
        rows = self.rows() + self.rows(identifier="terafort.com CDN", type="CDN")
        check = self.post("estate_import_validate", rows)
        # One provider, one account, one property — not two of each.
        self.assertEqual(len(check.data["will_create"]), 3, check.data["will_create"])

        self.post("estate_import_commit", rows)
        self.assertEqual(Provider.objects.count(), 1)
        self.assertEqual(ProviderAccount.objects.count(), 1)
        self.assertEqual(Property.objects.count(), 1)
        self.assertEqual(Service.objects.count(), 2)

    def test_an_existing_provider_is_reused_case_insensitively(self):
        Provider.objects.create(name="Cloudflare", slug="cloudflare")
        self.post("estate_import_commit", self.rows(provider="cloudflare"))
        self.assertEqual(Provider.objects.count(), 1)

    def test_nothing_is_created_when_a_row_fails(self):
        rows = self.rows() + self.rows(identifier="second", type="NOT_A_TYPE")
        response = self.post("estate_import_commit", rows)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(Provider.objects.count(), 0, "no half-built chain")
        self.assertEqual(ProviderAccount.objects.count(), 0)
        self.assertEqual(Service.objects.count(), 0)

    def test_a_blank_property_leaves_the_service_orphaned(self):
        self.post("estate_import_commit", self.rows(property="", kind=""))
        self.assertEqual(Property.objects.count(), 0)
        self.assertIsNone(Service.objects.get().property)

    def test_reimporting_updates_the_service_and_creates_nothing_new(self):
        self.post("estate_import_commit", self.rows())
        self.post("estate_import_commit", self.rows(cost="1500"))

        self.assertEqual(Provider.objects.count(), 1)
        self.assertEqual(Service.objects.count(), 1)
        self.assertEqual(Service.objects.get().cost, Decimal("1500"))


class ExtensibleTypeTests(TestCase):
    """A type added in Settings has to work everywhere, or adding it did nothing."""

    def setUp(self):
        self.client = APIClient()
        self.client.force_authenticate(create_user("types@example.invalid", "ADMIN"))

    def add_type(self, code, label):
        from core.models import ListOfValues

        return ListOfValues.objects.create(
            group="subscription_category", code=code, label=label, is_active=True
        )

    def test_a_custom_service_type_becomes_selectable(self):
        self.add_type("PODCAST", "Podcast hosting")
        from core import estate

        self.assertIn("PODCAST", estate.service_type_codes())
        self.assertEqual(estate.service_type_label("PODCAST"), "Podcast hosting")

    def test_a_custom_type_passes_model_validation(self):
        """Choices are a callable, so this works without a migration."""
        self.add_type("PODCAST", "Podcast hosting")
        provider = Provider.objects.create(name="Acme", slug="acme")
        account = ProviderAccount.objects.create(provider=provider, account_email="a@example.invalid")

        service = Service(
            service_type="PODCAST", identifier="Show", provider=provider,
            provider_account=account, cost=0, billing_cycle="MONTHLY",
        )
        service.full_clean()
        service.save()
        self.assertEqual(Service.objects.get().service_type, "PODCAST")

    def test_the_importer_accepts_a_custom_type(self):
        """It read the frozen tuple before, so a custom type was rejected."""
        self.add_type("PODCAST", "Podcast hosting")
        spec = estate_import.build_specs()["master"]
        column = spec.column("Service type")
        self.assertIn("PODCAST", [c for c, _ in column.choices])

    def test_a_custom_type_never_counts_as_a_stack_gap(self):
        """Only the seven built-in roles are stack positions."""
        self.add_type("PODCAST", "Podcast hosting")
        from core import estate

        self.assertFalse(estate.is_stack_type("PODCAST"))


    def test_adding_a_type_takes_effect_immediately(self):
        """Cached for query count, so invalidation has to be wired up."""
        from core import estate

        estate.service_type_codes()  # warm the cache
        self.add_type("NEWSLETTER", "Newsletter platform")
        self.assertIn("NEWSLETTER", estate.service_type_codes())

    def test_removing_a_type_takes_effect_immediately(self):
        from core import estate

        value = self.add_type("TEMPORARY", "Temporary")
        self.assertIn("TEMPORARY", estate.service_type_codes())
        value.delete()
        self.assertNotIn("TEMPORARY", estate.service_type_codes())

    def test_reading_types_does_not_query_per_call(self):
        """This is what turned a 10-query endpoint into 1380."""
        from core import estate

        self.add_type("PODCAST", "Podcast hosting")
        estate.service_type_codes()  # warm
        with self.assertNumQueries(0):
            for _ in range(50):
                estate.service_type_codes()


    def test_a_custom_type_can_be_tracked_as_a_layer(self):
        """It was accepted in the UI and dropped on save — a silent no-op."""
        from core import estate_reports

        self.add_type("PODCAST", "Podcast hosting")
        settings = estate_reports.estate_settings()
        settings.enabled_layers = ["DNS", "PODCAST"]
        settings.save()

        self.assertIn("PODCAST", estate_reports.tracked_layers(settings))

    def test_tracking_a_custom_type_makes_it_count_as_a_gap(self):
        """Ticking it means "every property should have one" — honour that."""
        from core import estate_reports

        self.add_type("PODCAST", "Podcast hosting")
        settings = estate_reports.estate_settings()
        settings.enabled_layers = ["DNS", "PODCAST"]
        settings.save()

        self.assertIn("PODCAST", estate_reports.tracked_stack_types(settings))

    def test_the_four_builtin_categories_still_never_count_as_gaps(self):
        """The original bug: a stale setting gave every property three amber gaps."""
        from core import estate_reports

        settings = estate_reports.estate_settings()
        settings.enabled_layers = ["DNS", "SAAS", "STORAGE", "MONITORING", "OTHER"]
        settings.save()

        stack = estate_reports.tracked_stack_types(settings)
        self.assertEqual(stack, ["DNS"])

    def test_a_custom_type_appears_in_the_layer_catalog(self):
        """So it is offered under "Not tracked" in Settings."""
        from core import estate_reports

        self.add_type("PODCAST", "Podcast hosting")
        codes = [row["layer"] for row in estate_reports.layer_catalog()]
        self.assertIn("PODCAST", codes)


    def test_reading_types_costs_one_query_per_request_not_one_per_row(self):
        """The regression that hid here twice, pinned at the source.

        First as a missing cache, then as a circular seed: the List of Values
        group seeded itself from `Service.service_type`, whose choices read
        that group. The recursion ended in a RecursionError swallowed by a
        broad except, so it degraded into a query per row rather than failing.
        """
        from django.db import connection
        from django.test.utils import CaptureQueriesContext

        from core import estate

        estate.clear_type_cache()
        with CaptureQueriesContext(connection) as ctx:
            for _ in range(30):
                estate.service_type_choices()
        self.assertLessEqual(
            len(ctx.captured_queries), 1,
            "reading types must not query per call",
        )

    def test_the_builtin_seed_does_not_depend_on_the_model_field(self):
        """Seeding from the field would recreate the circular lookup."""
        from core import estate
        from core.lov import seed_values

        codes = [code for code, _ in seed_values("subscription_category")]
        self.assertEqual(codes, list(estate.SERVICE_TYPE_CODES))
