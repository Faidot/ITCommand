from datetime import timedelta
from decimal import Decimal
from io import BytesIO

from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APIClient

from core import rbac
from core.models import (
    Role,
    SoftwareLicense,
    SoftwareProduct,
    Subscription,
    SubscriptionAssignment,
    LicenseAssignment,
)
from core.test_subscription_assignments import make_subscription
from core.test_subscriptions import create_user


def create_reports_role(slug, *, subscriptions=True, licenses=True):
    permissions = rbac.blank_permissions()
    permissions["reports"] = {"view": True, "add": True, "edit": True, "delete": True}
    permissions["dashboard"] = {"view": True, "add": False, "edit": False, "delete": False}
    permissions["subscriptions"] = {
        "view": subscriptions,
        "add": False,
        "edit": False,
        "delete": False,
    }
    permissions["licenses"] = {
        "view": licenses,
        "add": False,
        "edit": False,
        "delete": False,
    }
    return Role.objects.create(
        slug=slug, name=slug.title(), permissions=permissions
    )


class MasterReportSubscriptionCountTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.role = create_reports_role("REPORTS_VIEWER")
        self.viewer = create_user("reports-viewer@example.com", self.role.slug)
        self.member = create_user("reports-member@example.com", self.role.slug)
        self.client.force_authenticate(self.viewer)

        self.product = SoftwareProduct.objects.create(name="Figma", vendor="Figma Inc")

    def make_license(self, *, license_type="SUBSCRIPTION"):
        return SoftwareLicense.objects.create(
            product=self.product, license_type=license_type
        )

    def counts_for(self, user):
        response = self.client.get(reverse("reports_master_user"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        for row in response.data["users"]:
            if row["email"] == user.email:
                return row
        self.fail(f"No report row for {user.email}")

    def test_a_cross_linked_license_and_subscription_count_once(self):
        """The headline guard: one service must not be counted twice."""
        subscription = make_subscription(name="Figma Org", seats_total=5)
        license_obj = self.make_license()
        subscription.linked_license = license_obj
        subscription.save(update_fields=["linked_license"])

        SubscriptionAssignment.objects.create(
            subscription=subscription, user=self.member
        )
        LicenseAssignment.objects.create(license=license_obj, user=self.member)

        row = self.counts_for(self.member)
        self.assertEqual(
            row["counts"]["subscriptions"],
            1,
            "The same service was counted as both a subscription and a "
            "subscription-type licence.",
        )
        # The licence itself is still reported as a licence.
        self.assertEqual(row["counts"]["licenses"], 1)

    def test_an_unlinked_legacy_subscription_license_still_counts(self):
        license_obj = self.make_license()
        LicenseAssignment.objects.create(license=license_obj, user=self.member)

        row = self.counts_for(self.member)
        self.assertEqual(row["counts"]["subscriptions"], 1)

    def test_a_real_subscription_alone_counts(self):
        subscription = make_subscription(name="Notion", seats_total=5)
        SubscriptionAssignment.objects.create(
            subscription=subscription, user=self.member
        )
        row = self.counts_for(self.member)
        self.assertEqual(row["counts"]["subscriptions"], 1)
        self.assertEqual(row["subscriptions"][0]["name"], "Notion")

    def test_a_perpetual_license_is_never_a_subscription(self):
        license_obj = self.make_license(license_type="PERPETUAL")
        LicenseAssignment.objects.create(license=license_obj, user=self.member)
        row = self.counts_for(self.member)
        self.assertEqual(row["counts"]["subscriptions"], 0)
        self.assertEqual(row["counts"]["licenses"], 1)

    def test_revoked_seats_are_not_counted(self):
        subscription = make_subscription(name="Slack", seats_total=5)
        SubscriptionAssignment.objects.create(
            subscription=subscription, user=self.member, is_active=False
        )
        row = self.counts_for(self.member)
        self.assertEqual(row["counts"]["subscriptions"], 0)
        self.assertEqual(row["subscriptions"], [])


class DashboardSubscriptionBlockTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.today = timezone.localdate()

    def dashboard(self, user):
        self.client.force_authenticate(user)
        response = self.client.get(reverse("main_dashboard"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        return response.data

    def test_block_reports_totals_and_expiring(self):
        user = create_user(
            "dash-viewer@example.com", create_reports_role("DASH_VIEWER").slug
        )
        make_subscription(name="Active soon", expiry_date=self.today + timedelta(days=10))
        make_subscription(
            name="Active later", expiry_date=self.today + timedelta(days=300)
        )
        make_subscription(name="Paused", status="PAUSED")

        block = self.dashboard(user)["subscriptions"]
        self.assertEqual(block["total"], 3)
        self.assertEqual(block["active"], 2)
        self.assertEqual(block["expiring_soon"], 1)
        self.assertEqual(self.dashboard(user)["kpis"]["active_subscriptions"], 2)

    def test_annual_cost_is_never_summed_across_currencies(self):
        user = create_user(
            "dash-currency@example.com", create_reports_role("DASH_CURRENCY").slug
        )
        make_subscription(
            name="USD monthly", cost=Decimal("100.00"), currency="USD",
            billing_cycle="MONTHLY",
        )
        make_subscription(
            name="EUR yearly", cost=Decimal("600.00"), currency="EUR",
            billing_cycle="YEARLY",
        )

        block = self.dashboard(user)["subscriptions"]
        by_currency = {
            row["currency"]: row["annual_cost"] for row in block["annual_cost_by_currency"]
        }
        self.assertEqual(by_currency, {"USD": 1200.0, "EUR": 600.0})

    def test_block_is_empty_without_the_subscriptions_permission(self):
        blocked = create_user(
            "dash-blocked@example.com",
            create_reports_role("DASH_BLOCKED", subscriptions=False).slug,
        )
        make_subscription(name="Hidden")

        block = self.dashboard(blocked)["subscriptions"]
        self.assertEqual(block["total"], 0)
        self.assertEqual(block["active"], 0)
        self.assertEqual(block["annual_cost_by_currency"], [])


class MasterExportSubscriptionSheetTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.role = create_reports_role("EXPORT_VIEWER")
        self.viewer = create_user("export-viewer@example.com", self.role.slug)
        self.member = create_user("export-member@example.com", self.role.slug)
        self.client.force_authenticate(self.viewer)

    def test_workbook_has_a_subscriptions_sheet_with_one_row_per_seat(self):
        subscription = make_subscription(name="Linear", seats_total=3)
        SubscriptionAssignment.objects.create(
            subscription=subscription, user=self.member
        )

        response = self.client.get(reverse("reports_export_master_user"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        workbook = load_workbook(BytesIO(response.content))

        self.assertIn("Subscriptions", workbook.sheetnames)
        sheet = workbook["Subscriptions"]
        self.assertEqual(sheet.max_row, 2)  # header + one seat
        header = [cell.value for cell in sheet[1]]
        self.assertIn("Subscription", header)
        self.assertIn("Currency", header)
        row = [cell.value for cell in sheet[2]]
        self.assertIn("Linear", row)
        self.assertIn(self.member.email, row)
