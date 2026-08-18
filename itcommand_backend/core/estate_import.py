"""Bulk import for the Digital Estate, from a spreadsheet people fill in.

The shape of this is deliberate: **download a template, fill it, validate it,
then commit it.** Validation is a separate step that writes nothing, because
the failure mode that matters is a 300-row sheet with a typo on row 147 — and
finding that out by having rows 1-146 imported and the rest abandoned is worse
than not importing at all.

So:

* the template carries the exact headers, an example row and a Reference sheet
  listing every code that will be accepted, because "what can I put in this
  column?" is otherwise a support question;
* validation reports every bad row at once, not the first one, since fixing a
  spreadsheet one error per round trip is miserable;
* commit runs in a single transaction and re-validates. Nothing lands unless
  all of it lands.

Lookups are by human-readable value (a provider's name, a person's email, a
property's name) rather than database id. Nobody filling in a spreadsheet
knows an id, and a template full of them would be unusable.
"""
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Callable

from django.db import transaction

from core import estate


# ---------------------------------------------------------------------------
# Column specification
# ---------------------------------------------------------------------------

@dataclass
class Column:
    """One spreadsheet column and how to turn its text into a field value."""

    name: str
    help: str
    required: bool = False
    #: Accepted codes, shown on the Reference sheet. Empty means free text.
    choices: tuple = ()
    example: str = ""
    #: Resolves the cleaned string to a model value. Raises ValueError to
    #: reject the row with a message the person can act on.
    resolve: Callable | None = None


@dataclass
class ImportSpec:
    key: str
    label: str
    model_path: str
    columns: list[Column]
    #: Fields that identify an existing row, for update-instead-of-duplicate.
    match_on: tuple = ()
    #: Extra work after the field map is built, e.g. deriving one field from
    #: another. Receives (values dict) and may raise ValueError.
    finalise: Callable | None = None
    notes: str = ""
    #: Record types this sheet may create as a side effect, named in the UI so
    #: the consequence of a typo is visible before committing.
    creates: tuple = ()
    #: Cross-row pass, after every row is coerced. Receives
    #: (spec, results, context) and may add errors or `plan_new` entries.
    #: `context` is shared by every sheet of one workbook, which is how the
    #: Services tab knows about an account the Accounts tab has not written yet.
    resolve_rows: Callable | None = None
    #: Writer, for sheets that create their own dependencies. Receives
    #: (results) and returns (created, updated). None uses the generic writer.
    commit_rows: Callable | None = None

    def column(self, name):
        for c in self.columns:
            if c.name == name:
                return c
        return None


# ---------------------------------------------------------------------------
# Cell coercion
# ---------------------------------------------------------------------------

TRUE_WORDS = {"yes", "y", "true", "1", "on"}
FALSE_WORDS = {"no", "n", "false", "0", "off"}


def as_text(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        # openpyxl reads unformatted numbers as floats; "4242.0" is never what
        # somebody typed into a last-four or a quantity column.
        return str(int(value))
    return str(value).strip()


def as_bool(raw, column):
    text = raw.strip().lower()
    if text in TRUE_WORDS:
        return True
    if text in FALSE_WORDS:
        return False
    raise ValueError(f"{column}: use Yes or No, not {raw!r}.")


def as_date(raw, column):
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"{column}: {raw!r} is not a date. Use YYYY-MM-DD.")


def as_decimal(raw, column):
    try:
        # Thousands separators are what a spreadsheet produces when somebody
        # formats the column, and rejecting them would be pedantry.
        return Decimal(raw.replace(",", "").replace(" ", ""))
    except (InvalidOperation, AttributeError):
        raise ValueError(f"{column}: {raw!r} is not a number.")


def as_ip(raw, column):
    """Validate an address here rather than letting the model reject it later.

    A bad IP is worth stopping at the row it came from: the alternative is a
    model-level failure at commit time that names no row, and an address
    somebody will eventually try to connect to.
    """
    import ipaddress

    try:
        return str(ipaddress.ip_address(raw.strip()))
    except ValueError:
        raise ValueError(f"{column}: {raw!r} is not an IP address.")


def as_choice(raw, column, choices):
    upper = raw.strip().upper().replace(" ", "_")
    valid = {c[0] for c in choices}
    if upper in valid:
        return upper
    # Accept the human label too — somebody reading "Monthly" in the Reference
    # sheet will reasonably type "Monthly" rather than "MONTHLY".
    by_label = {str(c[1]).strip().upper(): c[0] for c in choices}
    if raw.strip().upper() in by_label:
        return by_label[raw.strip().upper()]
    raise ValueError(
        f"{column}: {raw!r} is not valid. Use one of: {', '.join(sorted(valid))}."
    )


# ---------------------------------------------------------------------------
# Lookups
# ---------------------------------------------------------------------------

def find_provider(raw, column):
    from core.models import Provider

    provider = (
        Provider.objects.filter(name__iexact=raw).first()
        or Provider.objects.filter(slug__iexact=raw).first()
    )
    if not provider:
        raise ValueError(f"{column}: no provider called {raw!r}. Add it first.")
    return provider


def find_user(raw, column):
    from core.models import User

    user = User.objects.filter(email__iexact=raw).first()
    if not user:
        raise ValueError(f"{column}: no user with email {raw!r}.")
    return user


def find_property(raw, column):
    from core.models import Property

    prop = Property.objects.filter(name__iexact=raw).first()
    if not prop:
        raise ValueError(f"{column}: no property called {raw!r}.")
    return prop


def find_department(raw, column):
    from core.models import Department

    dept = Department.objects.filter(name__iexact=raw).first()
    if not dept:
        raise ValueError(f"{column}: no department called {raw!r}.")
    return dept


# ---------------------------------------------------------------------------
# Specs
# ---------------------------------------------------------------------------

def build_specs():
    """Built lazily so model imports stay out of module import time."""
    kinds = tuple(estate.property_kind_choices())

    properties = ImportSpec(
        key="properties",
        label="Properties",
        model_path="core.Property",
        match_on=("name",),
        notes="A property is the thing you own — a domain, an app, a site.",
        columns=[
            Column("Name", "Unique. Existing names are updated, not duplicated.",
                   required=True, example="terafort.com"),
            Column("Kind", "What sort of property this is.", required=True,
                   choices=kinds, example=kinds[0][0] if kinds else "",
                   resolve=lambda raw, col: as_choice(raw, col, kinds)),
            Column("Owner email", "Must match an existing user.",
                   example="sam@example.com", resolve=find_user),
            Column("Department", "Must match an existing department name.",
                   example="IT", resolve=find_department),
            Column("Active", "Yes or No. Defaults to Yes.", example="Yes",
                   resolve=as_bool),
            Column("Notes", "Free text.", example=""),
        ],
    )

    accounts = ImportSpec(
        key="accounts",
        label="Provider accounts",
        model_path="core.ProviderAccount",
        match_on=("provider", "account_email"),
        notes=(
            "One login at a provider. Credentials are never imported — attach "
            "a vault entry from the account screen afterwards."
        ),
        columns=[
            Column("Provider", "Provider name or slug. Must already exist.",
                   required=True, example="Cloudflare", resolve=find_provider),
            Column("Account email", "The login. Unique per provider.",
                   required=True, example="ops@example.com"),
            Column("Auth type", "How you sign in.", choices=estate.AUTH_TYPES,
                   example="PASSWORD",
                   resolve=lambda raw, col: as_choice(raw, col, estate.AUTH_TYPES)),
            Column("MFA type", "Second factor, if any.", choices=estate.MFA_TYPES,
                   example="APP",
                   resolve=lambda raw, col: as_choice(raw, col, estate.MFA_TYPES)),
            Column("Owner email", "Must match an existing user.",
                   example="sam@example.com", resolve=find_user),
            Column("Console URL", "Where you log in.",
                   example="https://dash.cloudflare.com"),
            Column("Active", "Yes or No. Defaults to Yes.", example="Yes",
                   resolve=as_bool),
            Column("Notes", "Free text.", example=""),
        ],
    )

    def _derive_provider(values):
        # Service.provider is PROTECTed and non-null, and is always the
        # account's provider. Asking for it twice invites the two disagreeing.
        account = values.get("provider_account")
        if account is not None:
            values["provider"] = account.provider

    services = ImportSpec(
        key="services",
        label="Services",
        model_path="core.Service",
        match_on=("provider_account", "identifier"),
        notes=(
            "A billable thing. The provider is taken from the account, so it "
            "is not a column — it cannot disagree with the login that pays."
        ),
        columns=[
            Column("Provider", "Provider name or slug of the paying account.",
                   required=True, example="Cloudflare"),
            Column("Account email", "Which login at that provider pays for it.",
                   required=True, example="ops@example.com"),
            Column("Service type", "Its role in the stack.", required=True,
                   choices=tuple(estate.service_type_choices()), example="DNS",
                   resolve=lambda raw, col: as_choice(raw, col, estate.service_type_choices())),
            Column("Identifier", "What it is called on the invoice or console.",
                   required=True, example="terafort.com DNS"),
            Column("Property", "What it keeps running. Blank means orphaned.",
                   example="terafort.com", resolve=find_property),
            Column("Status", "Current state.", choices=estate.SERVICE_STATUSES,
                   example="ACTIVE",
                   resolve=lambda raw, col: as_choice(raw, col, estate.SERVICE_STATUSES)),
            Column("Renewal date", "YYYY-MM-DD. Blank if it does not renew.",
                   example="2027-01-31", resolve=as_date),
            Column("Auto renew", "Yes or No.", example="Yes", resolve=as_bool),
            Column("Cost", "Per billing cycle. Numbers only.", example="1200",
                   resolve=as_decimal),
            Column("Currency", "Three-letter code.", example="PKR"),
            Column("Billing cycle", "How often it is charged.",
                   choices=estate.BILLING_CYCLES, example="YEARLY",
                   resolve=lambda raw, col: as_choice(raw, col, estate.BILLING_CYCLES)),
            Column("Console URL", "Where it is managed.", example=""),
            Column("Billing descriptor", "How it appears on the card statement.",
                   example=""),
            Column("Notes", "Free text.", example=""),
        ],
        finalise=_derive_provider,
        resolve_rows=_resolve_service_accounts,
    )

    # ── the master sheet ────────────────────────────────────────────────
    #
    # The three sheets above each require their dependencies to exist first:
    # you cannot import a service until its account exists, and not that until
    # its provider does. For an estate being entered from scratch that means
    # three passes in a fixed order, and knowing that order.
    #
    # This one takes a whole service in a single row and creates whatever is
    # missing beneath it. The cost is that a typo in a provider name silently
    # creates a second provider, so the validation report says exactly what it
    # is about to create *before* anything is written — see `plan_new` below.

    master = ImportSpec(
        key="master",
        label="Master sheet (creates everything)",
        model_path="core.Service",
        match_on=("provider_account", "identifier"),
        creates=("Provider", "Provider account", "Property"),
        notes=(
            "One row per service. Providers, accounts and properties named "
            "here are created if they do not exist yet, so an estate can be "
            "entered from a blank system in one upload. Check the 'will "
            "create' list before importing — a misspelled provider makes a "
            "second provider rather than matching the first."
        ),
        columns=[
            Column("Provider", "Created if it does not exist.", required=True,
                   example="Cloudflare"),
            Column("Account email", "The login that pays. Created under the "
                   "provider if new.", required=True, example="ops@example.com"),
            Column("Service type", "Its role in the stack.", required=True,
                   choices=tuple(estate.service_type_choices()), example="DNS",
                   resolve=lambda raw, col: as_choice(raw, col, estate.service_type_choices())),
            Column("Identifier", "What it is called on the invoice or console.",
                   required=True, example="terafort.com DNS"),
            Column("Property", "Created if new. Blank leaves the service orphaned.",
                   example="terafort.com"),
            Column("Property kind", "Only used when the property is created.",
                   choices=kinds, example=kinds[0][0] if kinds else "",
                   resolve=lambda raw, col: as_choice(raw, col, kinds)),
            Column("Status", "Current state.", choices=estate.SERVICE_STATUSES,
                   example="ACTIVE",
                   resolve=lambda raw, col: as_choice(raw, col, estate.SERVICE_STATUSES)),
            Column("Renewal date", "YYYY-MM-DD. Blank if it does not renew.",
                   example="2027-01-31", resolve=as_date),
            Column("Auto renew", "Yes or No.", example="Yes", resolve=as_bool),
            Column("Cost", "Per billing cycle. Numbers only.", example="1200",
                   resolve=as_decimal),
            Column("Currency", "Three-letter code.", example="PKR"),
            Column("Billing cycle", "How often it is charged.",
                   choices=estate.BILLING_CYCLES, example="YEARLY",
                   resolve=lambda raw, col: as_choice(raw, col, estate.BILLING_CYCLES)),
            Column("Account owner email", "Only used when the account is "
                   "created. Must match an existing user.",
                   example="sam@example.com", resolve=find_user),
            Column("MFA type", "Only used when the account is created.",
                   choices=estate.MFA_TYPES, example="APP",
                   resolve=lambda raw, col: as_choice(raw, col, estate.MFA_TYPES)),
            Column("Console URL", "Where the service is managed.", example=""),
            Column("Notes", "Free text.", example=""),
        ],
        resolve_rows=_resolve_master,
        commit_rows=_commit_master,
    )

    return {s.key: s for s in (master, properties, accounts, services)}


# ---------------------------------------------------------------------------
# The workbook — one file, one tab per record type
# ---------------------------------------------------------------------------
#
# The single sheets above each solve half the problem. The three separate ones
# demand that you already have providers, accounts and properties, in that
# order, across three uploads. The master sheet fixes the ordering but pays for
# it by folding a whole account into every service row: enter twelve services
# on one login and you type that login's owner and MFA twelve times, and the
# twelve had better agree.
#
# This is the shape people actually want. Accounts are entered once, on the
# Accounts tab, with their own columns. Services name the account that pays and
# nothing more. One upload, imported in dependency order inside one
# transaction.
#
# What makes it work is `context`: a dict passed to every tab's resolver in
# order, carrying the names the earlier tabs are about to create. Without it
# the Services tab would reject an account that is sitting three tabs to the
# left, unwritten, and the whole point would be lost.

#: Tab title -> spec key, in dependency order. Order is load-bearing twice:
#: validation walks it so later tabs see earlier tabs' pending records, and
#: commit walks it so a row's dependencies exist by the time it is written.
WORKBOOK_TABS = (
    ("Properties", "wb_properties"),
    ("Accounts", "wb_accounts"),
    ("People", "wb_people"),
    ("Services", "wb_services"),
    ("Servers", "wb_servers"),
)


def _seen(context, bucket):
    return context.setdefault(bucket, set())


def _resolve_wb_properties(spec, results, context):
    """Register property names, so the Services tab does not plan to create them."""
    for result in results:
        name = str(result.values.get("name", "") or "").strip()
        if name:
            _seen(context, "properties").add(name.lower())


def _find_provider_cached(name, memo):
    """Provider by name or slug, looked up once per name per pass.

    Rows overwhelmingly share providers — twenty services on one Cloudflare
    login — so without this a 2000-row workbook runs the same two queries two
    thousand times to reach the same answer.
    """
    key = name.lower()
    if key not in memo:
        from core.models import Provider

        memo[key] = (
            Provider.objects.filter(name__iexact=name).first()
            or Provider.objects.filter(slug__iexact=name).first()
        )
    return memo[key]


def _resolve_wb_accounts(spec, results, context):
    """Resolve the raw provider name, and record the account for later tabs."""
    from core.models import ProviderAccount

    memo = {}
    for result in results:
        provider_raw = str(result.values.pop("_Provider", "") or "").strip()
        email = str(result.values.get("account_email", "") or "").strip()
        if result.errors:
            continue

        provider = _find_provider_cached(provider_raw, memo)
        if not provider and provider_raw.lower() not in _seen(context, "providers"):
            _seen(context, "providers").add(provider_raw.lower())
            result.plan_new.append(f"Provider \u201c{provider_raw}\u201d")
        elif provider:
            _seen(context, "providers").add(provider_raw.lower())

        key = (provider_raw.lower(), email.lower())
        _seen(context, "accounts").add(key)
        # Kept as text: the provider may not exist yet, so there is nothing to
        # point a foreign key at until commit.
        result.values["_provider_name"] = provider_raw
        result.values["_account_key"] = f"{key[0]}|{key[1]}"

        existing = (
            ProviderAccount.objects.filter(
                provider=provider, account_email__iexact=email
            ).first()
            if provider
            else None
        )
        result.action = "update" if existing else "create"


def _resolve_account_ref(result, context, memo, *, what):
    """Resolve a row's Provider + Account email to an account.

    Returns (provider_raw, email, account) with `account` None when the
    Accounts tab is going to create it. Adds an error and returns None when
    the row names an account that exists nowhere — deliberately an error
    rather than a silent create, because these tabs carry no MFA or owner
    column and inventing a login here would record a way in that nobody chose.
    """
    from core.models import ProviderAccount

    provider_raw = str(result.values.pop("_Provider", "") or "").strip()
    email = str(result.values.pop("_Account email", "") or "").strip()
    if result.errors:
        return None

    provider = _find_provider_cached(provider_raw, memo)
    account = (
        ProviderAccount.objects.filter(
            provider=provider, account_email__iexact=email
        ).first()
        if provider
        else None
    )
    key = (provider_raw.lower(), email.lower())
    if account is None and key not in _seen(context, "accounts"):
        result.errors.append(
            f"Account email: no account {email!r} at {provider_raw!r}. "
            f"Add it on the Accounts tab, or fix the spelling ({what})."
        )
        return None

    result.values["_account_key"] = f"{key[0]}|{key[1]}"
    return provider_raw, email, account


def _resolve_wb_people(spec, results, context):
    """Point each login at the account it gets into."""
    from core.models import AccountUser

    memo = {}
    for result in results:
        ref = _resolve_account_ref(result, context, memo, what="this login")
        if ref is None:
            continue
        provider_raw, email, account = ref
        result.values["_new"] = {"provider": provider_raw, "email": email}

        if account is not None:
            result.action = (
                "update"
                if AccountUser.objects.filter(
                    provider_account=account, login=result.values.get("login")
                ).exists()
                else "create"
            )


def _resolve_wb_servers(spec, results, context):
    """Point each server at its account, and plan any property it names."""
    from core.models import Property, Server

    memo = {}
    for result in results:
        property_raw = str(result.values.pop("_Property", "") or "").strip()
        ref = _resolve_account_ref(result, context, memo, what="this server")
        if ref is None:
            continue
        provider_raw, email, account = ref

        if property_raw:
            prop = Property.objects.filter(name__iexact=property_raw).first()
            if not prop and property_raw.lower() not in _seen(context, "properties"):
                _seen(context, "properties").add(property_raw.lower())
                result.plan_new.append(f"Property \u201c{property_raw}\u201d")

        result.values["_new"] = {
            "provider": provider_raw, "email": email, "property": property_raw,
        }
        if account is not None:
            result.action = (
                "update"
                if Server.objects.filter(
                    provider_account=account, name=result.values.get("name")
                ).exists()
                else "create"
            )


def _resolve_wb_services(spec, results, context):
    """Point each service at its account — existing, or pending on the Accounts tab."""
    from core.models import Property, ProviderAccount, Service

    memo = {}
    for result in results:
        provider_raw = str(result.values.pop("_Provider", "") or "").strip()
        email = str(result.values.pop("_Account email", "") or "").strip()
        property_raw = str(result.values.pop("_Property", "") or "").strip()
        if result.errors:
            continue

        key = (provider_raw.lower(), email.lower())
        provider = _find_provider_cached(provider_raw, memo)
        account = (
            ProviderAccount.objects.filter(
                provider=provider, account_email__iexact=email
            ).first()
            if provider
            else None
        )
        if account is None and key not in _seen(context, "accounts"):
            # Deliberately an error rather than a silent create. This tab has
            # no owner or MFA columns, so inventing the account here would
            # quietly produce a login nobody is recorded as holding — and the
            # far likelier cause is a misspelling of a row on the Accounts tab.
            result.errors.append(
                f"Account email: no account {email!r} at {provider_raw!r}. "
                "Add it on the Accounts tab, or fix the spelling."
            )
            continue

        if property_raw:
            prop = Property.objects.filter(name__iexact=property_raw).first()
            if not prop and property_raw.lower() not in _seen(context, "properties"):
                _seen(context, "properties").add(property_raw.lower())
                result.plan_new.append(f"Property \u201c{property_raw}\u201d")

        result.values["_new"] = {
            "provider": provider_raw, "email": email, "property": property_raw,
        }
        result.values["_account_key"] = f"{key[0]}|{key[1]}"

        if account is not None:
            result.action = (
                "update"
                if Service.objects.filter(
                    provider_account=account, identifier=result.values.get("identifier")
                ).exists()
                else "create"
            )


def _get_or_create_provider(name):
    """One provider per name, case-insensitively. Shared by both writers."""
    from django.utils.text import slugify

    from core.models import Provider

    provider = (
        Provider.objects.filter(name__iexact=name).first()
        or Provider.objects.filter(slug__iexact=name).first()
    )
    if provider:
        return provider
    base = slugify(name)[:60] or "provider"
    slug, suffix = base, 2
    while Provider.objects.filter(slug=slug).exists():
        slug = f"{base[:56]}-{suffix}"
        suffix += 1
    return Provider.objects.create(name=name, slug=slug)


def _commit_wb_accounts(results):
    from core.models import ProviderAccount

    created = updated = 0
    for result in results:
        values = dict(result.values)
        provider_name = values.pop("_provider_name", "")
        values.pop("_account_key", None)
        provider = _get_or_create_provider(provider_name)
        email = values.pop("account_email", "")

        account = ProviderAccount.objects.filter(
            provider=provider, account_email__iexact=email
        ).first()
        if account:
            for k, v in values.items():
                setattr(account, k, v)
            account.full_clean()
            account.save()
            updated += 1
        else:
            account = ProviderAccount(provider=provider, account_email=email, **values)
            account.full_clean()
            account.save()
            created += 1
    return created, updated


def _account_for(plan):
    """The account a People or Servers row named, now that it exists."""
    from core.models import ProviderAccount

    provider = _get_or_create_provider(plan.get("provider", ""))
    email = plan.get("email", "")
    account = ProviderAccount.objects.filter(
        provider=provider, account_email__iexact=email
    ).first()
    if account is None:
        # Only reachable if the Accounts tab row that promised this account was
        # not committed. Refuse rather than invent a way in.
        raise ValueError(
            f"Account {email!r} at {provider.name} was not created; "
            "nothing was imported."
        )
    return account


def _commit_wb_people(results):
    from core.models import AccountUser

    created = updated = 0
    for result in results:
        values = dict(result.values)
        account = _account_for(values.pop("_new", {}))
        values.pop("_account_key", None)
        login = values.pop("login", "")

        person = AccountUser.objects.filter(
            provider_account=account, login=login
        ).first()
        if person:
            for key, value in values.items():
                setattr(person, key, value)
            person.full_clean()
            person.save()
            updated += 1
        else:
            person = AccountUser(provider_account=account, login=login, **values)
            person.full_clean()
            person.save()
            created += 1
    return created, updated


def _commit_wb_servers(results):
    from core.models import Property, Server

    created = updated = 0
    for result in results:
        values = dict(result.values)
        plan = values.pop("_new", {})
        account = _account_for(plan)
        values.pop("_account_key", None)
        name = values.pop("name", "")

        property_name = plan.get("property", "")
        if property_name:
            prop = Property.objects.filter(name__iexact=property_name).first()
            if not prop:
                prop = Property.objects.create(name=property_name, kind="OTHER")
            values["property"] = prop

        server = Server.objects.filter(provider_account=account, name=name).first()
        if server:
            for key, value in values.items():
                setattr(server, key, value)
            server.full_clean()
            server.save()
            updated += 1
        else:
            server = Server(provider_account=account, name=name, **values)
            server.full_clean()
            server.save()
            created += 1
    return created, updated


def _commit_wb_services(results):
    from core.models import Property, ProviderAccount, Service

    created = updated = 0
    for result in results:
        values = dict(result.values)
        plan = values.pop("_new", {})
        values.pop("_account_key", None)

        provider = _get_or_create_provider(plan.get("provider", ""))
        email = plan.get("email", "")
        account = ProviderAccount.objects.filter(
            provider=provider, account_email__iexact=email
        ).first()
        if account is None:
            # Only reachable if the Accounts tab row that promised this account
            # was not committed. Refuse rather than invent a login.
            raise ValueError(
                f"Account {email!r} at {provider.name} was not created; "
                "nothing was imported."
            )

        property_name = plan.get("property", "")
        prop = None
        if property_name:
            prop = Property.objects.filter(name__iexact=property_name).first()
            if not prop:
                prop = Property.objects.create(name=property_name, kind="OTHER")

        payload = {**values, "provider": provider, "provider_account": account}
        if prop is not None:
            payload["property"] = prop

        service = Service.objects.filter(
            provider_account=account, identifier=values.get("identifier")
        ).first()
        if service:
            for k, v in payload.items():
                setattr(service, k, v)
            service.full_clean()
            service.save()
            updated += 1
        else:
            service = Service(**payload)
            service.full_clean()
            service.save()
            created += 1
    return created, updated


def build_workbook_specs():
    """The tabs of the workbook, keyed by spec key."""
    kinds = tuple(estate.property_kind_choices())
    types = tuple(estate.service_type_choices())

    properties = ImportSpec(
        key="wb_properties",
        label="Properties",
        model_path="core.Property",
        match_on=("name",),
        notes="Optional. A service can name a property that is not listed here — "
              "it is created as 'Other'. Fill this tab in to set the kind and owner.",
        columns=[
            Column("Name", "Unique. Existing names are updated, not duplicated.",
                   required=True, example="terafort.com"),
            Column("Kind", "What sort of property this is.", required=True,
                   choices=kinds, example=kinds[0][0] if kinds else "",
                   resolve=lambda raw, col: as_choice(raw, col, kinds)),
            Column("Owner email", "Must match an existing user.",
                   example="sam@example.com", resolve=find_user),
            Column("Department", "Must match an existing department name.",
                   example="IT", resolve=find_department),
            Column("Active", "Yes or No. Defaults to Yes.", example="Yes",
                   resolve=as_bool),
            Column("Notes", "Free text.", example=""),
        ],
        resolve_rows=_resolve_wb_properties,
    )

    accounts = ImportSpec(
        key="wb_accounts",
        label="Accounts",
        model_path="core.ProviderAccount",
        match_on=("_account_key",),
        creates=("Provider",),
        notes="One row per login. The provider is created if it does not exist. "
              "Credentials are never imported — attach a vault entry afterwards.",
        columns=[
            Column("Provider", "Created if it does not exist yet.",
                   required=True, example="Cloudflare"),
            Column("Account email", "The login. Unique per provider.",
                   required=True, example="ops@example.com"),
            Column("Auth type", "How you sign in.", choices=estate.AUTH_TYPES,
                   example="PASSWORD",
                   resolve=lambda raw, col: as_choice(raw, col, estate.AUTH_TYPES)),
            Column("MFA type", "Second factor, if any.", choices=estate.MFA_TYPES,
                   example="APP",
                   resolve=lambda raw, col: as_choice(raw, col, estate.MFA_TYPES)),
            Column("Owner email", "Must match an existing user.",
                   example="sam@example.com", resolve=find_user),
            Column("Console URL", "Where you log in.",
                   example="https://dash.cloudflare.com"),
            Column("Active", "Yes or No. Defaults to Yes.", example="Yes",
                   resolve=as_bool),
            Column("Notes", "Free text.", example=""),
        ],
        resolve_rows=_resolve_wb_accounts,
        commit_rows=_commit_wb_accounts,
    )

    services = ImportSpec(
        key="wb_services",
        label="Services",
        model_path="core.Service",
        match_on=("_account_key", "identifier"),
        creates=("Property",),
        notes="One row per billable thing. Provider and Account email must name "
              "a row on the Accounts tab, or an account that already exists.",
        columns=[
            Column("Provider", "Pick from the Accounts tab.", required=True,
                   example="Cloudflare"),
            Column("Account email", "Which login at that provider pays for it.",
                   required=True, example="ops@example.com"),
            Column("Service type", "Its role in the stack.", required=True,
                   choices=types, example="DNS",
                   resolve=lambda raw, col: as_choice(raw, col, types)),
            Column("Identifier", "What it is called on the invoice or console.",
                   required=True, example="terafort.com DNS"),
            Column("Property", "What it keeps running. Created if new.",
                   example="terafort.com"),
            Column("Status", "Current state.", choices=estate.SERVICE_STATUSES,
                   example="ACTIVE",
                   resolve=lambda raw, col: as_choice(raw, col, estate.SERVICE_STATUSES)),
            Column("Renewal date", "YYYY-MM-DD. Blank if it does not renew.",
                   example="2027-01-31", resolve=as_date),
            Column("Auto renew", "Yes or No.", example="Yes", resolve=as_bool),
            Column("Cost", "Per billing cycle. Numbers only.", example="1200",
                   resolve=as_decimal),
            Column("Currency", "Three-letter code.", example="PKR"),
            Column("Billing cycle", "How often it is charged.",
                   choices=estate.BILLING_CYCLES, example="YEARLY",
                   resolve=lambda raw, col: as_choice(raw, col, estate.BILLING_CYCLES)),
            Column("Console URL", "Where it is managed.", example=""),
            Column("Billing descriptor", "How it appears on the card statement.",
                   example=""),
            Column("Notes", "Free text.", example=""),
        ],
        resolve_rows=_resolve_wb_services,
        commit_rows=_commit_wb_services,
    )

    people = ImportSpec(
        key="wb_people",
        label="People",
        model_path="core.AccountUser",
        match_on=("_account_key", "login"),
        notes="One row per person who can sign in. The account is one bill; "
              "these are the ways into it, and they do not share a second factor.",
        columns=[
            Column("Provider", "Which account they get into.", required=True,
                   example="AWS"),
            Column("Account email", "The account's own login, from the Accounts tab.",
                   required=True, example="ops@example.com"),
            Column("Login", "What this person types. A username is fine.",
                   required=True, example="iam:alice"),
            Column("Login is a", "What that string is.", choices=estate.LOGIN_KINDS,
                   example="USERNAME",
                   resolve=lambda raw, col: as_choice(raw, col, estate.LOGIN_KINDS)),
            Column("Person email", "Links this login to someone in IT Command. "
                   "Leave blank for a service account.",
                   example="alice@example.com", resolve=find_user),
            Column("Name", "Only used when Person email is blank.",
                   example="Deploy robot"),
            Column("Role", "What they can do.", choices=estate.ACCOUNT_ROLES,
                   example="ADMIN",
                   resolve=lambda raw, col: as_choice(raw, col, estate.ACCOUNT_ROLES)),
            Column("Second factor", "Theirs, not the account's.",
                   choices=estate.MFA_TYPES, example="APP",
                   resolve=lambda raw, col: as_choice(raw, col, estate.MFA_TYPES)),
            Column("Notes", "Free text.", example=""),
        ],
        resolve_rows=_resolve_wb_people,
        commit_rows=_commit_wb_people,
    )

    servers = ImportSpec(
        key="wb_servers",
        label="Servers",
        model_path="core.Server",
        match_on=("_account_key", "name"),
        creates=("Property",),
        notes="One row per machine. It belongs to the account that pays for it, "
              "and to the property it keeps running.",
        columns=[
            Column("Provider", "Which account pays for it.", required=True,
                   example="AWS"),
            Column("Account email", "The account's own login, from the Accounts tab.",
                   required=True, example="ops@example.com"),
            Column("Server name", "Hostname or console name.", required=True,
                   example="web-01"),
            Column("Role", "What the box is for.", choices=estate.SERVER_ROLES,
                   example="WEB",
                   resolve=lambda raw, col: as_choice(raw, col, estate.SERVER_ROLES)),
            Column("Environment", "Production, staging and so on.",
                   choices=estate.SERVER_ENVIRONMENTS, example="PRODUCTION",
                   resolve=lambda raw, col: as_choice(raw, col, estate.SERVER_ENVIRONMENTS)),
            Column("Status", "Current state.", choices=estate.SERVER_STATUSES,
                   example="RUNNING",
                   resolve=lambda raw, col: as_choice(raw, col, estate.SERVER_STATUSES)),
            Column("Public IP", "Blank if it has none.", example="203.0.113.9",
                   resolve=as_ip),
            Column("Private IP", "Blank if it has none.", example="10.0.0.4",
                   resolve=as_ip),
            Column("Hostname", "Full DNS name, if it has one.",
                   example="web-01.terafort.com"),
            Column("Region", "Where it runs.", example="eu-west-1"),
            Column("Size", "Instance type or plan.", example="t3.medium"),
            Column("Operating system", "", example="Ubuntu 24.04"),
            Column("Property", "What it keeps running. Created if new.",
                   example="terafort.com"),
            Column("Owner email", "Must match an existing user.",
                   example="sam@example.com", resolve=find_user),
            Column("Notes", "Free text.", example=""),
        ],
        resolve_rows=_resolve_wb_servers,
        commit_rows=_commit_wb_servers,
    )

    return {s.key: s for s in (properties, accounts, people, services, servers)}


#: Column name -> model field. Kept apart from the spec so the sheet can be
#: worded for people while the model keeps its own names.
FIELD_MAP = {
    "properties": {
        "Name": "name", "Kind": "kind", "Owner email": "owner",
        "Department": "department", "Active": "is_active", "Notes": "notes",
    },
    "accounts": {
        "Provider": "provider", "Account email": "account_email",
        "Auth type": "auth_type", "MFA type": "mfa_type", "Owner email": "owner",
        "Console URL": "console_url", "Active": "is_active", "Notes": "notes",
    },
    "master": {
        "Service type": "service_type", "Identifier": "identifier",
        "Status": "status", "Renewal date": "renewal_date",
        "Auto renew": "auto_renew", "Cost": "cost", "Currency": "currency",
        "Billing cycle": "billing_cycle", "Console URL": "console_url",
        "Notes": "notes",
    },
    "services": {
        "Service type": "service_type", "Identifier": "identifier",
        "Property": "property", "Status": "status", "Renewal date": "renewal_date",
        "Auto renew": "auto_renew", "Cost": "cost", "Currency": "currency",
        "Billing cycle": "billing_cycle", "Console URL": "console_url",
        "Billing descriptor": "billing_descriptor", "Notes": "notes",
    },
    # Workbook tabs. Columns naming a record that may not exist yet — Provider,
    # Account email, Property — are deliberately absent, so they arrive as
    # `_<column name>` raw text for the resolvers above to deal with rather
    # than being assigned straight to a foreign key.
    "wb_properties": {
        "Name": "name", "Kind": "kind", "Owner email": "owner",
        "Department": "department", "Active": "is_active", "Notes": "notes",
    },
    "wb_accounts": {
        "Account email": "account_email", "Auth type": "auth_type",
        "MFA type": "mfa_type", "Owner email": "owner",
        "Console URL": "console_url", "Active": "is_active", "Notes": "notes",
    },
    "wb_people": {
        "Login": "login", "Login is a": "login_kind", "Person email": "user",
        "Name": "display_name", "Role": "role", "Second factor": "mfa_type",
        "Notes": "notes",
    },
    "wb_servers": {
        "Server name": "name", "Role": "server_role", "Environment": "environment",
        "Status": "status", "Public IP": "public_ip", "Private IP": "private_ip",
        "Hostname": "hostname", "Region": "region", "Size": "size",
        "Operating system": "operating_system", "Owner email": "owner",
        "Notes": "notes",
    },
    "wb_services": {
        "Service type": "service_type", "Identifier": "identifier",
        "Status": "status", "Renewal date": "renewal_date",
        "Auto renew": "auto_renew", "Cost": "cost", "Currency": "currency",
        "Billing cycle": "billing_cycle", "Console URL": "console_url",
        "Billing descriptor": "billing_descriptor", "Notes": "notes",
    },
}


# ---------------------------------------------------------------------------
# Template
# ---------------------------------------------------------------------------

def build_template(spec):
    """An .xlsx with headers, an example row, guidance and valid codes."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = spec.label[:31]

    header_fill = PatternFill("solid", fgColor="1F2937")
    required_fill = PatternFill("solid", fgColor="7F1D1D")

    for i, col in enumerate(spec.columns, start=1):
        cell = ws.cell(row=1, column=i, value=col.name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = required_fill if col.required else header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Row 2 explains the column, row 3 shows one filled-in example. Both
        # are removed on import, so the sheet can be handed back as-is.
        ws.cell(row=2, column=i, value=col.help).font = Font(italic=True, size=9, color="6B7280")
        ws.cell(row=3, column=i, value=col.example)

        width = max(len(col.name), min(38, len(col.help) // 2), 14)
        ws.column_dimensions[get_column_letter(i)].width = width + 2

    ws.freeze_panes = "A4"

    guide = wb.create_sheet("Reference")
    guide.append(["Read this first"])
    guide["A1"].font = Font(bold=True, size=13)
    for line in [
        spec.notes,
        "",
        "Row 1 is the header — do not rename or reorder it.",
        "Row 2 is guidance and row 3 is an example. Both are ignored on import;",
        "you may overwrite them or leave them in place.",
        "Columns shaded dark red are required.",
        "Blank cells are left at their default rather than cleared.",
        "",
        "Upload validates the whole sheet before writing anything. If any row",
        "fails, nothing is imported — fix the sheet and upload it again.",
        "",
    ]:
        guide.append([line])

    for col in spec.columns:
        if not col.choices:
            continue
        guide.append([])
        guide.append([f"{col.name} — accepted values"])
        guide.cell(row=guide.max_row, column=1).font = Font(bold=True)
        for code, label in col.choices:
            guide.append([code, str(label)])
    guide.column_dimensions["A"].width = 30
    guide.column_dimensions["B"].width = 46

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Workbook template
# ---------------------------------------------------------------------------

#: How many rows the dropdowns and date formats reach. Validation is stored as
#: a range, so this costs nothing in file size; it is a limit on how far down
#: the conveniences apply, not on how many rows may be imported.
TEMPLATE_ROWS = 400

#: First row of real data. Row 1 is the header, 2 the guidance, 3 the example.
FIRST_DATA_ROW = 4

#: Columns whose dropdown reads a sibling tab instead of a fixed list. This is
#: what makes the workbook feel joined up: type a login on the Accounts tab and
#: it is in the Services tab's dropdown immediately, with no upload in between.
LIVE_RANGES = {
    ("wb_services", "Provider"): "Accounts!$A${first}:$A${last}",
    ("wb_services", "Account email"): "Accounts!$B${first}:$B${last}",
    ("wb_services", "Property"): "Properties!$A${first}:$A${last}",
    ("wb_people", "Provider"): "Accounts!$A${first}:$A${last}",
    ("wb_people", "Account email"): "Accounts!$B${first}:$B${last}",
    ("wb_servers", "Provider"): "Accounts!$A${first}:$A${last}",
    ("wb_servers", "Account email"): "Accounts!$B${first}:$B${last}",
    ("wb_servers", "Property"): "Properties!$A${first}:$A${last}",
}

def _workbook_lists(specs):
    """Every fixed dropdown list, and which column uses which.

    Keyed by the *values*, not by the column name. Two tabs can use the same
    word for different things — "Role" is an account role on People and a
    server role on Servers, "Status" is a service status on Services and a
    server status on Servers — and keying by name silently pointed the second
    one at the first one's list. Strictly, so Excel then refused the very
    values the sheet was asking for.

    Identical lists still share one column, which is the behaviour worth
    keeping: Accounts' "MFA type" and People's "Second factor" hold the same
    codes and have no business being stored twice.

    Returns (lists, assignments): name -> values, and (spec key, column name)
    -> list name.
    """
    from core.models import Department, Provider, User

    lists = {}
    by_values = {}
    assignments = {}

    def register(preferred, values):
        """Name this set of values, reusing an identical set if there is one."""
        values = tuple(values)
        if not values:
            return None
        if values in by_values:
            return by_values[values]
        name, suffix = preferred, 2
        while name in lists:
            name = f"{preferred} {suffix}"
            suffix += 1
        lists[name] = values
        by_values[values] = name
        return name

    for _, key in WORKBOOK_TABS:
        spec = specs[key]
        for col in spec.columns:
            if col.choices:
                name = register(col.name, (code for code, _ in col.choices))
            elif col.resolve is as_bool:
                name = register("Yes or No", ("Yes", "No"))
            else:
                continue
            if name:
                assignments[(spec.key, col.name)] = name

    # Lists drawn from the database rather than from a spec. Registered after
    # the choice columns so a name collision renames these, not the fixed ones.
    dynamic = {
        "Existing providers": tuple(
            Provider.objects.order_by("name").values_list("name", flat=True)[:400]
        ),
        "User emails": tuple(
            User.objects.filter(is_active=True).order_by("email")
            .values_list("email", flat=True)[:400]
        ),
        "Departments": tuple(
            Department.objects.order_by("name").values_list("name", flat=True)[:400]
        ),
    }
    try:
        from core.lov import get_values

        dynamic["Currencies"] = tuple(code for code, _ in get_values("currency"))
    except Exception:
        dynamic["Currencies"] = ("USD", "EUR", "GBP", "PKR")

    for name, values in dynamic.items():
        if values and name not in lists:
            lists[name] = tuple(values)

    return lists, assignments


#: Free-text columns that still deserve a dropdown of what already exists.
#: Lenient — a value not in the list is accepted, because these columns are
#: allowed to name something new. Strict ones would make the template refuse
#: the first provider anybody ever adds.
_SUGGESTED = {
    "Provider": ("Existing providers", False),
    "Currency": ("Currencies", False),
    "Owner email": ("User emails", True),
    "Person email": ("User emails", True),
    "Department": ("Departments", True),
}


def _column_validation(spec, col, lists, letters, assignments):
    """(formula1, strict) for a column's dropdown, or None for free text."""
    live = LIVE_RANGES.get((spec.key, col.name))
    if live:
        return live.format(first=FIRST_DATA_ROW, last=TEMPLATE_ROWS), False

    assigned = assignments.get((spec.key, col.name))
    if assigned:
        name, strict = assigned, True
    elif col.name in _SUGGESTED:
        name, strict = _SUGGESTED[col.name]
    else:
        return None

    if name not in letters:
        # Nothing to offer — no users, no departments, a fresh install. An
        # empty dropdown blocks the column entirely, which is worse than none.
        return None
    column = letters[name]
    return f"Lists!${column}$2:${column}${len(lists[name]) + 1}", strict


def _write_tab(ws, spec, lists, letters, assignments):
    """Headers, guidance, example, dropdowns and date cells for one tab."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    header_fill = PatternFill("solid", fgColor="1F2937")
    required_fill = PatternFill("solid", fgColor="7F1D1D")

    for i, col in enumerate(spec.columns, start=1):
        letter = get_column_letter(i)
        cell = ws.cell(row=1, column=i, value=col.name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = required_fill if col.required else header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

        ws.cell(row=2, column=i, value=col.help).font = Font(
            italic=True, size=9, color="6B7280"
        )
        ws.cell(row=3, column=i, value=col.example)
        ws.column_dimensions[letter].width = (
            max(len(col.name), min(38, len(col.help) // 2), 14) + 2
        )

        span = f"{letter}{FIRST_DATA_ROW}:{letter}{TEMPLATE_ROWS}"

        validation = _column_validation(spec, col, lists, letters, assignments)
        if validation:
            formula, strict = validation
            rule = DataValidation(
                type="list", formula1=formula, allow_blank=True,
                showErrorMessage=strict, showInputMessage=True,
            )
            rule.promptTitle = col.name[:32]
            rule.prompt = col.help[:255]
            if strict:
                rule.errorTitle = "Not an accepted value"
                rule.error = f"Pick one of the values offered for {col.name}."[:255]
            ws.add_data_validation(rule)
            rule.add(span)

        if col.resolve is as_date:
            # Formatted rather than validated. A hard date rule rejects the
            # text "2027-01-31", which `as_date` reads perfectly well, and
            # different spreadsheet apps disagree about what counts as a date.
            for row in range(FIRST_DATA_ROW, TEMPLATE_ROWS + 1):
                ws.cell(row=row, column=i).number_format = "yyyy-mm-dd"
            rule = DataValidation(
                type="date", operator="between",
                formula1="DATE(1990,1,1)", formula2="DATE(2100,12,31)",
                allow_blank=True, showErrorMessage=False, showInputMessage=True,
            )
            rule.promptTitle = "Date"
            rule.prompt = "Pick a date, or type it as YYYY-MM-DD."
            ws.add_data_validation(rule)
            rule.add(span)

    ws.freeze_panes = f"A{FIRST_DATA_ROW}"


def build_workbook_template():
    """The multi-tab workbook: one tab per record type, dropdowns in the cells."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    specs = build_workbook_specs()
    lists, assignments = _workbook_lists(specs)
    letters = {
        name: get_column_letter(i) for i, name in enumerate(lists, start=1)
    }

    wb = Workbook()
    readme = wb.active
    readme.title = "Read me"
    readme["A1"] = "Digital Estate workbook"
    readme["A1"].font = Font(bold=True, size=15)
    for line in [
        "",
        "Fill in the tabs left to right. You do not have to fill in all of them.",
        "",
        "  Properties   Optional. Only needed to set a property's kind and owner.",
        "  Accounts     One row per account. The provider is created if it is new.",
        "  People       One row per person who can sign in to an account above.",
        "  Services     One row per billable thing, pointing at an account above.",
        "  Servers      One row per machine, pointing at an account above.",
        "",
        "Cells with a small arrow have a dropdown. On the Services tab the",
        "Provider, Account email and Property dropdowns read the tabs to their",
        "left, so a login you type on the Accounts tab can be picked here",
        "straight away — there is no need to import in between.",
        "",
        "Date cells are formatted as YYYY-MM-DD. Typing that text works too.",
        "",
        "Row 1 is the header — do not rename or reorder it.",
        "Row 2 is guidance and row 3 is an example. Both are ignored on import.",
        "Columns shaded dark red are required.",
        "Blank cells keep the record's default rather than clearing it.",
        "",
        "Upload checks every tab before writing anything. If one row anywhere",
        "in the workbook fails, nothing at all is imported — fix it and upload",
        "again. Credentials are never imported; attach vault entries afterwards.",
        "",
        "The Lists tab holds the dropdown values. Leave it in place.",
    ]:
        readme.append([line])
    readme.column_dimensions["A"].width = 96

    for title, key in WORKBOOK_TABS:
        _write_tab(wb.create_sheet(title), specs[key], lists, letters, assignments)

    sheet = wb.create_sheet("Lists")
    for i, (name, values) in enumerate(lists.items(), start=1):
        head = sheet.cell(row=1, column=i, value=name)
        head.font = Font(bold=True, color="FFFFFF")
        head.fill = PatternFill("solid", fgColor="1F2937")
        head.alignment = Alignment(wrap_text=True)
        for row, value in enumerate(values, start=2):
            sheet.cell(row=row, column=i, value=str(value))
        sheet.column_dimensions[get_column_letter(i)].width = max(
            14, min(34, len(name) + 4)
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Reading and validating
# ---------------------------------------------------------------------------

@dataclass
class RowResult:
    row: int
    values: dict = field(default_factory=dict)
    errors: list = field(default_factory=list)
    action: str = "create"  # or "update"
    #: Human descriptions of records this row will bring into existence, e.g.
    #: ["Provider 'Cloudflare'"]. Shown before committing, because a
    #: misspelled name silently creating a second provider is the one real
    #: hazard of a sheet that creates its own dependencies.
    plan_new: list = field(default_factory=list)


#: Row 1 headers, 2 guidance, 3 example. Data starts at 4 — but a sheet where
#: somebody deleted the guidance rows is still accepted, see `_first_data_row`.
HEADER_ROW = 1


def _first_data_row(ws, spec):
    """Skip the guidance and example rows if they are still present.

    People delete them, keep them, or overwrite the example with real data.
    All three have to work, so the rows are detected rather than assumed.
    """
    first_col = spec.columns[0].name
    for row_idx in (2, 3):
        value = as_text(ws.cell(row=row_idx, column=1).value)
        if value in (spec.columns[0].help, spec.columns[0].example) and value != "":
            continue
        return row_idx
    return 4


def validate_records(spec, records, limit=2000, context=None):
    """Validate rows given as {column name: raw text}. Returns (results, errors).

    The one validation path. Each sheet in the workbook turns its rows into
    these dicts and calls this. A second path would be a second set of rules to
    keep in step, and the one that skipped a check would be the one that
    corrupted data.

    Writes nothing.
    """
    field_map = FIELD_MAP[spec.key]
    results = []
    seen_keys = {}

    for offset, record in enumerate(records):
        if len(results) >= limit:
            return results, [f"That is more than {limit} rows. Split it into batches."]
        cells = {key: as_text(value) for key, value in (record or {}).items()}
        if not any(cells.values()):
            continue

        result = RowResult(row=int(cells.pop("__row__", 0) or offset + 1))
        for col in spec.columns:
            raw = cells.get(col.name, "")

            if not raw:
                if col.required:
                    result.errors.append(f"{col.name} is required.")
                continue

            try:
                value = col.resolve(raw, col.name) if col.resolve else raw
            except ValueError as exc:
                result.errors.append(str(exc))
                continue

            target = field_map.get(col.name)
            if target:
                result.values[target] = value
            else:
                result.values[f"_{col.name}"] = value

        if spec.finalise and not result.errors:
            try:
                spec.finalise(result.values)
            except ValueError as exc:
                result.errors.append(str(exc))

        results.append(result)

    if spec.resolve_rows:
        spec.resolve_rows(spec, results, context if context is not None else {})
    _mark_duplicates(spec, results, seen_keys)
    if spec.commit_rows is None:
        # A sheet with its own writer works out create-vs-update itself; the
        # generic check would look up raw text against a foreign key column.
        _mark_existing(spec, results)
    return results, []


def read_rows(file_obj, spec, limit=2000, context=None):
    """Parse and validate a single-sheet .xlsx. Returns (results, sheet_errors)."""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(file_obj, data_only=True, read_only=True)
    except Exception:
        return [], ["That file could not be read as an .xlsx workbook."]

    ws = wb[wb.sheetnames[0]]
    headers = [as_text(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]

    expected = [c.name for c in spec.columns]
    missing = [h for h in expected if h not in headers]
    if missing:
        return [], [
            "The header row does not match the template. Missing column(s): "
            + ", ".join(missing)
            + ". Download a fresh template rather than editing the headers."
        ]

    return validate_records(
        spec, _records_from_sheet(ws, spec, headers), limit=limit, context=context
    )


def _records_from_sheet(ws, spec, headers):
    """Column-keyed dicts from a worksheet. Reading only, no validation."""
    index = {h: i for i, h in enumerate(headers)}
    start = _first_data_row(ws, spec)

    records = []
    for row_idx, raw_row in enumerate(
        ws.iter_rows(min_row=start, values_only=True), start=start
    ):
        cells = [as_text(v) for v in raw_row]
        if not any(cells):
            continue  # blank spacer row
        record = {"__row__": row_idx}
        for col in spec.columns:
            pos = index[col.name]
            record[col.name] = cells[pos] if pos < len(cells) else ""
        records.append(record)
    return records


def _resolve_master(spec, results, context):
    """Work out, per row, what already exists and what will be created.

    Nothing is written here — validation must stay side-effect free. Each row
    records the objects it *would* create so the report can show them, and
    keeps the raw text for `commit` to act on.

    Names seen earlier in the same sheet count as existing, so twenty services
    at one provider plan one provider, not twenty.
    """
    from core.models import Property, Provider, ProviderAccount

    planned_providers = set()
    planned_properties = set()
    planned_accounts = set()

    for result in results:
        provider_raw = str(result.values.pop("_Provider", "") or "").strip()
        email_raw = str(result.values.pop("_Account email", "") or "").strip()
        property_raw = str(result.values.pop("_Property", "") or "").strip()
        kind = result.values.pop("_Property kind", "") or "OTHER"
        owner = result.values.pop("_Account owner email", None)
        mfa = result.values.pop("_MFA type", "") or ""

        if result.errors:
            continue

        result.values["_new"] = {
            "provider": provider_raw, "email": email_raw,
            "property": property_raw, "kind": kind,
            "owner": owner, "mfa": mfa,
        }

        provider = (
            Provider.objects.filter(name__iexact=provider_raw).first()
            or Provider.objects.filter(slug__iexact=provider_raw).first()
        )
        if not provider and provider_raw.lower() not in planned_providers:
            planned_providers.add(provider_raw.lower())
            result.plan_new.append(f"Provider “{provider_raw}”")

        account_key = (provider_raw.lower(), email_raw.lower())
        account = None
        if provider:
            account = ProviderAccount.objects.filter(
                provider=provider, account_email__iexact=email_raw
            ).first()
        if not account and account_key not in planned_accounts:
            planned_accounts.add(account_key)
            result.plan_new.append(f"Account “{email_raw}” at {provider_raw}")

        if property_raw:
            existing = Property.objects.filter(name__iexact=property_raw).first()
            if not existing and property_raw.lower() not in planned_properties:
                planned_properties.add(property_raw.lower())
                result.plan_new.append(f"Property “{property_raw}”")

        # A row is an update only when the exact service already exists.
        if account:
            match = __import__("django.apps", fromlist=["apps"]).apps.get_model(
                "core.Service"
            ).objects.filter(
                provider_account=account, identifier=result.values.get("identifier")
            ).exists()
            result.action = "update" if match else "create"


def _resolve_service_accounts(spec, results, context):
    """Turn the service sheet's Provider + Account email into one account."""
    from core.models import ProviderAccount

    for result in results:
        provider_raw = result.values.pop("_Provider", None)
        email_raw = result.values.pop("_Account email", None)
        if result.errors and (provider_raw is None or email_raw is None):
            continue
        if not provider_raw or not email_raw:
            continue
        try:
            provider = find_provider(provider_raw, "Provider")
        except ValueError as exc:
            result.errors.append(str(exc))
            continue
        account = ProviderAccount.objects.filter(
            provider=provider, account_email__iexact=email_raw
        ).first()
        if not account:
            result.errors.append(
                f"Account email: {email_raw!r} is not an account at {provider.name}. "
                "Import the account first, or fix the spelling."
            )
            continue
        result.values["provider_account"] = account
        result.values["provider"] = provider


def _row_key(spec, values):
    parts = []
    for f in spec.match_on:
        value = values.get(f)
        parts.append(str(getattr(value, "pk", value)).lower() if value is not None else "")
    return tuple(parts)


def _mark_duplicates(spec, results, seen):
    """Two rows describing the same record is a mistake worth naming."""
    for result in results:
        if result.errors or not spec.match_on:
            continue
        key = _row_key(spec, result.values)
        if not any(key):
            continue
        if key in seen:
            result.errors.append(
                f"This is the same record as row {seen[key]}. Remove one of them."
            )
        else:
            seen[key] = result.row


def _mark_existing(spec, results):
    """Flag rows that will update rather than create."""
    from django.apps import apps

    if not spec.match_on:
        return
    model = apps.get_model(spec.model_path)
    for result in results:
        if result.errors:
            continue
        lookup = {}
        for f in spec.match_on:
            value = result.values.get(f)
            if value is None:
                lookup = {}
                break
            lookup[f] = value
        if not lookup:
            continue
        if model.objects.filter(**lookup).exists():
            result.action = "update"


def read_workbook(file_obj, limit=2000):
    """Read and validate every tab of the workbook. Returns (sheets, errors).

    `sheets` is a list of (spec, results) in dependency order, holding only the
    tabs that carry data. Tabs left blank are skipped rather than reported as a
    problem: filling in only the Services tab against accounts that already
    exist is a perfectly ordinary use of this file.

    The `context` dict is threaded through every tab in order. That is what
    lets the Services tab accept a login the Accounts tab has not written yet.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(file_obj, data_only=True, read_only=True)
    except Exception:
        return [], ["That file could not be read as an .xlsx workbook."]

    specs = build_workbook_specs()
    present = [t for t, _ in WORKBOOK_TABS if t in wb.sheetnames]
    if not present:
        return [], [
            "No recognised tabs. This workbook needs at least one of: "
            + ", ".join(t for t, _ in WORKBOOK_TABS)
            + ". Download the template again."
        ]

    context = {}
    sheets = []
    errors = []
    for title, key in WORKBOOK_TABS:
        if title not in wb.sheetnames:
            continue
        spec = specs[key]
        ws = wb[title]
        headers = [
            as_text(c.value)
            for c in next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW))
        ]
        missing = [c.name for c in spec.columns if c.name not in headers]
        if missing:
            errors.append(
                f"{title} tab is missing these columns: {', '.join(missing)}. "
                "Download the template again."
            )
            continue

        records = _records_from_sheet(ws, spec, headers)
        if not records:
            continue  # tab left blank
        results, sheet_errors = validate_records(
            spec, records, limit=limit, context=context
        )
        errors.extend(f"{title} tab: {e}" for e in sheet_errors)
        sheets.append((spec, results))

    if not sheets and not errors:
        errors.append("Every tab is empty. Fill one in and upload it again.")
    return sheets, errors


@transaction.atomic
def commit_workbook(sheets):
    """Import every tab, in dependency order, or none of them.

    One transaction across all three tabs, not one per tab. A workbook whose
    Services tab fails must not leave its accounts behind — the person would
    have to work out which half landed before they could safely upload again.
    """
    if any(r.errors for _, results in sheets for r in results):
        raise ValueError("Refusing to import a workbook that has errors.")

    totals = {}
    for spec, results in sheets:
        created, updated = commit(spec, results)
        totals[spec.key] = {
            "label": spec.label, "created": created, "updated": updated,
        }
    return totals


# ---------------------------------------------------------------------------
# Commit
# ---------------------------------------------------------------------------

@transaction.atomic
def commit(spec, results):
    """Write every row, or none. Returns (created, updated).

    Atomic on purpose: a partial import leaves somebody reconciling a
    spreadsheet against a database by hand, which is worse than a failed
    upload.
    """
    from django.apps import apps

    if any(r.errors for r in results):
        raise ValueError("Refusing to import a sheet that has errors.")

    if spec.commit_rows:
        return spec.commit_rows(results)

    model = apps.get_model(spec.model_path)
    created = updated = 0

    for result in results:
        values = dict(result.values)
        lookup = {f: values.pop(f) for f in spec.match_on if f in values}
        if lookup and model.objects.filter(**lookup).exists():
            obj = model.objects.get(**lookup)
            for k, v in values.items():
                setattr(obj, k, v)
            obj.full_clean(exclude=[f.name for f in obj._meta.fields if f.name not in values])
            obj.save()
            updated += 1
        else:
            obj = model(**{**lookup, **values})
            obj.full_clean()
            obj.save()
            created += 1

    return created, updated


def _commit_master(results):
    """Create the provider, account and property a row needs, then the service.

    Called from inside `commit`, so it is already in that transaction: a row
    that fails halfway does not leave a provider and an account behind with no
    service attached.

    get_or_create on a case-insensitive lookup rather than a bare create, so
    "Cloudflare" and "cloudflare" in the same sheet resolve to one provider.
    """
    from django.utils.text import slugify

    from core.models import Property, Provider, ProviderAccount, Service

    created = updated = 0

    for result in results:
        plan = result.values.pop("_new", {})
        values = dict(result.values)

        provider_name = plan.get("provider", "")
        provider = (
            Provider.objects.filter(name__iexact=provider_name).first()
            or Provider.objects.filter(slug__iexact=provider_name).first()
        )
        if not provider:
            base = slugify(provider_name)[:60] or "provider"
            slug = base
            # Two different names can slugify the same; the slug is unique.
            suffix = 2
            while Provider.objects.filter(slug=slug).exists():
                slug = f"{base[:56]}-{suffix}"
                suffix += 1
            provider = Provider.objects.create(name=provider_name, slug=slug)

        email = plan.get("email", "")
        account = ProviderAccount.objects.filter(
            provider=provider, account_email__iexact=email
        ).first()
        if not account:
            account = ProviderAccount.objects.create(
                provider=provider,
                account_email=email,
                owner=plan.get("owner"),
                mfa_type=plan.get("mfa") or "UNKNOWN",
            )

        property_name = plan.get("property", "")
        prop = None
        if property_name:
            prop = Property.objects.filter(name__iexact=property_name).first()
            if not prop:
                prop = Property.objects.create(
                    name=property_name, kind=plan.get("kind") or "OTHER"
                )

        service = Service.objects.filter(
            provider_account=account, identifier=values.get("identifier")
        ).first()
        payload = {**values, "provider": provider, "provider_account": account}
        if prop is not None:
            payload["property"] = prop

        if service:
            for key, value in payload.items():
                setattr(service, key, value)
            service.full_clean()
            service.save()
            updated += 1
        else:
            service = Service(**payload)
            service.full_clean()
            service.save()
            created += 1

    return created, updated
