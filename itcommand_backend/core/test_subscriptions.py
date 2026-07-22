from datetime import timedelta
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APIClient

from core import rbac
from core.models import (
    AppSettings,
    AuditLog,
    BudgetCategory,
    Department,
    Expense,
    Role,
    Subscription,
    SubscriptionSettings,
    Vendor,
    VendorContract,
)
from core.serializers import ExpenseSerializer


User = get_user_model()


def create_role(slug, *, view=False, add=False, edit=False, delete=False):
    permissions = rbac.blank_permissions()
    permissions["subscriptions"] = {
        "view": view,
        "add": add,
        "edit": edit,
        "delete": delete,
    }
    return Role.objects.create(
        slug=slug,
        name=slug.replace("_", " ").title(),
        permissions=permissions,
    )


def create_user(email, role):
    return User.objects.create_user(
        email=email,
        password="SubscriptionTestPassword!1",
        full_name=email.split("@")[0].title(),
        role=role,
    )


def subscription_payload(**overrides):
    today = timezone.localdate()
    payload = {
        "name": "ChatGPT Business",
        "platform": "OpenAI",
        "plan_type": "Business",
        "category": "AI",
        "cost": "100.00",
        "currency": "USD",
        "billing_cycle": "MONTHLY",
        "start_date": today.isoformat(),
        "expiry_date": (today + timedelta(days=30)).isoformat(),
        "purpose": "Customer support automation",
        "team": "Support",
        "status": "ACTIVE",
        "auto_renew": True,
        "renewal_reminder_enabled": True,
        "renewal_reminder_days": 14,
        "cancellation_reminder_enabled": True,
        "cancellation_reminder_days": 7,
        "notes": "Review seats before renewal.",
    }
    payload.update(overrides)
    return payload


class SubscriptionCrudAndPermissionTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.manager_role = create_role(
            "SUB_MANAGER", view=True, add=True, edit=True, delete=True
        )
        self.viewer_role = create_role("SUB_VIEWER", view=True)
        self.blocked_role = create_role("SUB_BLOCKED")
        self.manager = create_user("manager@example.com", self.manager_role.slug)
        self.viewer = create_user("viewer@example.com", self.viewer_role.slug)
        self.blocked = create_user("blocked@example.com", self.blocked_role.slug)
        self.department = Department.objects.create(name="Engineering")

    def test_crud_sets_creator_and_audits_changes(self):
        self.client.force_authenticate(self.manager)
        owner = create_user("owner@example.com", self.viewer_role.slug)
        response = self.client.post(
            reverse("subscription-list"),
            subscription_payload(
                department=self.department.pk,
                owner=owner.pk,
                admin=self.manager.pk,
            ),
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED, response.data)
        subscription = Subscription.objects.get(pk=response.data["id"])
        self.assertEqual(subscription.created_by, self.manager)
        self.assertEqual(response.data["department_name"], "Engineering")
        self.assertEqual(response.data["owner_email"], owner.email)
        self.assertEqual(response.data["effective_status"], "ACTIVE")
        self.assertEqual(response.data["monthly_cost"], "100.00")
        self.assertEqual(response.data["annual_cost"], "1200.00")
        self.assertTrue(
            AuditLog.objects.filter(
                user=self.manager,
                action="CREATE",
                model_name="Subscription",
                object_id=str(subscription.pk),
            ).exists()
        )

        response = self.client.patch(
            reverse("subscription-detail", args=[subscription.pk]),
            {"cost": "125.50", "currency": "usd"},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK, response.data)
        self.assertEqual(response.data["currency"], "USD")
        self.assertEqual(response.data["annual_cost"], "1506.00")

        response = self.client.delete(reverse("subscription-detail", args=[subscription.pk]))
        self.assertEqual(response.status_code, status.HTTP_204_NO_CONTENT)
        self.assertFalse(Subscription.objects.filter(pk=subscription.pk).exists())

    def test_rbac_is_enforced_for_crud_and_collection_actions(self):
        Subscription.objects.create(
            created_by=self.manager,
            **{
                key: value
                for key, value in subscription_payload().items()
                if key not in {"start_date", "expiry_date"}
            },
            start_date=timezone.localdate(),
            expiry_date=timezone.localdate() + timedelta(days=30),
        )

        self.client.force_authenticate(self.blocked)
        self.assertEqual(
            self.client.get(reverse("subscription-list")).status_code,
            status.HTTP_403_FORBIDDEN,
        )
        self.assertEqual(
            self.client.get(reverse("subscription-dashboard")).status_code,
            status.HTTP_403_FORBIDDEN,
        )

        self.client.force_authenticate(self.viewer)
        for route in (
            "subscription-list",
            "subscription-dashboard",
            "subscription-options",
            "subscription-settings",
        ):
            self.assertEqual(self.client.get(reverse(route)).status_code, status.HTTP_200_OK)
        self.assertEqual(
            self.client.post(reverse("subscription-list"), subscription_payload(), format="json").status_code,
            status.HTTP_403_FORBIDDEN,
        )
        self.assertEqual(
            self.client.patch(
                reverse("subscription-settings"),
                {"monthly_budget_threshold": "500.00"},
                format="json",
            ).status_code,
            status.HTTP_403_FORBIDDEN,
        )

    def test_options_only_expose_all_users_to_subscription_editors(self):
        Subscription.objects.create(
            created_by=self.manager,
            owner=self.manager,
            admin=self.viewer,
            **{
                key: value
                for key, value in subscription_payload().items()
                if key not in {"start_date", "expiry_date"}
            },
            start_date=timezone.localdate(),
            expiry_date=timezone.localdate() + timedelta(days=30),
        )

        self.client.force_authenticate(self.manager)
        response = self.client.get(reverse("subscription-options"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            {item["email"] for item in response.data["users"]},
            {self.manager.email, self.viewer.email, self.blocked.email},
        )
        alert_eligibility = {
            item["email"]: item["can_receive_subscription_alerts"]
            for item in response.data["users"]
        }
        self.assertTrue(alert_eligibility[self.manager.email])
        self.assertTrue(alert_eligibility[self.viewer.email])
        self.assertFalse(alert_eligibility[self.blocked.email])
        self.assertEqual(response.data["departments"], [{"id": self.department.pk, "name": "Engineering"}])
        self.assertIn({"value": "AI", "label": "AI tools"}, response.data["categories"])

        self.client.force_authenticate(self.viewer)
        response = self.client.get(reverse("subscription-options"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["users"], [])

        # Read-only users still receive identities actually referenced by each
        # subscription; they do not need the company-wide user directory.
        response = self.client.get(reverse("subscription-list"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        item = response.data["results"][0]
        self.assertEqual(item["owner_name"], self.manager.full_name)
        self.assertEqual(item["owner_email"], self.manager.email)
        self.assertEqual(item["admin_name"], self.viewer.full_name)
        self.assertEqual(item["admin_email"], self.viewer.email)

    def test_api_create_uses_current_company_reminder_defaults(self):
        SubscriptionSettings.objects.update_or_create(
            pk=1,
            defaults={
                "default_renewal_reminder_days": 21,
                "default_cancellation_reminder_days": 5,
            },
        )
        payload = subscription_payload()
        payload.pop("renewal_reminder_days")
        payload.pop("cancellation_reminder_days")
        self.client.force_authenticate(self.manager)

        response = self.client.post(
            reverse("subscription-list"),
            payload,
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED, response.data)
        self.assertEqual(response.data["renewal_reminder_days"], 21)
        self.assertEqual(response.data["cancellation_reminder_days"], 5)


class SubscriptionValidationTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.role = create_role("SUB_EDITOR", view=True, add=True, edit=True)
        self.user = create_user("editor@example.com", self.role.slug)
        self.client.force_authenticate(self.user)

    def test_dates_cost_currency_and_cancellation_deadline_are_validated(self):
        today = timezone.localdate()
        invalid_cases = [
            subscription_payload(cost="-0.01"),
            subscription_payload(currency="US"),
            subscription_payload(currency="ZZZ"),
            subscription_payload(
                start_date=today.isoformat(),
                expiry_date=(today - timedelta(days=1)).isoformat(),
            ),
            subscription_payload(
                cancellation_deadline=(today + timedelta(days=31)).isoformat()
            ),
        ]
        for payload in invalid_cases:
            with self.subTest(payload=payload):
                response = self.client.post(
                    reverse("subscription-list"), payload, format="json"
                )
                self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST, response.data)

    def test_current_iso_currency_codes_are_accepted(self):
        response = self.client.post(
            reverse("subscription-list"),
            subscription_payload(currency="xcg"),
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED, response.data)
        self.assertEqual(response.data["currency"], "XCG")

    def test_unchanged_legacy_currency_remains_editable(self):
        today = timezone.localdate()
        subscription = Subscription.objects.create(
            name="Legacy billed service",
            platform="Legacy",
            category="OTHER",
            cost=Decimal("10.00"),
            currency="ABC",
            billing_cycle="MONTHLY",
            start_date=today,
            expiry_date=today + timedelta(days=30),
            created_by=self.user,
        )

        response = self.client.patch(
            reverse("subscription-detail", args=[subscription.pk]),
            {"currency": "abc", "cost": "11.00"},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK, response.data)
        self.assertEqual(response.data["currency"], "ABC")

        response = self.client.patch(
            reverse("subscription-detail", args=[subscription.pk]),
            {"currency": "ZZZ"},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST, response.data)

    def test_settings_reject_non_iso_currency(self):
        response = self.client.patch(
            reverse("subscription-settings"),
            {"budget_currency": "ZZZ"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST, response.data)
        self.assertIn("budget_currency", response.data)

    def test_settings_with_legacy_currency_remain_editable(self):
        SubscriptionSettings.objects.update_or_create(
            pk=1,
            defaults={"budget_currency": "ABC", "notifications_enabled": True},
        )

        response = self.client.patch(
            reverse("subscription-settings"),
            {"notifications_enabled": False},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK, response.data)
        self.assertEqual(response.data["budget_currency"], "ABC")
        self.assertFalse(response.data["notifications_enabled"])

        response = self.client.patch(
            reverse("subscription-settings"),
            {"budget_currency": "abc"},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK, response.data)
        self.assertEqual(response.data["budget_currency"], "ABC")

    def test_computed_status_and_normalized_costs(self):
        today = timezone.localdate()
        subscription = Subscription.objects.create(
            name="AWS",
            platform="Amazon Web Services",
            category="CLOUD",
            cost=Decimal("1200.00"),
            currency="USD",
            billing_cycle="YEARLY",
            start_date=today - timedelta(days=365),
            expiry_date=today - timedelta(days=1),
        )
        self.assertEqual(subscription.effective_status, "EXPIRED")
        self.assertEqual(subscription.days_until_expiry, -1)
        self.assertEqual(subscription.monthly_cost, Decimal("100.00"))
        self.assertEqual(subscription.annual_cost, Decimal("1200.00"))

        subscription.status = "PAUSED"
        self.assertEqual(subscription.effective_status, "PAUSED")


class SubscriptionSettingsInitializationTests(TestCase):
    def test_first_load_is_singleton_and_uses_company_currency(self):
        AppSettings.objects.update_or_create(
            key="default_currency",
            defaults={"value": "pkr"},
        )
        SubscriptionSettings.objects.all().delete()

        first = SubscriptionSettings.get_solo()
        second = SubscriptionSettings.get_solo()

        self.assertEqual(first.pk, second.pk)
        self.assertEqual(first.budget_currency, "PKR")
        self.assertEqual(SubscriptionSettings.objects.count(), 1)

    def test_invalid_company_currency_falls_back_to_usd(self):
        AppSettings.objects.update_or_create(
            key="default_currency",
            defaults={"value": "ZZZ"},
        )
        SubscriptionSettings.objects.all().delete()

        settings = SubscriptionSettings.get_solo()

        self.assertEqual(settings.budget_currency, "USD")


class SubscriptionDashboardAndExportTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.role = create_role("SUB_REPORTER", view=True, edit=True)
        self.user = create_user("reporter@example.com", self.role.slug)
        self.client.force_authenticate(self.user)
        today = timezone.localdate()
        common = {
            "category": "AI",
            "start_date": today - timedelta(days=10),
            "status": "ACTIVE",
            "created_by": self.user,
        }
        self.monthly = Subscription.objects.create(
            name="ChatGPT",
            platform="OpenAI",
            plan_type="Business",
            cost=Decimal("100.00"),
            currency="USD",
            billing_cycle="MONTHLY",
            expiry_date=today + timedelta(days=20),
            purpose="Support automation",
            **common,
        )
        Subscription.objects.create(
            name="Claude",
            platform="Anthropic",
            plan_type="Team",
            cost=Decimal("1200.00"),
            currency="USD",
            billing_cycle="YEARLY",
            expiry_date=today + timedelta(days=90),
            **common,
        )
        Subscription.objects.create(
            name="Figma",
            platform="Figma",
            category="DESIGN",
            cost=Decimal("240.00"),
            currency="EUR",
            billing_cycle="YEARLY",
            start_date=today - timedelta(days=100),
            expiry_date=today + timedelta(days=100),
            status="ACTIVE",
            created_by=self.user,
        )
        Subscription.objects.create(
            name="Old tool",
            platform="Legacy",
            category="OTHER",
            cost=Decimal("10.00"),
            currency="USD",
            billing_cycle="MONTHLY",
            start_date=today - timedelta(days=100),
            expiry_date=today - timedelta(days=1),
            status="ACTIVE",
            created_by=self.user,
        )
        SubscriptionSettings.objects.create(
            pk=1,
            budget_currency="USD",
            monthly_budget_threshold=Decimal("190.00"),
            yearly_budget_threshold=Decimal("2300.00"),
        )

    def test_dashboard_never_sums_different_currencies(self):
        response = self.client.get(reverse("subscription-dashboard"), {"days": 30})
        self.assertEqual(response.status_code, status.HTTP_200_OK, response.data)
        self.assertEqual(response.data["default_currency"], "USD")
        self.assertEqual(response.data["monthly_spend"], "200.00")
        self.assertEqual(response.data["yearly_spend"], "2400.00")
        self.assertEqual(response.data["status_counts"]["ACTIVE"], 3)
        self.assertEqual(response.data["status_counts"]["EXPIRED"], 1)
        self.assertEqual(response.data["upcoming_count"], 1)
        self.assertEqual(response.data["upcoming_renewals"][0]["name"], "ChatGPT")
        by_currency = {item["currency"]: item for item in response.data["spend_by_currency"]}
        self.assertEqual(by_currency["EUR"]["yearly_spend"], "240.00")
        self.assertEqual(by_currency["USD"]["yearly_spend"], "2400.00")
        self.assertTrue(response.data["budget"]["monthly_exceeded"])
        self.assertTrue(response.data["budget"]["yearly_exceeded"])

    def test_settings_update_and_threshold_validation(self):
        response = self.client.patch(
            reverse("subscription-settings"),
            {
                "budget_currency": "eur",
                "monthly_budget_threshold": "500.00",
                "yearly_budget_threshold": "5000.00",
            },
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK, response.data)
        self.assertEqual(response.data["budget_currency"], "EUR")
        self.assertEqual(response.data["updated_by"], self.user.pk)

        for invalid_threshold in ("0.00", "-1.00"):
            with self.subTest(invalid_threshold=invalid_threshold):
                response = self.client.patch(
                    reverse("subscription-settings"),
                    {"monthly_budget_threshold": invalid_threshold},
                    format="json",
                )
                self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_search_and_effective_status_filters(self):
        response = self.client.get(reverse("subscription-list"), {"search": "openai"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 1)
        response = self.client.get(reverse("subscription-list"), {"status": "EXPIRED"})
        self.assertEqual(response.data["count"], 1)
        self.assertEqual(response.data["results"][0]["name"], "Old tool")

        today = timezone.localdate()
        response = self.client.get(
            reverse("subscription-list"),
            {
                "expiry_from": (today + timedelta(days=80)).isoformat(),
                "expiry_to": (today + timedelta(days=95)).isoformat(),
            },
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 1)
        self.assertEqual(response.data["results"][0]["name"], "Claude")

    def test_dashboard_rounds_after_aggregating_small_yearly_costs(self):
        today = timezone.localdate()
        for index in range(12):
            Subscription.objects.create(
                name=f"Tiny yearly tool {index}",
                platform="Small SaaS",
                category="SAAS",
                cost=Decimal("0.01"),
                currency="USD",
                billing_cycle="YEARLY",
                start_date=today - timedelta(days=1),
                expiry_date=today + timedelta(days=200),
                status="ACTIVE",
                created_by=self.user,
            )

        response = self.client.get(reverse("subscription-dashboard"))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["monthly_spend"], "200.01")
        self.assertEqual(response.data["yearly_spend"], "2400.12")

    def test_xlsx_and_pdf_are_real_downloads(self):
        self.monthly.notes = "=HYPERLINK(\"https://example.invalid\",\"click\")"
        self.monthly.save(update_fields=["notes"])
        xlsx = self.client.get(reverse("subscription-export"), {"format": "xlsx"})
        self.assertEqual(xlsx.status_code, status.HTTP_200_OK)
        self.assertEqual(
            xlsx["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(xlsx.content), read_only=True)
        self.assertEqual(workbook["Subscriptions"]["A1"].value, "Name")
        self.assertEqual(workbook["Subscriptions"].max_row, 5)
        chatgpt_row = next(
            row for row in workbook["Subscriptions"].iter_rows(values_only=True)
            if row[0] == "ChatGPT"
        )
        self.assertTrue(chatgpt_row[19].startswith("'="))

        pdf = self.client.get(reverse("subscription-export"), {"format": "pdf"})
        self.assertEqual(pdf.status_code, status.HTTP_200_OK)
        self.assertEqual(pdf["Content-Type"], "application/pdf")
        self.assertTrue(pdf.content.startswith(b"%PDF-"))
        self.assertGreater(len(pdf.content), 1000)

    def test_export_rejects_unknown_format(self):
        response = self.client.get(reverse("subscription-export"), {"format": "csv"})
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)


class SubscriptionLinkTests(TestCase):
    """Phase 1: FKs into vendors, contracts, budgets, vault and licenses."""

    def setUp(self):
        self.client = APIClient()
        self.manager_role = create_role(
            "LINK_MANAGER", view=True, add=True, edit=True, delete=True
        )
        self.viewer_role = create_role("LINK_VIEWER", view=True)
        self.manager = create_user("link-manager@example.com", self.manager_role.slug)
        self.viewer = create_user("link-viewer@example.com", self.viewer_role.slug)

        self.vendor = Vendor.objects.create(name="OpenAI Inc")
        self.other_vendor = Vendor.objects.create(name="Vercel Inc")
        self.contract = VendorContract.objects.create(
            vendor=self.vendor, title="OpenAI Enterprise Agreement"
        )
        self.other_contract = VendorContract.objects.create(
            vendor=self.other_vendor, title="Vercel Pro Agreement"
        )
        self.budget_category = BudgetCategory.objects.create(name="Software")

    def _create(self, **overrides):
        self.client.force_authenticate(self.manager)
        return self.client.post(
            reverse("subscription-list"),
            subscription_payload(**overrides),
            format="json",
        )

    def test_links_round_trip_with_display_names(self):
        response = self._create(
            vendor=self.vendor.pk,
            vendor_contract=self.contract.pk,
            budget_category=self.budget_category.pk,
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED, response.data)
        self.assertEqual(response.data["vendor_name"], "OpenAI Inc")
        self.assertEqual(
            response.data["vendor_contract_title"], "OpenAI Enterprise Agreement"
        )
        self.assertEqual(
            response.data["vendor_contract_number"], self.contract.contract_number
        )
        self.assertEqual(response.data["budget_category_name"], "Software")
        # platform is unchanged by the vendor link — both coexist
        self.assertEqual(response.data["platform"], "OpenAI")

    def test_links_default_to_null_when_omitted(self):
        response = self._create()
        self.assertEqual(response.status_code, status.HTTP_201_CREATED, response.data)
        for key in (
            "vendor",
            "vendor_contract",
            "budget_category",
            "vault_credential",
            "linked_license",
        ):
            self.assertIsNone(response.data[key], key)

    def test_contract_from_a_different_vendor_is_rejected(self):
        response = self._create(
            vendor=self.vendor.pk, vendor_contract=self.other_contract.pk
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("vendor_contract", response.data)

    def test_contract_without_a_vendor_is_allowed(self):
        response = self._create(vendor_contract=self.contract.pk)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED, response.data)

    def test_deleting_a_vendor_leaves_the_subscription_intact(self):
        subscription_id = self._create(vendor=self.vendor.pk).data["id"]
        self.vendor.delete()
        subscription = Subscription.objects.get(pk=subscription_id)
        self.assertIsNone(subscription.vendor_id)

    def test_filter_and_search_by_vendor(self):
        linked = self._create(vendor=self.vendor.pk).data["id"]
        self._create(name="Unlinked Service")

        response = self.client.get(
            reverse("subscription-list"), {"vendor": self.vendor.pk}
        )
        self.assertEqual([row["id"] for row in response.data["results"]], [linked])

        response = self.client.get(reverse("subscription-list"), {"search": "OpenAI Inc"})
        self.assertEqual([row["id"] for row in response.data["results"]], [linked])

    def test_filter_by_budget_category(self):
        linked = self._create(budget_category=self.budget_category.pk).data["id"]
        self._create(name="No Budget Category")
        response = self.client.get(
            reverse("subscription-list"), {"budget_category": self.budget_category.pk}
        )
        self.assertEqual([row["id"] for row in response.data["results"]], [linked])

    def test_expense_links_to_a_subscription(self):
        subscription = Subscription.objects.get(pk=self._create().data["id"])
        expense = Expense.objects.create(
            title="ChatGPT February",
            amount=Decimal("100.00"),
            expense_date=timezone.localdate(),
            paid_to="OpenAI",
            linked_subscription=subscription,
        )
        self.assertEqual(list(subscription.expenses.all()), [expense])
        data = ExpenseSerializer(expense).data
        self.assertEqual(data["linked_subscription"], subscription.pk)
        self.assertEqual(data["linked_subscription_name"], subscription.name)

    def test_options_gate_cross_module_selectors(self):
        self.client.force_authenticate(self.manager)
        response = self.client.get(reverse("subscription-options"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            [row["name"] for row in response.data["vendors"]],
            ["OpenAI Inc", "Vercel Inc"],
        )
        self.assertEqual(
            {row["title"] for row in response.data["contracts"]},
            {"OpenAI Enterprise Agreement", "Vercel Pro Agreement"},
        )
        self.assertEqual(
            [row["name"] for row in response.data["budget_categories"]], ["Software"]
        )
        # Vault and licenses have their own RBAC modules; a subscriptions
        # manager without those permissions must not be able to enumerate them.
        self.assertEqual(response.data["vault_credentials"], [])
        self.assertEqual(response.data["licenses"], [])

    def test_viewer_cannot_enumerate_vendors_or_budgets(self):
        self.client.force_authenticate(self.viewer)
        response = self.client.get(reverse("subscription-options"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["vendors"], [])
        self.assertEqual(response.data["contracts"], [])
        self.assertEqual(response.data["budget_categories"], [])


class SubscriptionDetailPayloadTests(TestCase):
    """The retrieve payload must stay a strict superset of the list payload."""

    def setUp(self):
        self.client = APIClient()
        self.manager_role = create_role(
            "DETAIL_MANAGER", view=True, add=True, edit=True, delete=True
        )
        self.manager = create_user("detail-manager@example.com", self.manager_role.slug)
        self.client.force_authenticate(self.manager)
        self.subscription = Subscription.objects.create(
            created_by=self.manager,
            **{
                key: value
                for key, value in subscription_payload().items()
                if key not in {"start_date", "expiry_date"}
            },
            start_date=timezone.localdate(),
            expiry_date=timezone.localdate() + timedelta(days=30),
        )

    def test_detail_is_a_superset_of_the_list_row(self):
        list_response = self.client.get(reverse("subscription-list"))
        list_row = list_response.data["results"][0]

        detail = self.client.get(
            reverse("subscription-detail", args=[self.subscription.pk])
        )
        self.assertEqual(detail.status_code, status.HTTP_200_OK)

        missing = set(list_row) - set(detail.data)
        self.assertEqual(missing, set(), f"Detail dropped list fields: {missing}")
        for key in ("assignments", "renewals", "expenses"):
            self.assertIn(key, detail.data)

    def test_the_list_stays_lean(self):
        response = self.client.get(reverse("subscription-list"))
        row = response.data["results"][0]
        for key in ("assignments", "renewals", "expenses"):
            self.assertNotIn(key, row, f"List payload should not embed {key}")

    def test_detail_retrieval_stays_query_bounded(self):
        for index in range(3):
            member = create_user(f"detail-seat-{index}@example.com", self.manager_role.slug)
            self.client.post(
                reverse("subscription-assign-seat", args=[self.subscription.pk]),
                {"user_id": member.pk},
                format="json",
            )
        url = reverse("subscription-detail", args=[self.subscription.pk])
        self.client.get(url)  # warm any lazily-cached lookups
        # Prefetched, so the roster does not cost one query per seat holder.
        # One query per nested relation: assignments, renewals, expenses, payments.
        with self.assertNumQueries(6):
            self.client.get(url)

    def test_list_query_count_does_not_grow_with_rows(self):
        """seats_used is annotated, not counted per row."""
        url = reverse("subscription-list")
        self.client.get(url)  # warm
        with self.assertNumQueries(3):
            self.client.get(url)

        for index in range(4):
            Subscription.objects.create(
                created_by=self.manager,
                **{
                    key: value
                    for key, value in subscription_payload(name=f"Extra {index}").items()
                    if key not in {"start_date", "expiry_date"}
                },
                start_date=timezone.localdate(),
                expiry_date=timezone.localdate() + timedelta(days=30),
            )
        # Same count with 5 rows as with 1.
        with self.assertNumQueries(3):
            self.client.get(url)
