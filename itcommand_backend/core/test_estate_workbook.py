"""The multi-tab estate workbook: one file, Accounts and Services tabs.

What is worth testing here is not "does openpyxl write a file". It is the two
promises the workbook makes that the single sheets could not:

* a service may name an account that does not exist yet, because the Accounts
  tab three columns to the left is about to create it; and
* if anything anywhere in the workbook is wrong, *nothing* is written — not the
  accounts, not the providers they would have created.

Every test below drives the real endpoints with a real .xlsx built by the real
template builder, because the failure this replaces was a file that validated
in Python and then would not open in Excel.
"""
from datetime import date
from io import BytesIO

from django.contrib.auth import get_user_model
from django.urls import reverse
from openpyxl import Workbook, load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from core import estate_import
from core.models import (
    AccountUser, Property, Provider, ProviderAccount, Server, Service,
)

User = get_user_model()


def _sheet(rows_by_tab):
    """A workbook holding only headers and the given rows, per tab.

    Built from the specs rather than hardcoded headers, so a column renamed in
    the spec fails these tests loudly instead of silently importing nothing.
    """
    specs = estate_import.build_workbook_specs()
    wb = Workbook()
    wb.remove(wb.active)
    for title, key in estate_import.WORKBOOK_TABS:
        if title not in rows_by_tab:
            continue
        spec = specs[key]
        ws = wb.create_sheet(title)
        ws.append([c.name for c in spec.columns])
        ws.append([c.help for c in spec.columns])       # guidance row
        ws.append([c.example for c in spec.columns])    # example row
        for row in rows_by_tab[title]:
            ws.append([row.get(c.name, "") for c in spec.columns])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = "workbook.xlsx"
    return buf


class EstateWorkbookTemplateTests(APITestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            email="wb-admin@example.com",
            password="pw", role="ADMIN",
        )
        self.client.force_authenticate(self.admin)

    def test_the_template_has_a_tab_per_record_type(self):
        response = self.client.get(
            reverse("estate_import_template"), {"resource": "workbook"}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        wb = load_workbook(BytesIO(response.content))
        self.assertEqual(
            wb.sheetnames,
            ["Read me", "Properties", "Accounts", "People", "Services", "Servers", "Lists"]
        )

    def test_choice_columns_get_a_dropdown_that_refuses_anything_else(self):
        """A closed list is entered by picking, not by remembering the code."""
        wb = load_workbook(BytesIO(estate_import.build_workbook_template()))
        rules = {
            str(dv.sqref): dv for dv in wb["Accounts"].data_validations.dataValidation
        }
        # Column D is MFA type on the Accounts tab.
        mfa = rules["D4:D400"]
        self.assertEqual(mfa.type, "list")
        self.assertTrue(mfa.formula1.startswith("Lists!"))
        self.assertTrue(mfa.showErrorMessage, "a closed list must reject typos")

    def test_a_column_that_may_name_something_new_suggests_without_refusing(self):
        """The provider dropdown must not block the first provider you ever add."""
        Provider.objects.create(name="Cloudflare", slug="cloudflare")
        wb = load_workbook(BytesIO(estate_import.build_workbook_template()))
        rules = {
            str(dv.sqref): dv for dv in wb["Accounts"].data_validations.dataValidation
        }
        self.assertFalse(rules["A4:A400"].showErrorMessage)

    def test_the_services_tab_reads_its_dropdowns_from_the_tabs_before_it(self):
        """Type a login on Accounts, pick it on Services — no upload in between."""
        wb = load_workbook(BytesIO(estate_import.build_workbook_template()))
        rules = {
            str(dv.sqref): dv for dv in wb["Services"].data_validations.dataValidation
        }
        self.assertEqual(rules["A4:A400"].formula1, "Accounts!$A$4:$A$400")
        self.assertEqual(rules["B4:B400"].formula1, "Accounts!$B$4:$B$400")
        self.assertEqual(rules["E4:E400"].formula1, "Properties!$A$4:$A$400")

    def test_date_cells_are_formatted_and_prompt_rather_than_reject(self):
        """A hard date rule would reject the text `as_date` reads perfectly well."""
        wb = load_workbook(BytesIO(estate_import.build_workbook_template()))
        ws = wb["Services"]
        self.assertEqual(ws["G4"].number_format, "yyyy-mm-dd")
        rule = {str(dv.sqref): dv for dv in ws.data_validations.dataValidation}["G4:G400"]
        self.assertEqual(rule.type, "date")
        self.assertFalse(rule.showErrorMessage)

    def test_columns_sharing_a_name_get_their_own_list(self):
        """"Role" is an account role on People and a server role on Servers;
        "Status" is a service status on Services and a server status on
        Servers. Keying the lists by column name pointed the second of each
        pair at the first one's values — strictly, so Excel then refused the
        very codes the sheet was asking for."""
        wb = load_workbook(BytesIO(estate_import.build_workbook_template()))
        lists = wb["Lists"]

        def values_behind(tab, column_letter):
            rule = {
                str(dv.sqref): dv
                for dv in wb[tab].data_validations.dataValidation
            }[f"{column_letter}4:{column_letter}400"]
            letter = rule.formula1.split("$")[1]
            index = ord(letter) - ord("A") + 1
            return [
                lists.cell(row=row, column=index).value
                for row in range(2, 8)
                if lists.cell(row=row, column=index).value
            ]

        self.assertIn("OWNER", values_behind("People", "G"))
        self.assertIn("WEB", values_behind("Servers", "D"))
        self.assertIn("RUNNING", values_behind("Servers", "F"))
        self.assertIn("ACTIVE", values_behind("Services", "F"))

    def test_identical_lists_are_still_stored_once(self):
        """Accounts' "MFA type" and People's "Second factor" hold the same
        codes and have no business being two columns."""
        wb = load_workbook(BytesIO(estate_import.build_workbook_template()))

        def source(tab, column_letter):
            rule = {
                str(dv.sqref): dv
                for dv in wb[tab].data_validations.dataValidation
            }[f"{column_letter}4:{column_letter}400"]
            return rule.formula1.split("$")[1]

        self.assertEqual(source("Accounts", "D"), source("People", "H"))

    def test_an_empty_list_produces_no_dropdown_at_all(self):
        """An empty dropdown would block the column it was meant to help with."""
        Provider.objects.all().delete()
        wb = load_workbook(BytesIO(estate_import.build_workbook_template()))
        rules = {
            str(dv.sqref): dv for dv in wb["Accounts"].data_validations.dataValidation
        }
        self.assertNotIn("A4:A400", rules)


class EstateWorkbookImportTests(APITestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            email="wb-imp@example.com",
            password="pw", role="ADMIN",
        )
        self.client.force_authenticate(self.admin)
        self.validate = reverse("estate_import_validate")
        self.commit = reverse("estate_import_commit")

    def _post(self, url, rows):
        return self.client.post(
            url, {"resource": "workbook", "file": _sheet(rows)}, format="multipart"
        )

    def test_a_service_may_name_an_account_the_accounts_tab_has_not_written_yet(self):
        """The whole point of one file: no import between the two tabs."""
        rows = {
            "Accounts": [{
                "Provider": "Cloudflare", "Account email": "ops@example.com",
                "MFA type": "APP",
            }],
            "Services": [{
                "Provider": "Cloudflare", "Account email": "ops@example.com",
                "Service type": "DNS", "Identifier": "terafort.com DNS",
                "Property": "terafort.com", "Cost": "1200", "Currency": "PKR",
                "Billing cycle": "YEARLY",
            }],
        }
        report = self._post(self.validate, rows).json()
        self.assertTrue(report["can_commit"], report)

        response = self._post(self.commit, rows)
        self.assertEqual(response.status_code, status.HTTP_200_OK, response.data)

        account = ProviderAccount.objects.get(account_email="ops@example.com")
        self.assertEqual(account.provider.name, "Cloudflare")
        service = Service.objects.get(identifier="terafort.com DNS")
        self.assertEqual(service.provider_account, account)
        # The provider on the service is the account's, never a second one.
        self.assertEqual(service.provider, account.provider)
        self.assertEqual(service.property.name, "terafort.com")

    def test_the_report_names_every_record_the_workbook_will_bring_into_existence(self):
        """A misspelled provider makes a second provider. Say so before writing."""
        report = self._post(self.validate, {
            "Accounts": [{"Provider": "Vercel", "Account email": "a@example.com"}],
            "Services": [{
                "Provider": "Vercel", "Account email": "a@example.com",
                "Service type": "HOSTING", "Identifier": "app hosting",
                "Property": "brand-new.com",
            }],
        }).json()
        self.assertIn("Provider “Vercel”", report["will_create"])
        self.assertIn("Property “brand-new.com”", report["will_create"])

    def test_twenty_services_on_one_new_provider_plan_one_provider(self):
        rows = {
            "Accounts": [{"Provider": "Fastly", "Account email": "n@example.com"}],
            "Services": [
                {
                    "Provider": "Fastly", "Account email": "n@example.com",
                    "Service type": "CDN", "Identifier": f"edge-{i}",
                }
                for i in range(20)
            ],
        }
        report = self._post(self.validate, rows).json()
        self.assertEqual(
            [w for w in report["will_create"] if w.startswith("Provider")],
            ["Provider “Fastly”"],
        )

    def test_a_service_naming_an_account_on_no_tab_is_an_error_not_a_silent_create(self):
        """This tab has no owner or MFA column. Inventing a login here would
        record a credential holder nobody chose."""
        report = self._post(self.validate, {
            "Accounts": [{"Provider": "Cloudflare", "Account email": "ops@example.com"}],
            "Services": [{
                "Provider": "Cloudflare", "Account email": "typo@example.com",
                "Service type": "DNS", "Identifier": "x",
            }],
        }).json()
        self.assertFalse(report["can_commit"])
        errors = report["tabs"][-1]["rows"][0]["errors"]
        self.assertTrue(
            any("Add it on the Accounts tab" in e for e in errors), errors
        )

    def test_one_bad_row_on_the_last_tab_leaves_the_first_tab_unwritten(self):
        """Half an import is worse than none — somebody has to reconcile it by hand."""
        response = self._post(self.commit, {
            "Accounts": [{"Provider": "Netlify", "Account email": "ops@example.com"}],
            "Services": [{
                "Provider": "Netlify", "Account email": "ops@example.com",
                "Service type": "NOT_A_TYPE", "Identifier": "x",
            }],
        })
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(Provider.objects.filter(name="Netlify").exists())
        self.assertFalse(ProviderAccount.objects.exists())

    def test_re_importing_the_same_workbook_updates_rather_than_duplicates(self):
        rows = {
            "Accounts": [{
                "Provider": "Cloudflare", "Account email": "ops@example.com",
                "MFA type": "SMS",
            }],
            "Services": [{
                "Provider": "Cloudflare", "Account email": "ops@example.com",
                "Service type": "DNS", "Identifier": "terafort.com DNS",
                "Cost": "1200", "Currency": "PKR", "Billing cycle": "YEARLY",
            }],
        }
        self.assertEqual(self._post(self.commit, rows).status_code, 200)
        rows["Accounts"][0]["MFA type"] = "SECURITY_KEY"
        rows["Services"][0]["Cost"] = "1500"
        response = self._post(self.commit, rows)
        self.assertEqual(response.status_code, status.HTTP_200_OK, response.data)

        self.assertEqual(Provider.objects.filter(name="Cloudflare").count(), 1)
        self.assertEqual(ProviderAccount.objects.count(), 1)
        self.assertEqual(Service.objects.count(), 1)
        self.assertEqual(
            ProviderAccount.objects.get().mfa_type, "SECURITY_KEY"
        )
        self.assertEqual(str(Service.objects.get().cost), "1500.00")

    def test_two_rows_describing_the_same_account_are_rejected(self):
        report = self._post(self.validate, {
            "Accounts": [
                {"Provider": "Cloudflare", "Account email": "ops@example.com"},
                {"Provider": "cloudflare", "Account email": "OPS@example.com"},
            ],
        }).json()
        self.assertFalse(report["can_commit"])
        self.assertTrue(
            any("same record" in e for e in report["tabs"][0]["rows"][1]["errors"])
        )

    def test_the_services_tab_alone_works_against_accounts_that_already_exist(self):
        """Not every upload is a fresh estate. One tab must be enough."""
        provider = Provider.objects.create(name="Cloudflare", slug="cloudflare")
        ProviderAccount.objects.create(
            provider=provider, account_email="ops@example.com"
        )
        response = self._post(self.commit, {
            "Services": [{
                "Provider": "Cloudflare", "Account email": "ops@example.com",
                "Service type": "DNS", "Identifier": "terafort.com DNS",
            }],
        })
        self.assertEqual(response.status_code, status.HTTP_200_OK, response.data)
        self.assertEqual(Service.objects.count(), 1)

    def test_the_properties_tab_sets_the_kind_the_services_tab_would_have_guessed(self):
        rows = {
            "Properties": [{"Name": "game.example", "Kind": "MOBILE_GAME"}],
            "Accounts": [{"Provider": "Cloudflare", "Account email": "ops@example.com"}],
            "Services": [{
                "Provider": "Cloudflare", "Account email": "ops@example.com",
                "Service type": "DNS", "Identifier": "game dns",
                "Property": "game.example",
            }],
        }
        # Named on the Properties tab, so the Services tab must not claim it.
        report = self._post(self.validate, rows).json()
        self.assertNotIn("Property “game.example”", report["will_create"])

        self.assertEqual(self._post(self.commit, rows).status_code, 200)
        self.assertEqual(Property.objects.get(name="game.example").kind, "MOBILE_GAME")

    def test_an_empty_workbook_says_so_instead_of_reporting_success(self):
        report = self._post(self.validate, {"Accounts": []}).json()
        self.assertFalse(report["can_commit"])
        self.assertTrue(report["sheet_errors"])

    def test_a_workbook_missing_a_column_is_refused_with_the_column_named(self):
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("Accounts")
        ws.append(["Provider"])  # everything else missing
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "broken.xlsx"

        response = self.client.post(
            self.validate, {"resource": "workbook", "file": buf}, format="multipart"
        )
        report = response.json()
        self.assertFalse(report["can_commit"])
        self.assertTrue(
            any("Account email" in e for e in report["sheet_errors"]),
            report["sheet_errors"],
        )

    def test_the_downloaded_template_itself_fills_in_and_imports(self):
        """The end-to-end path, on the real file rather than a stand-in.

        This is the test that would have caught a template whose guidance rows
        were read as data, or whose Read me and Lists tabs confused the reader.
        """
        buf = BytesIO(estate_import.build_workbook_template())
        wb = load_workbook(buf)

        accounts = wb["Accounts"]
        accounts["A4"] = "Cloudflare"
        accounts["B4"] = "ops@example.com"
        accounts["D4"] = "APP"

        services = wb["Services"]
        services["A4"] = "Cloudflare"
        services["B4"] = "ops@example.com"
        services["C4"] = "DNS"
        services["D4"] = "terafort.com DNS"
        services["E4"] = "terafort.com"
        services["G4"] = date(2027, 1, 31)   # a real date cell, not text
        services["I4"] = 1200
        services["J4"] = "PKR"
        services["K4"] = "YEARLY"

        filled = BytesIO()
        wb.save(filled)
        filled.seek(0)
        filled.name = "filled.xlsx"

        response = self.client.post(
            self.commit, {"resource": "workbook", "file": filled}, format="multipart"
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK, response.data)

        service = Service.objects.get(identifier="terafort.com DNS")
        self.assertEqual(service.renewal_date, date(2027, 1, 31))
        self.assertEqual(str(service.cost), "1200.00")
        self.assertEqual(service.provider_account.mfa_type, "APP")
        # The example row shipped in the template must not have been imported.
        self.assertEqual(Service.objects.count(), 1)
        self.assertEqual(ProviderAccount.objects.count(), 1)

    def test_people_and_servers_import_alongside_their_account(self):
        """One upload: the account, the people in it, and the machines it pays
        for — with none of them existing beforehand."""
        rows = {
            "Accounts": [{"Provider": "AWS", "Account email": "1234-5678"}],
            "People": [
                {"Provider": "AWS", "Account email": "1234-5678",
                 "Login": "root@example.com", "Login is a": "EMAIL",
                 "Role": "OWNER", "Second factor": "SECURITY_KEY"},
                {"Provider": "AWS", "Account email": "1234-5678",
                 "Login": "iam:bob", "Login is a": "USERNAME",
                 "Role": "MEMBER", "Second factor": "NONE"},
            ],
            "Servers": [{
                "Provider": "AWS", "Account email": "1234-5678",
                "Server name": "web-01", "Role": "WEB",
                "Environment": "PRODUCTION", "Status": "RUNNING",
                "Public IP": "203.0.113.9", "Region": "eu-west-1",
                "Property": "terafort.com",
            }],
        }
        report = self._post(self.validate, rows).json()
        self.assertTrue(report["can_commit"], report)

        response = self._post(self.commit, rows)
        self.assertEqual(response.status_code, status.HTTP_200_OK, response.data)

        account = ProviderAccount.objects.get(account_email="1234-5678")
        self.assertEqual(account.people.count(), 2)
        # The rollup the whole model exists for: the account inherits its
        # weakest login without anybody setting it.
        self.assertEqual(account.effective_mfa_type, "NONE")

        server = Server.objects.get(name="web-01")
        self.assertEqual(server.provider_account, account)
        self.assertEqual(server.public_ip, "203.0.113.9")
        self.assertEqual(server.property.name, "terafort.com")

    def test_a_person_on_an_account_that_is_on_no_tab_is_an_error(self):
        report = self._post(self.validate, {
            "Accounts": [{"Provider": "AWS", "Account email": "1234-5678"}],
            "People": [{"Provider": "AWS", "Account email": "typo",
                        "Login": "iam:bob"}],
        }).json()
        self.assertFalse(report["can_commit"])
        errors = report["tabs"][-1]["rows"][0]["errors"]
        self.assertTrue(any("Add it on the Accounts tab" in e for e in errors), errors)

    def test_a_bad_ip_names_the_row_it_came_from(self):
        """Caught in validation, not by the model at commit time, so the report
        can say which row rather than failing the whole upload namelessly."""
        report = self._post(self.validate, {
            "Accounts": [{"Provider": "AWS", "Account email": "a@example.com"}],
            "Servers": [{"Provider": "AWS", "Account email": "a@example.com",
                         "Server name": "web-01", "Public IP": "203.0.113.999"}],
        }).json()
        self.assertFalse(report["can_commit"])
        errors = report["tabs"][-1]["rows"][0]["errors"]
        self.assertTrue(any("not an IP address" in e for e in errors), errors)

    def test_re_importing_people_updates_rather_than_duplicating(self):
        rows = {
            "Accounts": [{"Provider": "AWS", "Account email": "a@example.com"}],
            "People": [{"Provider": "AWS", "Account email": "a@example.com",
                        "Login": "iam:bob", "Second factor": "NONE"}],
        }
        self.assertEqual(self._post(self.commit, rows).status_code, 200)
        rows["People"][0]["Second factor"] = "APP"
        self.assertEqual(self._post(self.commit, rows).status_code, 200)

        self.assertEqual(AccountUser.objects.count(), 1)
        self.assertEqual(AccountUser.objects.get().mfa_type, "APP")

    def test_a_non_admin_cannot_import(self):
        staff = User.objects.create_user(
            email="wb-staff@example.com",
            password="pw", role="STAFF",
        )
        self.client.force_authenticate(staff)
        response = self._post(self.commit, {
            "Accounts": [{"Provider": "X", "Account email": "a@example.com"}],
        })
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertFalse(Provider.objects.filter(name="X").exists())
