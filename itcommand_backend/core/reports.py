import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from django.db.models import Sum, Count, Q, Avg, F, DurationField, ExpressionWrapper
from django.utils import timezone
from datetime import timedelta, datetime
from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import HttpResponse
from .models import (
    User, Department, Asset, Expense, Income, FinancialYear, Budget, BudgetCategory, PettyCashTransaction,
    DirectPayment, RecurringBill, AssetHistory,
    Ticket, TicketCategory,
    SoftwareLicense, SoftwareProduct,
    PurchaseRequest,
    Vendor, VendorContract, VendorPayment,
    Office, Floor, Seat, SeatAssignment,
    NetworkDevice, NetworkLocation, IPAddressPool,
    OnboardingRecord, OnboardingTask,
    KBArticle, KBCategory, KBFeedback,
)
from .permissions import IsAdminOrSuperadmin


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill('solid', fgColor='4F46E5')
_HEADER_FONT = Font(bold=True, color='FFFFFF')


def _last_n_months(n):
    """Return a list of (label, start_date, end_date) for the last n months,
    oldest first. Month boundaries are exact (first-of-month to first-of-next)."""
    today = timezone.now().date()
    months = []
    cursor = today.replace(day=1)
    starts = []
    for _ in range(n):
        starts.append(cursor)
        # step back one month
        if cursor.month == 1:
            cursor = cursor.replace(year=cursor.year - 1, month=12)
        else:
            cursor = cursor.replace(month=cursor.month - 1)
    for start in reversed(starts):
        if start.month == 12:
            end = start.replace(year=start.year + 1, month=1)
        else:
            end = start.replace(month=start.month + 1)
        months.append((start.strftime('%b %Y'), start, end))
    return months


def _parse_date(value):
    """Parse a 'YYYY-MM-DD' string to a date, or return None."""
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


def _resolve_period(request):
    """Resolve the reporting period from ?start_date & ?end_date query params.

    Falls back to the active financial year, then to the last 12 months.
    Returns (start_date, end_date) with start <= end.
    """
    today = timezone.now().date()
    start = _parse_date(request.query_params.get('start_date'))
    end = _parse_date(request.query_params.get('end_date'))
    if not start or not end:
        active_fy = FinancialYear.objects.filter(is_active=True).first()
        if active_fy:
            start = start or active_fy.start_date
            end = end or active_fy.end_date
        else:
            start = start or today.replace(year=today.year - 1)
            end = end or today
    if start > end:
        start, end = end, start
    return start, end


def _months_between(start, end):
    """Yield (label, seg_start, seg_end) monthly buckets across [start, end],
    where each segment is a half-open [seg_start, seg_end) interval clipped to
    the range. Oldest first."""
    buckets = []
    cursor = start.replace(day=1)
    end_exclusive = end + timedelta(days=1)
    while cursor <= end:
        nxt = cursor.replace(year=cursor.year + 1, month=1) if cursor.month == 12 \
            else cursor.replace(month=cursor.month + 1)
        seg_start = max(cursor, start)
        seg_end = min(nxt, end_exclusive)
        buckets.append((cursor.strftime('%b %Y'), seg_start, seg_end))
        cursor = nxt
    return buckets


def make_xlsx(filename, sheets):
    """Build a styled multi-sheet workbook.

    `sheets` is a list of (sheet_title, headers, rows) tuples where rows is a
    list of lists. Header row is bold/filled, columns are auto-sized.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for title, headers, rows in sheets:
        ws = wb.create_sheet(title=str(title)[:31])
        ws.append(list(headers))
        for cell in ws[1]:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(vertical='center')
        for row in rows:
            ws.append(list(row))
        ws.freeze_panes = 'A2'
        for i in range(1, len(headers) + 1):
            col = get_column_letter(i)
            longest = len(str(headers[i - 1]))
            for row in rows:
                if i - 1 < len(row) and row[i - 1] is not None:
                    longest = max(longest, len(str(row[i - 1])))
            ws.column_dimensions[col].width = min(longest + 4, 55)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

class FinancialSummaryView(APIView):
    permission_classes = [IsAdminOrSuperadmin]

    def get(self, request):
        start, end = _resolve_period(request)
        active_fy = FinancialYear.objects.filter(is_active=True).first()

        # Expenses falling inside the selected period.
        period_expenses = Expense.objects.filter(expense_date__gte=start, expense_date__lte=end)

        # Budget utilization per category — allocation is the full-year figure,
        # "spent" reflects the selected period.
        budget_utilization = []
        if active_fy:
            for b in Budget.objects.filter(financial_year=active_fy).select_related('category'):
                spent = period_expenses.filter(category=b.category).aggregate(s=Sum('amount'))['s'] or 0
                allocated = b.allocated_amount
                budget_utilization.append({
                    'category': b.category.name,
                    'allocated': float(allocated),
                    'spent': float(spent),
                    'remaining': float(allocated - spent),
                    'percentage': float((spent / allocated * 100) if allocated > 0 else 0)
                })

        # Monthly expense breakdown across the selected range.
        monthly_expenses = []
        for label, seg_start, seg_end in _months_between(start, end):
            amount = Expense.objects.filter(
                expense_date__gte=seg_start, expense_date__lt=seg_end
            ).aggregate(s=Sum('amount'))['s'] or 0
            monthly_expenses.append({'month': label, 'amount': float(amount)})

        # Top 5 expense categories (in period)
        top_categories = period_expenses.values('category__name').annotate(
            total=Sum('amount')).order_by('-total')[:5]
        top_5_categories = [{'name': c['category__name'] or 'Uncategorized', 'value': float(c['total'] or 0)}
                            for c in top_categories]

        # Petty cash (in period)
        pc = PettyCashTransaction.objects.filter(date__gte=start, date__lte=end)
        pc_topups = pc.filter(transaction_type='TOPUP').aggregate(s=Sum('amount'))['s'] or 0
        pc_expenses = pc.filter(transaction_type='EXPENSE').aggregate(s=Sum('amount'))['s'] or 0
        petty_cash_summary = {
            'balance': float(pc_topups - pc_expenses),
            'total_in': float(pc_topups),
            'total_out': float(pc_expenses),
        }

        # Direct payments (in period)
        direct_payments_total = DirectPayment.objects.filter(
            payment_date__gte=start, payment_date__lte=end
        ).aggregate(s=Sum('amount'))['s'] or 0

        # Recurring monthly commitment (current, not period-scoped)
        monthly_commitment = RecurringBill.objects.filter(is_active=True).aggregate(s=Sum('amount'))['s'] or 0

        # ---- Ledger: chronological money in/out with a running balance ----
        ledger, ledger_summary = self._build_ledger(start, end)

        return Response({
            'period': {'start': start.strftime('%Y-%m-%d'), 'end': end.strftime('%Y-%m-%d')},
            'budget_utilization': budget_utilization,
            'monthly_expenses': monthly_expenses,
            'top_categories': top_5_categories,
            'petty_cash': petty_cash_summary,
            'direct_payments_total': float(direct_payments_total),
            'monthly_commitment': float(monthly_commitment),
            'ledger': ledger,
            'ledger_summary': ledger_summary,
        })

    @staticmethod
    def _build_ledger(start, end):
        """Combined cash ledger: Income (credit) vs Expense + Direct Payment
        (debit), ordered by date with an opening/running/closing balance."""
        opening = (
            (Income.objects.filter(income_date__lt=start).aggregate(s=Sum('amount'))['s'] or 0)
            - (Expense.objects.filter(expense_date__lt=start).aggregate(s=Sum('amount'))['s'] or 0)
            - (DirectPayment.objects.filter(payment_date__lt=start).aggregate(s=Sum('amount'))['s'] or 0)
        )

        rows = []
        for inc in Income.objects.filter(income_date__gte=start, income_date__lte=end).select_related('source'):
            rows.append((inc.income_date, 'INCOME', inc.title,
                         inc.source.name if inc.source else '', inc.payment_method,
                         0.0, float(inc.amount)))
        for exp in Expense.objects.filter(expense_date__gte=start, expense_date__lte=end).select_related('category'):
            rows.append((exp.expense_date, 'EXPENSE', exp.title,
                         exp.paid_to, exp.payment_method,
                         float(exp.amount), 0.0))
        for dp in DirectPayment.objects.filter(payment_date__gte=start, payment_date__lte=end):
            rows.append((dp.payment_date, 'DIRECT PAYMENT', dp.title,
                         dp.paid_to, dp.payment_method,
                         float(dp.amount), 0.0))

        # Credits before debits on the same day so the balance reads naturally.
        rows.sort(key=lambda r: (r[0], r[5] > 0))

        running = float(opening)
        total_in = total_out = 0.0
        ledger = []
        for d, kind, desc, party, method, debit, credit in rows:
            running += credit - debit
            total_in += credit
            total_out += debit
            ledger.append({
                'date': d.strftime('%Y-%m-%d'),
                'type': kind,
                'description': desc,
                'party': party,
                'method': method,
                'debit': debit,
                'credit': credit,
                'balance': round(running, 2),
            })

        summary = {
            'opening_balance': round(float(opening), 2),
            'closing_balance': round(running, 2),
            'total_in': round(total_in, 2),
            'total_out': round(total_out, 2),
            'net': round(total_in - total_out, 2),
            'count': len(ledger),
        }
        return ledger, summary

class AssetSummaryView(APIView):
    permission_classes = [IsAdminOrSuperadmin]

    def get(self, request):
        today = timezone.now().date()
        
        # Assets by status
        statuses = Asset.objects.values('status').annotate(count=Count('id'))
        assets_by_status = [{'name': s['status'], 'value': s['count']} for s in statuses]

        # Assets by category
        categories = Asset.objects.values('category__name').annotate(count=Count('id'))
        assets_by_category = [{'name': c['category__name'], 'count': c['count']} for c in categories]

        # Recently assigned (last 30 days)
        thirty_days_ago = today - timedelta(days=30)
        recent_assignments = AssetHistory.objects.filter(action='ASSIGNED', timestamp__date__gte=thirty_days_ago).order_by('-timestamp')[:10]
        recently_assigned = [{
            'asset': a.asset.name,
            'tag': a.asset.asset_tag,
            'to_user': a.to_user.get_full_name() if a.to_user else 'Unknown',
            'date': a.timestamp.strftime('%Y-%m-%d')
        } for a in recent_assignments]

        # Expiring warranty (next 60 days)
        sixty_days = today + timedelta(days=60)
        expiring = Asset.objects.filter(warranty_expiry__gte=today, warranty_expiry__lte=sixty_days).order_by('warranty_expiry')
        expiring_warranties = [{
            'name': a.name,
            'tag': a.asset_tag,
            'warranty_date': a.warranty_expiry.strftime('%Y-%m-%d'),
            'days_remaining': (a.warranty_expiry - today).days
        } for a in expiring]

        # Total asset value by category
        value_by_category = Asset.objects.values('category__name').annotate(total=Sum('purchase_price'))
        total_value_by_category = [{'category': v['category__name'], 'value': float(v['total'] or 0)} for v in value_by_category]

        total_assets = Asset.objects.count()
        total_value = Asset.objects.aggregate(Sum('purchase_price'))['purchase_price__sum'] or 0
        assigned_count = Asset.objects.filter(status='ASSIGNED').count()
        assigned_percentage = (assigned_count / total_assets * 100) if total_assets > 0 else 0

        return Response({
            'total_assets': total_assets,
            'total_value': float(total_value),
            'assigned_percentage': float(assigned_percentage),
            'assets_by_status': assets_by_status,
            'assets_by_category': assets_by_category,
            'recently_assigned': recently_assigned,
            'expiring_warranties': expiring_warranties,
            'total_value_by_category': total_value_by_category
        })

class ExportFinancialView(APIView):
    permission_classes = [IsAdminOrSuperadmin]

    def get(self, request):
        start, end = _resolve_period(request)

        exp_headers = ['Date', 'Title', 'Category', 'Financial Year', 'Payment Method',
                       'Paid To', 'Receipt Number', 'Amount']
        exp_rows = [[
            e.expense_date.strftime('%Y-%m-%d') if e.expense_date else '',
            e.title,
            e.category.name if e.category else '',
            e.financial_year.name if e.financial_year else '',
            e.payment_method, e.paid_to, e.receipt_number or '', float(e.amount),
        ] for e in Expense.objects.filter(
            expense_date__gte=start, expense_date__lte=end
        ).select_related('category', 'financial_year').order_by('-expense_date')]

        # Ledger sheet — chronological with running balance.
        ledger, summary = FinancialSummaryView._build_ledger(start, end)
        led_headers = ['Date', 'Type', 'Description', 'Party', 'Method', 'Debit', 'Credit', 'Balance']
        led_rows = [['', '', 'Opening Balance', '', '', '', '', summary['opening_balance']]]
        led_rows += [[
            r['date'], r['type'], r['description'], r['party'], r['method'],
            r['debit'] or '', r['credit'] or '', r['balance'],
        ] for r in ledger]
        led_rows.append(['', '', 'Closing Balance', '', '',
                         summary['total_out'], summary['total_in'], summary['closing_balance']])

        return make_xlsx(f'financial_export_{start}_{end}.xlsx', [
            ('Expenses', exp_headers, exp_rows),
            ('Ledger', led_headers, led_rows),
        ])

class ExportAssetsView(APIView):
    permission_classes = [IsAdminOrSuperadmin]

    def get(self, request):
        assets = Asset.objects.all().order_by('asset_tag')
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Assets"

        headers = ['Asset Tag', 'Name', 'Category', 'Type', 'Brand', 'Model', 'Status', 'Condition', 'Purchase Price', 'Warranty Expiry', 'Assigned To', 'Location']
        ws.append(headers)

        for a in assets:
            ws.append([
                a.asset_tag,
                a.name,
                a.category.name if a.category else '',
                a.asset_type,
                a.brand,
                a.model,
                a.status,
                a.condition,
                float(a.purchase_price) if a.purchase_price else 0,
                a.warranty_expiry.strftime('%Y-%m-%d') if a.warranty_expiry else '',
                a.assigned_to.email if a.assigned_to else '',
                a.location
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=assets_export.xlsx'
        wb.save(response)
        return response

class MainDashboardView(APIView):
    permission_classes = [IsAdminOrSuperadmin]

    def get(self, request):
        now = timezone.now()
        today = now.date()
        soon_30 = today + timedelta(days=30)
        soon_60 = today + timedelta(days=60)
        soon_90 = today + timedelta(days=90)
        seven_days = today + timedelta(days=7)

        # ---- People & Assets ----
        total_users = User.objects.count()
        active_users = User.objects.filter(is_active=True).count()
        total_assets = Asset.objects.count()
        asset_value = Asset.objects.aggregate(s=Sum('purchase_price'))['s'] or 0
        assets_assigned = Asset.objects.filter(status='ASSIGNED').count()
        assets_by_status = [{'name': s['status'], 'value': s['count']}
                            for s in Asset.objects.values('status').annotate(count=Count('id'))]
        warranties_expiring = [{
            'name': a.name, 'tag': a.asset_tag,
            'date': a.warranty_expiry.strftime('%Y-%m-%d'),
            'days': (a.warranty_expiry - today).days,
        } for a in Asset.objects.filter(warranty_expiry__gte=today, warranty_expiry__lte=soon_60).order_by('warranty_expiry')[:8]]

        # ---- Finance ----
        active_fy = FinancialYear.objects.filter(is_active=True).first()
        total_budget = total_spent = 0
        if active_fy:
            total_budget = Budget.objects.filter(financial_year=active_fy).aggregate(s=Sum('allocated_amount'))['s'] or 0
            total_spent = Expense.objects.filter(financial_year=active_fy).aggregate(s=Sum('amount'))['s'] or 0
        budget_used_pct = (total_spent / total_budget * 100) if total_budget else 0

        upcoming_bills = RecurringBill.objects.filter(is_active=True, next_due_date__gte=today, next_due_date__lte=seven_days)
        upcoming_bills_count = upcoming_bills.count()
        upcoming_bills_amount = upcoming_bills.aggregate(s=Sum('amount'))['s'] or 0

        # Income vs expense, last 6 months.
        income_vs_expense = []
        monthly_expenses = []
        for label, start, end in _last_n_months(6):
            inc = Income.objects.filter(income_date__gte=start, income_date__lt=end).aggregate(s=Sum('amount'))['s'] or 0
            exp = Expense.objects.filter(expense_date__gte=start, expense_date__lt=end).aggregate(s=Sum('amount'))['s'] or 0
            short = label.split(' ')[0]
            income_vs_expense.append({'month': short, 'income': float(inc), 'expense': float(exp)})
            monthly_expenses.append({'month': short, 'amount': float(exp)})

        # ---- Helpdesk ----
        open_statuses = ['OPEN', 'IN_PROGRESS', 'PENDING']
        closed_statuses = ['RESOLVED', 'CLOSED']
        tickets = Ticket.objects.all()
        helpdesk = {
            'total': tickets.count(),
            'open': tickets.filter(status__in=open_statuses).count(),
            'overdue': tickets.filter(due_date__lt=now).exclude(status__in=closed_statuses).count(),
            'unassigned': tickets.filter(assigned_to__isnull=True).exclude(status__in=closed_statuses).count(),
            'resolved': tickets.filter(status__in=closed_statuses).count(),
        }
        tickets_by_status = [{'name': s['status'], 'value': s['count']}
                             for s in tickets.values('status').annotate(count=Count('id'))]

        # ---- Licenses ----
        licenses = SoftwareLicense.objects.all()
        lic_expiring = licenses.filter(expiry_date__gte=today, expiry_date__lte=soon_60).count()
        lic_annual = sum(l.annual_cost for l in licenses.filter(is_active=True))
        license_block = {
            'total': licenses.count(),
            'active': licenses.filter(is_active=True).count(),
            'expiring_soon': lic_expiring,
            'annual_cost': round(lic_annual, 2),
        }

        # ---- Procurement ----
        prs = PurchaseRequest.objects.all()
        procurement = {
            'total': prs.count(),
            'pending': prs.filter(status__in=['SUBMITTED', 'UNDER_REVIEW']).count(),
            'approved': prs.filter(status='APPROVED').count(),
            'est_total': float(prs.exclude(status__in=['REJECTED', 'CANCELLED']).aggregate(s=Sum('total_estimated_cost'))['s'] or 0),
        }

        # ---- Vendors ----
        contracts_expiring = [{
            'vendor': c.vendor.name, 'title': c.title,
            'date': c.end_date.strftime('%Y-%m-%d'),
            'days': (c.end_date - today).days,
        } for c in VendorContract.objects.select_related('vendor').filter(
            end_date__gte=today, end_date__lte=soon_90
        ).exclude(status__in=['TERMINATED', 'EXPIRED']).order_by('end_date')[:8]]
        vendors_block = {
            'total': Vendor.objects.count(),
            'active': Vendor.objects.filter(is_active=True).count(),
            'contracts_expiring': len(contracts_expiring),
        }

        # ---- Network ----
        devices = NetworkDevice.objects.all()
        network = {
            'total': devices.count(),
            'online': devices.filter(status='ONLINE').count(),
            'offline': devices.filter(status='OFFLINE').count(),
            'warranty_expiring': devices.filter(warranty_expiry__gte=today, warranty_expiry__lte=soon_30).exclude(status='DECOMMISSIONED').count(),
        }

        # ---- Onboarding ----
        records = OnboardingRecord.objects.all()
        onboarding = {
            'in_progress': records.filter(status='IN_PROGRESS').count(),
            'not_started': records.filter(status='NOT_STARTED').count(),
            'overdue': records.filter(target_completion_date__lt=today).exclude(status__in=['COMPLETED', 'CANCELLED']).count(),
        }

        # ---- Seating ----
        total_seats = Seat.objects.count()
        occupied = Seat.objects.filter(assignments__is_active=True).distinct().count()
        seating = {
            'total': total_seats, 'occupied': occupied,
            'pct': round(occupied / total_seats * 100, 1) if total_seats else 0,
        }

        # ---- Knowledge Base ----
        kb = {
            'published': KBArticle.objects.filter(status='PUBLISHED').count(),
            'total': KBArticle.objects.count(),
            'views': KBArticle.objects.aggregate(s=Sum('view_count'))['s'] or 0,
        }

        # ---- Cross-module recent activity ----
        recent_activity = []
        for e in Expense.objects.select_related('category').order_by('-created_at')[:6]:
            recent_activity.append({'type': 'EXPENSE', 'title': f"Expense: {e.title}",
                                    'amount': float(e.amount), 'date': e.created_at})
        for a in AssetHistory.objects.filter(action='ASSIGNED').select_related('asset', 'to_user').order_by('-timestamp')[:5]:
            recent_activity.append({'type': 'ASSET',
                                    'title': f"Asset assigned: {a.asset.name} → {a.to_user.get_full_name() if a.to_user else 'Unknown'}",
                                    'amount': None, 'date': a.timestamp})
        for t in Ticket.objects.select_related('requester').order_by('-created_at')[:5]:
            recent_activity.append({'type': 'TICKET',
                                    'title': f"Ticket {t.ticket_number}: {t.title}",
                                    'amount': None, 'date': t.created_at})
        for pr in PurchaseRequest.objects.order_by('-created_at')[:4]:
            recent_activity.append({'type': 'PROCUREMENT',
                                    'title': f"PR {pr.pr_number}: {pr.title}",
                                    'amount': float(pr.total_estimated_cost), 'date': pr.created_at})
        recent_activity.sort(key=lambda x: x['date'], reverse=True)
        for r in recent_activity:
            r['date'] = r['date'].strftime('%Y-%m-%d %H:%M')

        return Response({
            'kpis': {
                'total_users': total_users, 'active_users': active_users,
                'total_assets': total_assets, 'asset_value': float(asset_value),
                'assets_assigned': assets_assigned,
                'budget_used_pct': float(budget_used_pct),
                'total_budget': float(total_budget), 'total_spent': float(total_spent),
                'open_tickets': helpdesk['open'], 'overdue_tickets': helpdesk['overdue'],
                'active_licenses': license_block['active'],
                'pending_pos': procurement['pending'],
                'total_vendors': vendors_block['total'],
                'devices_total': network['total'], 'devices_online': network['online'],
                'seat_occupancy_pct': seating['pct'],
                'upcoming_bills_count': upcoming_bills_count,
                'upcoming_bills_amount': float(upcoming_bills_amount),
            },
            'monthly_expenses': monthly_expenses,
            'income_vs_expense': income_vs_expense,
            'assets_by_status': assets_by_status,
            'tickets_by_status': tickets_by_status,
            'helpdesk': helpdesk,
            'licenses': license_block,
            'procurement': procurement,
            'vendors': vendors_block,
            'network': network,
            'onboarding': onboarding,
            'seating': seating,
            'kb': kb,
            'warranties_expiring': warranties_expiring,
            'contracts_expiring': contracts_expiring,
            'recent_activity': recent_activity[:12],
            # legacy keys kept for back-compat
            'total_users_legacy': total_users,
        })


# ===========================================================================
# HELPDESK REPORT
# ===========================================================================
class HelpdeskSummaryView(APIView):
    permission_classes = [IsAdminOrSuperadmin]

    def get(self, request):
        now = timezone.now()
        qs = Ticket.objects.all()
        total = qs.count()
        open_statuses = ['OPEN', 'IN_PROGRESS', 'PENDING']
        closed_statuses = ['RESOLVED', 'CLOSED']

        open_count = qs.filter(status__in=open_statuses).count()
        resolved_count = qs.filter(status__in=closed_statuses).count()
        unassigned = qs.filter(assigned_to__isnull=True).exclude(status__in=closed_statuses).count()

        # SLA + overdue need the model properties — iterate open tickets only.
        sla = {'ON_TRACK': 0, 'AT_RISK': 0, 'BREACHED': 0}
        overdue = 0
        for t in qs.filter(status__in=open_statuses):
            sla[t.sla_status] = sla.get(t.sla_status, 0) + 1
            if t.is_overdue:
                overdue += 1

        by_status = [{'name': s['status'], 'value': s['count']}
                     for s in qs.values('status').annotate(count=Count('id'))]
        by_priority = [{'name': p['priority'], 'value': p['count']}
                       for p in qs.values('priority').annotate(count=Count('id')).order_by('-count')]
        by_category = [{'name': c['category__name'] or 'Uncategorized', 'value': c['count']}
                       for c in qs.values('category__name').annotate(count=Count('id')).order_by('-count')]

        # Created vs resolved over last 6 months
        monthly = []
        for label, start, end in _last_n_months(6):
            created = qs.filter(created_at__date__gte=start, created_at__date__lt=end).count()
            resolved = qs.filter(resolved_at__date__gte=start, resolved_at__date__lt=end).count()
            monthly.append({'month': label, 'created': created, 'resolved': resolved})

        # Average resolution time (hours) for resolved tickets
        avg_resolution_hours = 0
        resolved_qs = qs.filter(resolved_at__isnull=False)
        durations = resolved_qs.annotate(
            dur=ExpressionWrapper(F('resolved_at') - F('created_at'), output_field=DurationField())
        ).aggregate(avg=Avg('dur'))['avg']
        if durations:
            avg_resolution_hours = round(durations.total_seconds() / 3600, 1)

        # Top agents by assigned tickets
        agents = qs.filter(assigned_to__isnull=False).values(
            'assigned_to__full_name'
        ).annotate(
            total=Count('id'),
            resolved=Count('id', filter=Q(status__in=closed_statuses)),
        ).order_by('-total')[:8]
        top_agents = [{
            'agent': a['assigned_to__full_name'] or 'Unknown',
            'total': a['total'],
            'resolved': a['resolved'],
        } for a in agents]

        return Response({
            'totals': {
                'total': total, 'open': open_count, 'resolved': resolved_count,
                'overdue': overdue, 'unassigned': unassigned,
                'avg_resolution_hours': avg_resolution_hours,
            },
            'sla': [{'name': k, 'value': v} for k, v in sla.items()],
            'by_status': by_status,
            'by_priority': by_priority,
            'by_category': by_category,
            'monthly': monthly,
            'top_agents': top_agents,
        })


# ===========================================================================
# LICENSE REPORT
# ===========================================================================
class LicenseSummaryView(APIView):
    permission_classes = [IsAdminOrSuperadmin]

    def get(self, request):
        today = timezone.now().date()
        qs = SoftwareLicense.objects.select_related('product').all()

        total = qs.count()
        active = qs.filter(is_active=True).count()
        expired = 0
        expiring_soon = 0
        total_annual_cost = 0.0
        soon_cutoff = today + timedelta(days=60)
        seat_rows = []
        expiring_rows = []

        for lic in qs:
            total_annual_cost += lic.annual_cost
            if lic.is_expired:
                expired += 1
            elif lic.expiry_date and today <= lic.expiry_date <= soon_cutoff:
                expiring_soon += 1
                expiring_rows.append({
                    'product': lic.product.name,
                    'type': lic.license_type,
                    'expiry_date': lic.expiry_date.strftime('%Y-%m-%d'),
                    'days_remaining': lic.days_until_expiry,
                    'cost': float(lic.cost),
                })
            if lic.seats_total:
                seat_rows.append({
                    'product': lic.product.name,
                    'seats_total': lic.seats_total,
                    'seats_used': lic.seats_used,
                    'usage_pct': lic.seats_usage_pct,
                })

        expiring_rows.sort(key=lambda r: r['days_remaining'])
        seat_rows.sort(key=lambda r: r['usage_pct'], reverse=True)

        by_type = [{'name': t['license_type'], 'value': t['count']}
                   for t in qs.values('license_type').annotate(count=Count('id')).order_by('-count')]
        by_category = [{'name': c['product__category'], 'value': c['count']}
                       for c in qs.values('product__category').annotate(count=Count('id')).order_by('-count')]
        cost_by_category = [{'name': c['product__category'], 'value': float(c['total'] or 0)}
                            for c in qs.values('product__category').annotate(total=Sum('cost')).order_by('-total')]

        return Response({
            'totals': {
                'total': total, 'active': active, 'expired': expired,
                'expiring_soon': expiring_soon,
                'total_annual_cost': round(total_annual_cost, 2),
            },
            'by_type': by_type,
            'by_category': by_category,
            'cost_by_category': cost_by_category,
            'expiring': expiring_rows,
            'seat_utilization': seat_rows[:15],
        })


# ===========================================================================
# PROCUREMENT REPORT
# ===========================================================================
class ProcurementSummaryView(APIView):
    permission_classes = [IsAdminOrSuperadmin]

    def get(self, request):
        qs = PurchaseRequest.objects.all()
        total = qs.count()
        pending = qs.filter(status__in=['SUBMITTED', 'UNDER_REVIEW']).count()
        approved = qs.filter(status='APPROVED').count()
        received = qs.filter(status='RECEIVED').count()
        est_total = qs.aggregate(s=Sum('total_estimated_cost'))['s'] or 0
        act_total = qs.aggregate(s=Sum('total_actual_cost'))['s'] or 0

        by_status = [{'name': s['status'], 'value': s['count']}
                     for s in qs.values('status').annotate(count=Count('id')).order_by('-count')]
        by_priority = [{'name': p['priority'], 'value': p['count']}
                       for p in qs.values('priority').annotate(count=Count('id')).order_by('-count')]
        by_department = [{'name': d['department__name'] or 'Unassigned',
                          'count': d['count'], 'value': float(d['total'] or 0)}
                         for d in qs.values('department__name').annotate(
                             count=Count('id'), total=Sum('total_estimated_cost')
                         ).order_by('-total')]

        monthly = []
        for label, start, end in _last_n_months(6):
            month_qs = qs.filter(created_at__date__gte=start, created_at__date__lt=end)
            monthly.append({
                'month': label,
                'count': month_qs.count(),
                'value': float(month_qs.aggregate(s=Sum('total_estimated_cost'))['s'] or 0),
            })

        # Average approval time (days)
        avg_approval_days = 0
        approved_qs = qs.filter(approved_at__isnull=False)
        dur = approved_qs.annotate(
            d=ExpressionWrapper(F('approved_at') - F('created_at'), output_field=DurationField())
        ).aggregate(avg=Avg('d'))['avg']
        if dur:
            avg_approval_days = round(dur.total_seconds() / 86400, 1)

        pending_list = [{
            'pr_number': pr.pr_number,
            'title': pr.title,
            'requested_by': pr.requested_by.full_name if pr.requested_by else '',
            'department': pr.department.name if pr.department else '',
            'status': pr.status,
            'priority': pr.priority,
            'estimated': float(pr.total_estimated_cost),
        } for pr in qs.filter(status__in=['SUBMITTED', 'UNDER_REVIEW']).select_related(
            'requested_by', 'department').order_by('-created_at')[:15]]

        requesters = qs.values('requested_by__full_name').annotate(
            count=Count('id'), total=Sum('total_estimated_cost')
        ).order_by('-count')[:8]
        top_requesters = [{
            'name': r['requested_by__full_name'] or 'Unknown',
            'count': r['count'], 'value': float(r['total'] or 0),
        } for r in requesters]

        return Response({
            'totals': {
                'total': total, 'pending': pending, 'approved': approved,
                'received': received,
                'estimated_total': float(est_total), 'actual_total': float(act_total),
                'avg_approval_days': avg_approval_days,
            },
            'by_status': by_status,
            'by_priority': by_priority,
            'by_department': by_department,
            'monthly': monthly,
            'pending_list': pending_list,
            'top_requesters': top_requesters,
        })


# ===========================================================================
# VENDOR REPORT
# ===========================================================================
class VendorSummaryView(APIView):
    permission_classes = [IsAdminOrSuperadmin]

    def get(self, request):
        today = timezone.now().date()
        vendors = Vendor.objects.all()
        contracts = VendorContract.objects.select_related('vendor').all()
        payments = VendorPayment.objects.all()

        total_vendors = vendors.count()
        active_vendors = vendors.filter(is_active=True).count()
        total_contracts = contracts.count()
        active_contracts = contracts.filter(status='ACTIVE').count()
        contract_value = contracts.aggregate(s=Sum('value'))['s'] or 0
        total_payments = payments.aggregate(s=Sum('amount'))['s'] or 0

        soon = today + timedelta(days=90)
        expiring_qs = contracts.filter(
            end_date__gte=today, end_date__lte=soon
        ).exclude(status__in=['TERMINATED', 'EXPIRED']).order_by('end_date')
        expiring = [{
            'vendor': c.vendor.name,
            'contract': c.title,
            'type': c.contract_type,
            'end_date': c.end_date.strftime('%Y-%m-%d'),
            'days_remaining': (c.end_date - today).days,
            'value': float(c.value or 0),
        } for c in expiring_qs[:20]]

        by_category = [{'name': c['category'], 'value': c['count']}
                       for c in vendors.values('category').annotate(count=Count('id')).order_by('-count')]
        contracts_by_status = [{'name': s['status'], 'value': s['count']}
                               for s in contracts.values('status').annotate(count=Count('id')).order_by('-count')]
        contracts_by_type = [{'name': t['contract_type'], 'value': t['count']}
                             for t in contracts.values('contract_type').annotate(count=Count('id')).order_by('-count')]

        spend = payments.values('vendor__name').annotate(
            total=Sum('amount'), count=Count('id')
        ).order_by('-total')[:8]
        top_vendors = [{
            'name': s['vendor__name'], 'value': float(s['total'] or 0), 'count': s['count'],
        } for s in spend]

        monthly = []
        for label, start, end in _last_n_months(12):
            amt = payments.filter(payment_date__gte=start, payment_date__lt=end).aggregate(s=Sum('amount'))['s'] or 0
            monthly.append({'month': label, 'value': float(amt)})

        return Response({
            'totals': {
                'total_vendors': total_vendors, 'active_vendors': active_vendors,
                'total_contracts': total_contracts, 'active_contracts': active_contracts,
                'contract_value': float(contract_value), 'total_payments': float(total_payments),
                'expiring_contracts': expiring_qs.count(),
            },
            'by_category': by_category,
            'contracts_by_status': contracts_by_status,
            'contracts_by_type': contracts_by_type,
            'expiring': expiring,
            'top_vendors': top_vendors,
            'monthly_payments': monthly,
        })


# ===========================================================================
# SEATING REPORT
# ===========================================================================
class SeatingSummaryView(APIView):
    permission_classes = [IsAdminOrSuperadmin]

    def get(self, request):
        seats = Seat.objects.all()
        total_seats = seats.count()
        occupied = seats.filter(assignments__is_active=True).distinct().count()
        vacant = total_seats - occupied
        occupancy_pct = round((occupied / total_seats * 100), 1) if total_seats else 0

        by_office = []
        for office in Office.objects.all():
            o_seats = seats.filter(floor__office=office)
            o_total = o_seats.count()
            o_occ = o_seats.filter(assignments__is_active=True).distinct().count()
            by_office.append({
                'office': office.name,
                'total': o_total,
                'occupied': o_occ,
                'vacant': o_total - o_occ,
                'occupancy_pct': round((o_occ / o_total * 100), 1) if o_total else 0,
            })

        by_type = []
        for t in seats.values('seat_type').annotate(count=Count('id')).order_by('-count'):
            t_occ = seats.filter(seat_type=t['seat_type'], assignments__is_active=True).distinct().count()
            by_type.append({'name': t['seat_type'], 'total': t['count'], 'occupied': t_occ})

        return Response({
            'totals': {
                'total_seats': total_seats, 'occupied': occupied, 'vacant': vacant,
                'occupancy_pct': occupancy_pct,
                'offices': Office.objects.count(), 'floors': Floor.objects.count(),
            },
            'by_office': by_office,
            'by_type': by_type,
            'occupancy_chart': [
                {'name': 'Occupied', 'value': occupied},
                {'name': 'Vacant', 'value': vacant},
            ],
        })


# ===========================================================================
# NETWORK REPORT
# ===========================================================================
class NetworkSummaryView(APIView):
    permission_classes = [IsAdminOrSuperadmin]

    def get(self, request):
        today = timezone.now().date()
        devices = NetworkDevice.objects.all()
        total = devices.count()
        online = devices.filter(status='ONLINE').count()
        offline = devices.filter(status='OFFLINE').count()
        maintenance = devices.filter(status='MAINTENANCE').count()
        decommissioned = devices.filter(status='DECOMMISSIONED').count()

        soon = today + timedelta(days=60)
        warranty_qs = devices.filter(warranty_expiry__gte=today, warranty_expiry__lte=soon).order_by('warranty_expiry')
        warranty_expiring = [{
            'device': d.device_name,
            'code': d.device_code,
            'type': d.device_type,
            'expiry_date': d.warranty_expiry.strftime('%Y-%m-%d'),
            'days_remaining': (d.warranty_expiry - today).days,
        } for d in warranty_qs[:20]]

        by_type = [{'name': t['device_type'], 'value': t['count']}
                   for t in devices.values('device_type').annotate(count=Count('id')).order_by('-count')]
        by_status = [{'name': s['status'], 'value': s['count']}
                     for s in devices.values('status').annotate(count=Count('id')).order_by('-count')]
        by_location = [{'name': l['location__name'] or 'Unassigned', 'value': l['count']}
                       for l in devices.values('location__name').annotate(count=Count('id')).order_by('-count')]

        # IP pool utilization
        assigned_ips = set(
            devices.exclude(ip_address__isnull=True).exclude(ip_address='').values_list('ip_address', flat=True)
        )
        ip_pools = []
        total_capacity = 0
        total_used = 0
        for pool in IPAddressPool.objects.all():
            hosts = pool.get_ip_range()
            capacity = len(hosts)
            used = sum(1 for ip in hosts if ip in assigned_ips)
            total_capacity += capacity
            total_used += used
            ip_pools.append({
                'name': pool.name,
                'network': f"{pool.network_address}/{pool.cidr_prefix}",
                'capacity': capacity,
                'used': used,
                'usage_pct': round((used / capacity * 100), 1) if capacity else 0,
            })

        return Response({
            'totals': {
                'total': total, 'online': online, 'offline': offline,
                'maintenance': maintenance, 'decommissioned': decommissioned,
                'ip_pools': IPAddressPool.objects.count(),
                'ip_capacity': total_capacity, 'ip_used': total_used,
                'warranty_expiring': warranty_qs.count(),
            },
            'by_type': by_type,
            'by_status': by_status,
            'by_location': by_location,
            'warranty_expiring': warranty_expiring,
            'ip_pools': ip_pools,
        })


# ===========================================================================
# ONBOARDING REPORT
# ===========================================================================
class OnboardingSummaryView(APIView):
    permission_classes = [IsAdminOrSuperadmin]

    def get(self, request):
        today = timezone.now().date()
        records = OnboardingRecord.objects.select_related('employee').all()
        total = records.count()
        in_progress = records.filter(status='IN_PROGRESS').count()
        completed = records.filter(status='COMPLETED').count()
        not_started = records.filter(status='NOT_STARTED').count()
        overdue = records.filter(
            target_completion_date__lt=today
        ).exclude(status__in=['COMPLETED', 'CANCELLED']).count()

        by_status = [{'name': s['status'], 'value': s['count']}
                     for s in records.values('status').annotate(count=Count('id')).order_by('-count')]
        by_process = [{'name': p['process_type'], 'value': p['count']}
                      for p in records.values('process_type').annotate(count=Count('id')).order_by('-count')]

        tasks = OnboardingTask.objects.all()
        tasks_total = tasks.count()
        tasks_done = tasks.filter(status__in=['DONE', 'SKIPPED']).count()
        tasks_overdue = tasks.filter(due_date__lt=today).exclude(status__in=['DONE', 'SKIPPED']).count()
        completion_rate = round((tasks_done / tasks_total * 100), 1) if tasks_total else 0

        tasks_by_category = [{'name': c['category'], 'value': c['count']}
                             for c in tasks.values('category').annotate(count=Count('id')).order_by('-count')]

        # Average completion days for finished records
        avg_completion_days = 0
        done_qs = records.filter(actual_completion_date__isnull=False, start_date__isnull=False)
        deltas = [(r.actual_completion_date - r.start_date).days for r in done_qs]
        if deltas:
            avg_completion_days = round(sum(deltas) / len(deltas), 1)

        active_records = []
        for r in records.exclude(status__in=['COMPLETED', 'CANCELLED']).order_by('-created_at')[:15]:
            stats = r.progress_stats
            active_records.append({
                'employee': r.employee.full_name if r.employee else '',
                'process': r.process_type,
                'status': r.status,
                'progress': stats['percentage'],
                'target_date': r.target_completion_date.strftime('%Y-%m-%d') if r.target_completion_date else '',
                'overdue': bool(r.target_completion_date and r.target_completion_date < today),
            })

        return Response({
            'totals': {
                'total': total, 'in_progress': in_progress, 'completed': completed,
                'not_started': not_started, 'overdue': overdue,
                'completion_rate': completion_rate, 'avg_completion_days': avg_completion_days,
                'tasks_total': tasks_total, 'tasks_overdue': tasks_overdue,
            },
            'by_status': by_status,
            'by_process': by_process,
            'tasks_by_category': tasks_by_category,
            'active_records': active_records,
        })


# ===========================================================================
# KNOWLEDGE BASE REPORT
# ===========================================================================
class KBSummaryView(APIView):
    permission_classes = [IsAdminOrSuperadmin]

    def get(self, request):
        articles = KBArticle.objects.all()
        total = articles.count()
        published = articles.filter(status='PUBLISHED').count()
        draft = articles.filter(status='DRAFT').count()
        archived = articles.filter(status='ARCHIVED').count()
        total_views = articles.aggregate(s=Sum('view_count'))['s'] or 0

        by_status = [{'name': s['status'], 'value': s['count']}
                     for s in articles.values('status').annotate(count=Count('id')).order_by('-count')]
        by_category = [{'name': c['category__name'] or 'Uncategorized', 'value': c['count']}
                       for c in articles.values('category__name').annotate(count=Count('id')).order_by('-count')]

        top_viewed = [{
            'title': a.title, 'views': a.view_count,
            'category': a.category.name if a.category else '—',
            'status': a.status,
        } for a in articles.order_by('-view_count')[:10]]

        recently_updated = [{
            'title': a.title,
            'author': a.author.full_name if a.author else '—',
            'updated': a.updated_at.strftime('%Y-%m-%d'),
            'status': a.status,
        } for a in articles.order_by('-updated_at')[:10]]

        helpful = KBFeedback.objects.filter(is_helpful=True).count()
        not_helpful = KBFeedback.objects.filter(is_helpful=False).count()

        return Response({
            'totals': {
                'total': total, 'published': published, 'draft': draft,
                'archived': archived, 'total_views': total_views,
                'helpful': helpful, 'not_helpful': not_helpful,
            },
            'by_status': by_status,
            'by_category': by_category,
            'top_viewed': top_viewed,
            'recently_updated': recently_updated,
            'feedback': [
                {'name': 'Helpful', 'value': helpful},
                {'name': 'Not Helpful', 'value': not_helpful},
            ],
        })


# ===========================================================================
# USERS / PEOPLE REPORT
# ===========================================================================
class UserSummaryView(APIView):
    permission_classes = [IsAdminOrSuperadmin]

    def get(self, request):
        today = timezone.now().date()
        users = User.objects.all()
        total = users.count()
        active = users.filter(is_active=True).count()
        inactive = total - active

        by_role = [{'name': r['role'], 'value': r['count']}
                   for r in users.values('role').annotate(count=Count('id')).order_by('-count')]
        by_department = [{'name': d['department__name'] or 'Unassigned', 'value': d['count']}
                         for d in users.values('department__name').annotate(count=Count('id')).order_by('-count')]

        thirty = today - timedelta(days=30)
        recently_joined = [{
            'name': u.full_name or u.email,
            'role': u.role,
            'department': u.department.name if u.department else '—',
            'joined': u.created_at.strftime('%Y-%m-%d'),
        } for u in users.filter(created_at__date__gte=thirty).order_by('-created_at')[:15]]

        return Response({
            'totals': {
                'total': total, 'active': active, 'inactive': inactive,
                'departments': Department.objects.count(),
            },
            'by_role': by_role,
            'by_department': by_department,
            'recently_joined': recently_joined,
        })


# ===========================================================================
# EXPORTS
# ===========================================================================
class ExportHelpdeskView(APIView):
    permission_classes = [IsAdminOrSuperadmin]

    def get(self, request):
        headers = ['Ticket #', 'Title', 'Category', 'Priority', 'Status', 'Requester',
                   'Assigned To', 'Created', 'Due', 'Resolved', 'Overdue']
        rows = []
        for t in Ticket.objects.select_related('category', 'requester', 'assigned_to').order_by('-created_at'):
            rows.append([
                t.ticket_number, t.title,
                t.category.name if t.category else '',
                t.priority, t.status,
                t.requester.full_name if t.requester else '',
                t.assigned_to.full_name if t.assigned_to else '',
                t.created_at.strftime('%Y-%m-%d') if t.created_at else '',
                t.due_date.strftime('%Y-%m-%d') if t.due_date else '',
                t.resolved_at.strftime('%Y-%m-%d') if t.resolved_at else '',
                'Yes' if t.is_overdue else 'No',
            ])
        return make_xlsx('helpdesk_export.xlsx', [('Tickets', headers, rows)])


class ExportLicensesView(APIView):
    permission_classes = [IsAdminOrSuperadmin]

    def get(self, request):
        headers = ['Product', 'Category', 'Type', 'Seats Total', 'Seats Used', 'Cost',
                   'Billing Cycle', 'Annual Cost', 'Purchase Date', 'Expiry Date', 'Active']
        rows = []
        for l in SoftwareLicense.objects.select_related('product').order_by('product__name'):
            rows.append([
                l.product.name, l.product.category, l.license_type,
                l.seats_total if l.seats_total is not None else 'Unlimited',
                l.seats_used, float(l.cost), l.billing_cycle or '',
                round(l.annual_cost, 2),
                l.purchase_date.strftime('%Y-%m-%d') if l.purchase_date else '',
                l.expiry_date.strftime('%Y-%m-%d') if l.expiry_date else '',
                'Yes' if l.is_active else 'No',
            ])
        return make_xlsx('licenses_export.xlsx', [('Licenses', headers, rows)])


class ExportProcurementView(APIView):
    permission_classes = [IsAdminOrSuperadmin]

    def get(self, request):
        headers = ['PR #', 'Title', 'Requested By', 'Department', 'Priority', 'Status',
                   'Estimated Cost', 'Actual Cost', 'Required By', 'Approved At', 'Created']
        rows = []
        for pr in PurchaseRequest.objects.select_related('requested_by', 'department').order_by('-created_at'):
            rows.append([
                pr.pr_number, pr.title,
                pr.requested_by.full_name if pr.requested_by else '',
                pr.department.name if pr.department else '',
                pr.priority, pr.status,
                float(pr.total_estimated_cost),
                float(pr.total_actual_cost) if pr.total_actual_cost is not None else '',
                pr.required_by_date.strftime('%Y-%m-%d') if pr.required_by_date else '',
                pr.approved_at.strftime('%Y-%m-%d') if pr.approved_at else '',
                pr.created_at.strftime('%Y-%m-%d') if pr.created_at else '',
            ])
        return make_xlsx('procurement_export.xlsx', [('Purchase Requests', headers, rows)])


class ExportVendorsView(APIView):
    permission_classes = [IsAdminOrSuperadmin]

    def get(self, request):
        v_headers = ['Code', 'Name', 'Category', 'Email', 'Phone', 'Country', 'Rating', 'Active']
        v_rows = [[
            v.vendor_code, v.name, v.category, v.email or '', v.phone or '',
            v.country or '', v.rating or '', 'Yes' if v.is_active else 'No',
        ] for v in Vendor.objects.order_by('name')]

        c_headers = ['Contract #', 'Vendor', 'Title', 'Type', 'Status', 'Start', 'End', 'Value', 'Currency']
        c_rows = [[
            c.contract_number, c.vendor.name, c.title, c.contract_type, c.status,
            c.start_date.strftime('%Y-%m-%d') if c.start_date else '',
            c.end_date.strftime('%Y-%m-%d') if c.end_date else '',
            float(c.value) if c.value is not None else '', c.currency,
        ] for c in VendorContract.objects.select_related('vendor').order_by('vendor__name')]

        p_headers = ['Vendor', 'Amount', 'Date', 'Method', 'Invoice #', 'Reference #']
        p_rows = [[
            p.vendor.name, float(p.amount), p.payment_date.strftime('%Y-%m-%d'),
            p.payment_method, p.invoice_number or '', p.reference_number or '',
        ] for p in VendorPayment.objects.select_related('vendor').order_by('-payment_date')]

        return make_xlsx('vendors_export.xlsx', [
            ('Vendors', v_headers, v_rows),
            ('Contracts', c_headers, c_rows),
            ('Payments', p_headers, p_rows),
        ])


class ExportNetworkView(APIView):
    permission_classes = [IsAdminOrSuperadmin]

    def get(self, request):
        headers = ['Code', 'Name', 'Type', 'Status', 'Brand', 'Model', 'IP Address',
                   'MAC', 'Location', 'Warranty Expiry']
        rows = []
        for d in NetworkDevice.objects.select_related('location').order_by('device_code'):
            rows.append([
                d.device_code, d.device_name, d.device_type, d.status,
                d.brand or '', d.model or '', d.ip_address or '', d.mac_address or '',
                d.location.name if d.location else '',
                d.warranty_expiry.strftime('%Y-%m-%d') if d.warranty_expiry else '',
            ])
        return make_xlsx('network_export.xlsx', [('Devices', headers, rows)])


class ExportSeatingView(APIView):
    permission_classes = [IsAdminOrSuperadmin]

    def get(self, request):
        headers = ['Seat Code', 'Type', 'Office', 'Floor', 'Occupied By', 'Status']
        rows = []
        for s in Seat.objects.select_related('floor', 'floor__office').order_by('seat_code'):
            current = s.current_assignment
            rows.append([
                s.seat_code, s.seat_type,
                s.floor.office.name if s.floor and s.floor.office else '',
                s.floor.floor_name if s.floor else '',
                current.user.full_name if current and current.user else '',
                'Occupied' if s.is_occupied else 'Vacant',
            ])
        return make_xlsx('seating_export.xlsx', [('Seats', headers, rows)])


class ExportOnboardingView(APIView):
    permission_classes = [IsAdminOrSuperadmin]

    def get(self, request):
        headers = ['Employee', 'Process', 'Status', 'Progress %', 'Start Date',
                   'Target Date', 'Completed Date']
        rows = []
        for r in OnboardingRecord.objects.select_related('employee').order_by('-created_at'):
            stats = r.progress_stats
            rows.append([
                r.employee.full_name if r.employee else '',
                r.process_type, r.status, stats['percentage'],
                r.start_date.strftime('%Y-%m-%d') if r.start_date else '',
                r.target_completion_date.strftime('%Y-%m-%d') if r.target_completion_date else '',
                r.actual_completion_date.strftime('%Y-%m-%d') if r.actual_completion_date else '',
            ])
        return make_xlsx('onboarding_export.xlsx', [('Records', headers, rows)])


class ExportKBView(APIView):
    permission_classes = [IsAdminOrSuperadmin]

    def get(self, request):
        headers = ['Title', 'Category', 'Status', 'Visibility', 'Author', 'Views', 'Updated']
        rows = []
        for a in KBArticle.objects.select_related('category', 'author').order_by('-updated_at'):
            rows.append([
                a.title,
                a.category.name if a.category else '',
                a.status, a.visibility,
                a.author.full_name if a.author else '',
                a.view_count,
                a.updated_at.strftime('%Y-%m-%d') if a.updated_at else '',
            ])
        return make_xlsx('kb_export.xlsx', [('Articles', headers, rows)])


class ExportUsersView(APIView):
    permission_classes = [IsAdminOrSuperadmin]

    def get(self, request):
        headers = ['Name', 'Email', 'Role', 'Department', 'Designation', 'Active', 'Joined']
        rows = []
        for u in User.objects.select_related('department').order_by('full_name'):
            rows.append([
                u.full_name, u.email, u.role,
                u.department.name if u.department else '',
                u.designation or '', 'Yes' if u.is_active else 'No',
                u.created_at.strftime('%Y-%m-%d') if u.created_at else '',
            ])
        return make_xlsx('users_export.xlsx', [('Users', headers, rows)])
